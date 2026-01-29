# api.py

from io import BytesIO
from contextlib import asynccontextmanager
import json
import time
import threading
from fastapi import FastAPI, Query, UploadFile, File, Form, HTTPException, BackgroundTasks, Request
from fastapi.responses import JSONResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import os
import glob
import shutil
import logging
from typing import List, Optional, Literal
import warnings
from pathlib import Path
import numpy as np
import pandas as pd
from sqlalchemy import text
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import uvicorn
import datetime
from pred_reader import PowerDataImporter
from database import DatabaseManager

# Suppress noisy LibreSSL warning when urllib3 v2 is installed.
warnings.filterwarnings(
    "ignore",
    message="urllib3 v2 only supports OpenSSL 1.1.1+",
)

# 初始化导入器和数据库管理器
importer = PowerDataImporter()
db_manager = DatabaseManager()
logger = logging.getLogger("uvicorn.error")

@asynccontextmanager
async def _lifespan(app: FastAPI):
    _start_cos_daily_scheduler()
    _start_weather_scheduler()
    yield


app = FastAPI(
    title="Excel2SQL API",
    description="API for importing Excel data to SQL database",
    version="1.0.0",
    lifespan=_lifespan,
)

# 挂载静态文件 (use absolute path to avoid CWD issues in scripts)
_STATIC_DIR = Path(__file__).resolve().parent / "static"
app.mount("/static", StaticFiles(directory=str(_STATIC_DIR), check_dir=False), name="static")

# 设置模板
templates = Jinja2Templates(directory="templates")

@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    """返回前端页面"""
    return templates.TemplateResponse("index.html", {"request": request})

# 新增：表查询页面
@app.get("/table_query", response_class=HTMLResponse)
async def table_query_page(request: Request, table_name: str):
    """返回表查询页面"""
    return templates.TemplateResponse("table_query.html", {"request": request, "table_name": table_name})

# 新增：联表查询页面
@app.get("/join_query", response_class=HTMLResponse)
async def join_query_page(request: Request):
    """返回联表查询页面"""
    return templates.TemplateResponse("join_query.html", {"request": request})

@app.get("/health")
async def health_check():
    """健康检查接口"""
    db_status = db_manager.test_connection()
    return {
        "status": "healthy" if db_status else "unhealthy",
        "database": "connected" if db_status else "disconnected"
    }

# COS daily import status (used by the web UI reminder)
_BASE_DIR = Path(__file__).resolve().parent
_COS_DAILY_CONFIG = _BASE_DIR / "cos_daily_import.config.json"
_WEATHER_STATE_DEFAULT = _BASE_DIR / "state" / "weather_update_state.json"


def _cos_scheduler_enabled() -> bool:
    val = os.getenv("COS_DAILY_SCHEDULER", "1")
    return str(val).strip().lower() not in {"0", "false", "off", "no"}


def _try_acquire_lock(lock_path: str):
    try:
        import fcntl  # Unix only
    except Exception:
        return None
    try:
        fd = open(lock_path, "w")
    except Exception:
        return None
    try:
        fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
        fd.write(str(os.getpid()))
        fd.flush()
        return fd
    except Exception:
        try:
            fd.close()
        except Exception:
            pass
        return None


def _start_cos_daily_scheduler():
    if not _cos_scheduler_enabled():
        logger.info("COS daily scheduler disabled via env.")
        return

    lock_path = os.getenv("COS_DAILY_SCHEDULER_LOCK", "/tmp/excel2sql_cos_daily.lock")
    lock_fd = _try_acquire_lock(lock_path)
    if lock_fd is None:
        logger.info("COS daily scheduler lock busy; skip starting scheduler thread.")
        return

    def _loop():
        try:
            from zoneinfo import ZoneInfo
            tz = ZoneInfo("Asia/Shanghai")
        except Exception:
            tz = None

        while True:
            try:
                if not _cos_scheduler_enabled():
                    time.sleep(60)
                    continue

                try:
                    cfg = json.loads(_COS_DAILY_CONFIG.read_text(encoding="utf-8"))
                except Exception as e:
                    logger.error("COS daily scheduler: invalid config: %s", e)
                    time.sleep(300)
                    continue

                poll = cfg.get("polling") or {}
                start_hhmm = poll.get("start_hhmm", "11:20")
                end_hhmm = poll.get("end_hhmm", "12:00")
                interval_seconds = int(poll.get("interval_seconds", 60))

                now = datetime.datetime.now(tz) if tz else datetime.datetime.now()
                base_date = now.date()

                def _parse_hhmm(s: str):
                    parts = str(s).split(":")
                    return int(parts[0]), int(parts[1])

                sh, sm = _parse_hhmm(start_hhmm)
                eh, em = _parse_hhmm(end_hhmm)
                start_dt = datetime.datetime.combine(base_date, datetime.time(sh, sm, tzinfo=tz) if tz else datetime.time(sh, sm))
                end_dt = datetime.datetime.combine(base_date, datetime.time(eh, em, tzinfo=tz) if tz else datetime.time(eh, em))

                if now < start_dt:
                    sleep_s = (start_dt - now).total_seconds()
                    time.sleep(max(1.0, sleep_s))
                    continue
                if now >= end_dt:
                    # sleep to next day's start window
                    next_day = base_date + datetime.timedelta(days=1)
                    next_start = datetime.datetime.combine(next_day, datetime.time(sh, sm, tzinfo=tz) if tz else datetime.time(sh, sm))
                    sleep_s = (next_start - now).total_seconds()
                    time.sleep(max(60.0, sleep_s))
                    continue

                # Within window: run once per tick
                from cos_daily_auto_import import run_once as _cos_run_once, _all_targets_done  # noqa: E402

                _cos_run_once(cfg, base_date=base_date, dry_run=False)

                # Stop early when all targets are done
                try:
                    state_path = (_BASE_DIR / (cfg.get("local", {}).get("state_file") or "./state/cos_daily_state.json")).resolve()
                    state = json.loads(state_path.read_text(encoding="utf-8")) if state_path.exists() else {"days": {}}
                    if _all_targets_done(state, base_date.strftime("%Y-%m-%d"), cfg.get("targets") or {}):
                        # sleep to next day start
                        next_day = base_date + datetime.timedelta(days=1)
                        next_start = datetime.datetime.combine(next_day, datetime.time(sh, sm, tzinfo=tz) if tz else datetime.time(sh, sm))
                        sleep_s = (next_start - now).total_seconds()
                        time.sleep(max(60.0, sleep_s))
                        continue
                except Exception:
                    pass

                time.sleep(max(5, interval_seconds))
            except Exception as e:
                logger.error("COS daily scheduler loop error: %s", e)
                time.sleep(60)

    t = threading.Thread(target=_loop, name="cos_daily_scheduler", daemon=True)
    t.start()


def _weather_scheduler_enabled() -> bool:
    val = os.getenv("WEATHER_AUTO_SCHEDULER", "1")
    return str(val).strip().lower() not in {"0", "false", "off", "no"}


def _weather_update_interval_days() -> int:
    raw = os.getenv("WEATHER_UPDATE_INTERVAL_DAYS", "10")
    try:
        days = int(raw)
        return days if days > 0 else 10
    except Exception:
        return 10


def _resolve_weather_state_path() -> Path:
    raw = os.getenv("WEATHER_SCHEDULER_STATE")
    if not raw:
        return _WEATHER_STATE_DEFAULT
    p = Path(raw)
    return p if p.is_absolute() else (_BASE_DIR / p)


def _load_weather_state(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8")) or {}
    except Exception:
        return {}


def _write_weather_state(path: Path, payload: dict) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(path)


def _parse_iso_dt(value: str):
    if not value:
        return None
    try:
        return datetime.datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except Exception:
        return None


def _run_weather_update():
    import calendar_weather

    today = datetime.date.today()
    start_date = today - datetime.timedelta(days=30)
    end_date = today + datetime.timedelta(days=15)
    logger.info("Weather auto update start: %s -> %s", start_date, end_date)
    calendar_weather.update_calendar(start_date, end_date)
    logger.info("Weather auto update done: %s -> %s", start_date, end_date)
    return start_date, end_date


def _start_weather_scheduler():
    if not _weather_scheduler_enabled():
        logger.info("Weather scheduler disabled via env.")
        return

    lock_path = os.getenv("WEATHER_SCHEDULER_LOCK", "/tmp/excel2sql_weather_update.lock")
    lock_fd = _try_acquire_lock(lock_path)
    if lock_fd is None:
        logger.info("Weather scheduler lock busy; skip starting scheduler thread.")
        return

    state_path = _resolve_weather_state_path()

    def _loop():
        while True:
            try:
                if not _weather_scheduler_enabled():
                    time.sleep(60)
                    continue

                interval_days = _weather_update_interval_days()
                state = _load_weather_state(state_path)
                last_success_at = _parse_iso_dt(state.get("last_success_at"))
                now = datetime.datetime.now()

                due = True
                if last_success_at:
                    due = now >= (last_success_at + datetime.timedelta(days=interval_days))

                if due:
                    try:
                        start_date, end_date = _run_weather_update()
                        now = datetime.datetime.now()
                        state = {
                            "last_success_at": now.isoformat(),
                            "interval_days": interval_days,
                            "last_success_range": {
                                "start": start_date.isoformat(),
                                "end": end_date.isoformat(),
                            },
                        }
                        _write_weather_state(state_path, state)
                        last_success_at = now
                    except Exception as e:
                        logger.error("Weather scheduler update error: %s", e)
                        state = state or {}
                        state["last_error_at"] = datetime.datetime.now().isoformat()
                        state["last_error"] = str(e)
                        _write_weather_state(state_path, state)

                # Sleep until next due, but cap to re-check config periodically.
                if last_success_at:
                    next_due = last_success_at + datetime.timedelta(days=interval_days)
                    sleep_s = (next_due - datetime.datetime.now()).total_seconds()
                else:
                    sleep_s = 3600
                time.sleep(max(60.0, min(sleep_s, 6 * 3600)))
            except Exception as e:
                logger.error("Weather scheduler loop error: %s", e)
                time.sleep(60)

    t = threading.Thread(target=_loop, name="weather_scheduler", daemon=True)
    t.start()


def _load_cos_daily_state():
    if not _COS_DAILY_CONFIG.exists():
        return {
            "status": "missing_config",
            "message": f"config not found: {_COS_DAILY_CONFIG}",
        }
    try:
        with _COS_DAILY_CONFIG.open("r", encoding="utf-8") as f:
            cfg = json.load(f)
    except Exception as e:
        return {"status": "invalid_config", "message": str(e)}

    state_rel = (cfg.get("local") or {}).get("state_file") or "./state/cos_daily_state.json"
    state_path = (_BASE_DIR / state_rel).resolve()
    if not state_path.exists():
        return {
            "status": "no_state",
            "state_file": str(state_path),
            "days": {},
        }

    try:
        with state_path.open("r", encoding="utf-8") as f:
            state = json.load(f)
    except Exception as e:
        return {
            "status": "invalid_state",
            "state_file": str(state_path),
            "message": str(e),
        }

    days = state.get("days", {}) or {}
    if not days:
        return {
            "status": "empty",
            "state_file": str(state_path),
            "days": {},
        }

    latest_day = sorted(days.keys())[-1]
    latest_state = days.get(latest_day, {}) or {}
    targets = latest_state.get("targets", {}) or {}

    def _parse_iso(s):
        if not s:
            return None
        try:
            return datetime.datetime.fromisoformat(str(s).replace("Z", "+00:00"))
        except Exception:
            return None

    last_success_at = None
    last_success_targets = []
    for day_key, day_state in days.items():
        for target_name, t in (day_state.get("targets") or {}).items():
            if str(t.get("status")) != "done":
                continue
            t_time = _parse_iso(t.get("attempted_at"))
            if not t_time:
                continue
            if last_success_at is None or t_time > last_success_at:
                last_success_at = t_time
                last_success_targets = [target_name]
            elif last_success_at and t_time == last_success_at:
                last_success_targets.append(target_name)

    return {
        "status": "ok",
        "server_time": datetime.datetime.now().isoformat(),
        "state_file": str(state_path),
        "day": latest_day,
        "total_attempts": latest_state.get("total_attempts", 0),
        "targets": targets,
        "last_success_at": last_success_at.isoformat() if last_success_at else None,
        "last_success_targets": sorted(set(last_success_targets)),
    }


@app.get("/api/cos_daily/status")
async def cos_daily_status():
    return _load_cos_daily_state()

@app.get("/files")
async def list_files():
    """列出data目录中的所有Excel文件"""
    data_folder = "data"
    os.makedirs(data_folder, exist_ok=True)
    excel_files = glob.glob(os.path.join(data_folder, "*.xlsx"))
    excel_files.sort(reverse=True)  # 按文件名倒序排列（最新日期在前）
    
    return {
        "total": len(excel_files),
        "files": [os.path.basename(file) for file in excel_files]
    }

@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """上传Excel文件到data目录"""
    data_folder = "data"
    os.makedirs(data_folder, exist_ok=True)
    
    # 检查文件类型
    if not file.filename.endswith(('.xlsx')):
        raise HTTPException(status_code=400, detail="只支持.xlsx格式的Excel文件")
    
    # 保存文件
    file_path = os.path.join(data_folder, file.filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    return {"filename": file.filename, "status": "uploaded"}

import re
from pydantic import BaseModel, Field
from datetime import date as Date, timedelta

class SimilarDayRequest(BaseModel):
    target_date: str
    date_type: Optional[str] = None
    weights: Optional[dict] = None

@app.post("/api/similar-day")
async def find_similar_days(request: SimilarDayRequest):
    """
    查找相似日
    匹配维度：负荷预测、天气、温度、B类占比、新能源D日预测、日期类型
    """
    try:
        target_date_str = request.target_date
        weights = request.weights or {}
        
        # 默认权重
        w_load = float(weights.get('load', 0.4))
        w_weather = float(weights.get('weather', 0.1))
        w_temp = float(weights.get('temp', 0.1))
        w_b_ratio = float(weights.get('b_ratio', 0.15))
        w_ne = float(weights.get('ne', 0.1))
        # w_price = float(weights.get('price', 0.1)) # 移除日前电价权重
        w_date = float(weights.get('date', 0.05)) # 日期衰减系数
        
        # 新增权重
        w_month = float(weights.get('month', 0.15)) # 默认考虑月份相似性（二进制：同月=0，不同月=1）
        w_weekday = float(weights.get('weekday', 0.15)) # 默认考虑星期几相似性（二进制：同星期几=0，不同=1）
        w_day_type = float(weights.get('day_type', 0.2)) # 新增：日期类型权重（工作日/周末/节假日）

        # 1. 获取所有缓存数据
        table_name = "cache_daily_hourly"
        with db_manager.engine.connect() as conn:
            # 检查表是否存在
            tables = db_manager.get_tables()
            if table_name not in tables:
                return {"error": "缓存表不存在，请先生成缓存"}

            # 获取全量数据
            # 我们需要以下字段: 
            # record_date, hour, load_forecast, weather, temperature, 
            # class_b_forecast, spot_ne_d_forecast, day_type, price_diff
            
            # 构建查询字段
            fields = [
                "record_date", "hour", 
                "load_forecast", "weather", "temperature",
                "class_b_forecast", "spot_ne_d_forecast", "day_type", "price_diff" # 新增 price_diff
            ]
            
            # 检查字段是否存在 (防止报错)
            # 简单起见，使用 SELECT *，然后在 Pandas 里处理
            df = pd.read_sql(f"SELECT * FROM {table_name}", conn)
            
            if df.empty:
                return {"error": "缓存表中无数据"}

            # 转换日期格式
            df['record_date'] = pd.to_datetime(df['record_date']).dt.strftime('%Y-%m-%d')
        
        # 2. 提取目标日数据
        target_df = df[df['record_date'] == target_date_str].sort_values('hour')
        
        if target_df.empty:
            return {"error": f"目标日期 {target_date_str} 无数据，请先导入预测数据"}

        # 获取目标日期类型
        target_day_type = target_df['day_type'].iloc[0] if 'day_type' in target_df.columns else ''

        # 3. 数据预处理
        # 需要将长表(long)转为宽表(wide)，或者直接按日期分组计算
        
        # 辅助函数：计算两个向量的距离 (MAPE 或 归一化欧氏距离)
        # 这里使用 MAPE (平均绝对百分比误差) 的变体作为差异度量
        
        # 准备历史数据 (必须是目标日之前的日期)
        history_df = df[df['record_date'] < target_date_str].copy()
        
        # 调试：显示目标日期类型（不再强制过滤）
        print(f"[DEBUG] 目标日类型: {target_day_type or '无类型'}")
        # 不再强制过滤日期类型，允许匹配所有历史数据
        # 用户可通过设置月份/星期几权重为0来禁用相关过滤
        
        # 必须有24小时数据的日期才参与计算
        print(f"[DEBUG] 历史数据天数（24小时过滤前）: {len(history_df['record_date'].unique())}")
        valid_dates = history_df.groupby('record_date').count()['hour']
        valid_dates = valid_dates[valid_dates == 24].index.tolist()
        history_df = history_df[history_df['record_date'].isin(valid_dates)]
        print(f"[DEBUG] 历史数据天数（24小时过滤后）: {len(history_df['record_date'].unique())}")
        
        # 新增：排除没有价差数据的日期
        # 检查每个日期是否有完整的价差数据（price_diff 不为 None 且不为 NaN）
        if 'price_diff' in history_df.columns:
            # 按日期分组，检查是否有完整的24小时价差数据
            price_diff_counts = history_df.groupby('record_date')['price_diff'].apply(
                lambda x: x.notna().sum()  # 统计非空值数量
            )
            # 只保留有24个非空价差数据的日期
            dates_with_price_diff = price_diff_counts[price_diff_counts == 24].index.tolist()
            history_df = history_df[history_df['record_date'].isin(dates_with_price_diff)]
            print(f"[DEBUG] 历史数据天数（价差数据过滤后）: {len(history_df['record_date'].unique())}")
            print(f"[DEBUG] 有完整价差数据的日期: {dates_with_price_diff}")
        else:
            print(f"[DEBUG] 警告：缓存表中没有 price_diff 字段，无法进行价差数据过滤")
        
        if history_df.empty:
            return {"error": "没有足够的历史数据进行匹配"}

        # ---------------------------
        # 计算各项差异
        # ---------------------------
        
        results = []
        print(f"[DEBUG] 权重配置 - load:{w_load}, temp:{w_temp}, weather:{w_weather}, "
              f"b_ratio:{w_b_ratio}, ne:{w_ne}, "
              f"date:{w_date}, month:{w_month}, weekday:{w_weekday}, day_type:{w_day_type}")
        target_date_obj = datetime.datetime.strptime(target_date_str, "%Y-%m-%d").date()
        print(f"[DEBUG] 目标日期: {target_date_str}, 月份: {target_date_obj.month}, 星期几: {target_date_obj.weekday()}(0=周一)")

        # 预计算目标向量
        t_load = target_df['load_forecast'].fillna(0).values
        t_temp = target_df['temperature'].fillna(0).values
        # t_price = target_df['price_da'].fillna(0).values # 移除
        
        # B类占比
        t_b = target_df['class_b_forecast'].fillna(0).values
        # 避免除以0
        t_load_safe = np.where(t_load == 0, 1, t_load)
        t_b_ratio = t_b / t_load_safe
        
        # 新能源D日
        # 优先使用 spot_ne_d_forecast，如果没有则尝试用 new_energy_forecast
        if 'spot_ne_d_forecast' in target_df.columns and target_df['spot_ne_d_forecast'].sum() > 0:
            t_ne = target_df['spot_ne_d_forecast'].fillna(0).values
        elif 'new_energy_forecast' in target_df.columns:
            t_ne = target_df['new_energy_forecast'].fillna(0).values
        else:
            t_ne = np.zeros(24)

        # 天气 (字符串数组)
        t_weather = target_df['weather'].fillna("").values
        
        # 目标日期类型 (字符串)
        t_day_type = target_df['day_type'].iloc[0] if 'day_type' in target_df.columns else ""
        
        # 计算目标日期的统计信息
        target_weather_type = ""
        if len(t_weather) > 12:
            target_weather_type = t_weather[12]  # 取中午时段的天气作为代表
        elif len(t_weather) > 0:
            target_weather_type = t_weather[0]   # 如果没有12点数据，取第一个
        
        target_avg_temp = float(np.mean(t_temp)) if len(t_temp) > 0 else 0.0
        target_avg_load = float(np.mean(t_load)) if len(t_load) > 0 else 0.0
        # target_avg_price = float(np.mean(t_price)) if len(t_price) > 0 else 0.0 # 移除
        target_avg_b_ratio = float(np.mean(t_b_ratio)) if len(t_b_ratio) > 0 else 0.0
        target_avg_ne = float(np.mean(t_ne)) if len(t_ne) > 0 else 0.0

        # 遍历历史日期
        # 为了加速，可以使用 groupby Apply，但循环简单直观
        for date_val, group in history_df.groupby('record_date'):
            group = group.sort_values('hour')
            
            # 1. 负荷差异 (MAPE)
            h_load = group['load_forecast'].fillna(0).values
            # 如果负荷为空，跳过
            if np.sum(h_load) == 0:
                diff_load = 1.0 # 最大差异
            else:
                # MAPE: mean(abs(t - h) / t) -> 但 t 可能为0，且我们要的是相似度
                # 使用 归一化欧氏距离: dist / (norm(t) + norm(h)) 或 simple MAPE
                # 简单处理：mean(abs(diff)) / mean(target)
                mean_target = np.mean(t_load) if np.mean(t_load) > 0 else 1
                diff_load = np.mean(np.abs(t_load - h_load)) / mean_target
            
            # 2. 温度差异 (RMSE + 最高最低对比)
            h_temp = group['temperature'].fillna(0).values
            diff_temp = np.sqrt(np.mean((t_temp - h_temp)**2))
            # 最高温度差异
            max_diff = np.max(t_temp) - np.max(h_temp)
            diff_temp_max = abs(max_diff)
            # 最低温度差异
            min_diff = np.min(t_temp) - np.min(h_temp)
            diff_temp_min = abs(min_diff)
            # 综合温度差异归一化 (假设温差10度算大)
            diff_temp_norm = min((diff_temp / 10.0 + diff_temp_max / 10.0 + diff_temp_min / 10.0) / 3.0, 1.0)
            
            # 3. B类占比差异
            h_b = group['class_b_forecast'].fillna(0).values
            h_load_safe = np.where(h_load == 0, 1, h_load)
            h_b_ratio = h_b / h_load_safe
            diff_b_ratio = np.mean(np.abs(t_b_ratio - h_b_ratio)) # 本身就是比例，直接差值
            
            # 4. 新能源差异
            # 同样处理列名
            if 'spot_ne_d_forecast' in group.columns and group['spot_ne_d_forecast'].sum() > 0:
                h_ne = group['spot_ne_d_forecast'].fillna(0).values
            elif 'new_energy_forecast' in group.columns:
                h_ne = group['new_energy_forecast'].fillna(0).values
            else:
                h_ne = np.zeros(24)
            
            mean_ne_target = np.mean(t_ne) if np.mean(t_ne) > 0 else 1
            diff_ne = np.mean(np.abs(t_ne - h_ne)) / mean_ne_target
            
            # 5. 价格差异 (移除)
            # h_price = group['price_da'].fillna(0).values
            # mean_price_target = np.mean(t_price) if np.mean(t_price) > 0 else 1
            # diff_price = np.mean(np.abs(t_price - h_price)) / mean_price_target
            
            # 6. 天气差异 (不匹配的小时数比例)
            h_weather = group['weather'].fillna("").values
            # 简单比较字符串是否相等
            diff_weather = np.mean(t_weather != h_weather)
            
            # 7. 日期权重 (越近越好)
            # 计算天数差
            hist_date_obj = datetime.datetime.strptime(date_val, "%Y-%m-%d").date()
            days_diff = abs((target_date_obj - hist_date_obj).days)
            # 衰减因子: 1 - exp(-k * days) -> 距离
            # 或者 距离增加: days_diff / 365
            date_penalty = min(days_diff / 365.0, 1.0)
            
            # 8. 月份差异 (二进制: 同月=0, 不同月=1)
            target_month = target_date_obj.month
            hist_month = hist_date_obj.month
            diff_month = 0.0 if target_month == hist_month else 1.0
            
            # 9. 星期几差异 (二进制: 同为星期几=0, 不同=1)
            target_weekday = target_date_obj.weekday()  # Monday=0, Sunday=6
            hist_weekday = hist_date_obj.weekday()
            diff_weekday = 0.0 if target_weekday == hist_weekday else 1.0
            
            # 10. 日期类型差异 (新增)
            # 获取历史日期的 day_type
            h_day_type = group['day_type'].iloc[0] if 'day_type' in group.columns else ""
            # 如果目标或历史缺失日期类型，则视为差异大
            if not t_day_type or not h_day_type:
                diff_day_type = 1.0
            else:
                diff_day_type = 0.0 if t_day_type == h_day_type else 1.0
            
            # 总差异得分 (越小越好)
            # 各项 diff 都在 [0, 1] 左右 (MAPE可能大于1，但通常在0-0.5)
            total_score = (
                w_load * diff_load +
                w_temp * diff_temp_norm +
                w_b_ratio * diff_b_ratio +
                w_ne * diff_ne +
                # w_price * diff_price + # 移除
                w_weather * diff_weather +
                w_date * date_penalty +
                w_month * diff_month +
                w_weekday * diff_weekday +
                w_day_type * diff_day_type # 新增
            )
            
            results.append({
                "date": date_val,
                "score": total_score,
                "details": {
                    "diff_load": float(diff_load),
                    "diff_temp": float(diff_temp),
                    "diff_temp_norm": float(diff_temp_norm),
                    "diff_temp_max": float(diff_temp_max),
                    "diff_temp_min": float(diff_temp_min),
                    "diff_weather": float(diff_weather),
                    "diff_b_ratio": float(diff_b_ratio),
                    "diff_ne": float(diff_ne),
                    "date_penalty": float(date_penalty),
                    "diff_month": float(diff_month),
                    "diff_weekday": float(diff_weekday),
                    "diff_day_type": float(diff_day_type)
                },
                # 返回一些用于展示的数据
                "load_curve": h_load.tolist(),
                "temp_avg": float(np.mean(h_temp)),
                "weather_type": h_weather[12] if len(h_weather) > 12 else "", # 取中午天气作为代表
                "day_type": group['day_type'].iloc[0] if 'day_type' in group.columns else ""
            })
            
        # 排序并返回前5
        results.sort(key=lambda x: x['score'])
        top_matches = results[:5]
        
        # 转换得分为相似度 (1 / (1 + score)) 或者 (1 - score)
        for r in top_matches:
            r['similarity_score'] = max(0, 1 - r['score']) # 简单线性映射
            
        return {
            "target_date": target_date_str,
            "target_day_type": target_day_type,
            "target_weather_type": target_weather_type,
            "target_stats": {
                "avg_temp": target_avg_temp,
                "avg_load": target_avg_load,
                # "avg_price": target_avg_price, # 移除
                "avg_b_ratio": target_avg_b_ratio,
                "avg_ne": target_avg_ne
            },
            "target_load_curve": t_load.tolist(),
            # "target_price_curve": t_price.tolist(), # 移除
            "matches": top_matches
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


# =========================
# 策略报价/复盘（申报电量 + 胜率/收益）
# =========================

class StrategyActualHourlyIn(BaseModel):
    date: str = Field(..., description="日期 YYYY-MM-DD")
    hourly: List[float] = Field(..., description="24小时实际分时电量(00-23)，长度必须为24")
    source: Literal["actual", "settlement", "both"] = Field(
        "both",
        description="写入目标表：actual=strategy_actual_hourly, settlement=strategy_settlement_actual_hourly, both=两者都写",
    )
    platform: Optional[str] = Field(None, description="平台：天朗/辉华（默认天朗）")


class StrategyDaySettingsIn(BaseModel):
    date: str = Field(..., description="目标日期 YYYY-MM-DD")
    # 策略系数为 24 个时刻分别报；保留 strategy_coeff 作为兼容/快速填充（可选）。
    strategy_coeff: Optional[float] = Field(None, description="策略系数(可选：统一系数/兼容旧数据)")
    strategy_coeff_hourly: Optional[List[float]] = Field(
        None, description="24个时刻策略系数(00-23)，长度必须为24"
    )
    revenue_transfer: float = Field(0.0, description="收益转移（元，可正可负）")
    note: Optional[str] = Field(None, description="备注/策略说明")
    platform: Optional[str] = Field(None, description="平台：天朗/辉华（默认天朗）")


_STRATEGY_PLATFORM_LABELS = {"tianlang": "天朗", "huihua": "辉华"}
_STRATEGY_PLATFORM_ALIASES = {
    None: "tianlang",
    "": "tianlang",
    "tianlang": "tianlang",
    "tl": "tianlang",
    "天朗": "tianlang",
    "huihua": "huihua",
    "hh": "huihua",
    "辉华": "huihua",
}


def _normalize_platform(value: Optional[str]) -> str:
    """Normalize platform name/code into a safe internal code."""
    try:
        v = (value or "").strip()
    except Exception:
        v = ""
    key = v if v in _STRATEGY_PLATFORM_ALIASES else v.lower()
    p = _STRATEGY_PLATFORM_ALIASES.get(key)
    if not p:
        # Default to existing dataset to keep backward compatibility.
        p = "tianlang"
    if p not in _STRATEGY_PLATFORM_LABELS:
        p = "tianlang"
    return p


def _strategy_table(base: str, platform: Optional[str]) -> str:
    """
    Strategy tables are isolated per platform.
    - 天朗: use existing table names (no suffix) to keep current data intact.
    - 辉华: create/read tables with suffix `_huihua`.
    """
    p = _normalize_platform(platform)
    if p == "tianlang":
        return base
    # base is hard-coded; platform is normalized to a safe ascii identifier.
    return f"{base}_{p}"


def _parse_iso_date(value: str) -> Date:
    try:
        return Date.fromisoformat(value)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"日期格式错误: {value}") from e


def _hour_label_to_hour(label) -> Optional[int]:
    """Accept '00:00', '00:00:00', datetime.time(0,0) etc -> 0..23"""
    if label is None or (isinstance(label, float) and np.isnan(label)):
        return None
    if isinstance(label, datetime.time):
        return int(label.hour)
    s = str(label).strip()
    if not s:
        return None
    # Normalize common formats
    if s.endswith(":00") and len(s) == 5 and s[0:2].isdigit():
        return int(s[0:2])
    if s.endswith(":00:00") and len(s) >= 8 and s[0:2].isdigit():
        return int(s[0:2])
    # Fallback: try split
    parts = s.split(":")
    if parts and parts[0].isdigit():
        h = int(parts[0])
        if 0 <= h <= 23:
            return h
    return None


_STRATEGY_TABLES_READY_PLATFORMS = set()
_STRATEGY_TABLES_LOCK = threading.Lock()


def _ensure_strategy_tables(platform: Optional[str] = None):
    """Create tables on-demand (safe to call per request)."""
    # Table creation is idempotent, but repeatedly running these DDL statements is slow on remote DBs.
    p = _normalize_platform(platform)
    if p in _STRATEGY_TABLES_READY_PLATFORMS:
        return
    with _STRATEGY_TABLES_LOCK:
        if p in _STRATEGY_TABLES_READY_PLATFORMS:
            return
    with db_manager.engine.connect() as conn:
        with conn.begin():
            t_actual = _strategy_table("strategy_actual_hourly", p)
            conn.execute(
                text(
                    f"""
                    CREATE TABLE IF NOT EXISTS {t_actual} (
                        record_date DATE NOT NULL,
                        hour TINYINT NOT NULL,
                        actual_energy DOUBLE NULL,
                        updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
                            ON UPDATE CURRENT_TIMESTAMP,
                        PRIMARY KEY (record_date, hour)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                    """
                )
            )
            t_settle = _strategy_table("strategy_settlement_actual_hourly", p)
            conn.execute(
                text(
                    f"""
                    CREATE TABLE IF NOT EXISTS {t_settle} (
                        record_date DATE NOT NULL,
                        hour TINYINT NOT NULL,
                        actual_energy DOUBLE NULL,
                        updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
                            ON UPDATE CURRENT_TIMESTAMP,
                        PRIMARY KEY (record_date, hour)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                    """
                )
            )
            t_settings = _strategy_table("strategy_day_settings", p)
            conn.execute(
                text(
                    f"""
                    CREATE TABLE IF NOT EXISTS {t_settings} (
                        record_date DATE NOT NULL,
                        strategy_coeff DOUBLE NULL,
                        revenue_transfer DOUBLE NOT NULL DEFAULT 0,
                        note TEXT NULL,
                        updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
                            ON UPDATE CURRENT_TIMESTAMP,
                        PRIMARY KEY (record_date)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                    """
                )
            )
            t_coeff = _strategy_table("strategy_hourly_coeff", p)
            conn.execute(
                text(
                    f"""
                    CREATE TABLE IF NOT EXISTS {t_coeff} (
                        record_date DATE NOT NULL,
                        hour TINYINT NOT NULL,
                        coeff DOUBLE NULL,
                        updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
                            ON UPDATE CURRENT_TIMESTAMP,
                        PRIMARY KEY (record_date, hour)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                    """
                )
            )
            t_price = _strategy_table("strategy_price_hourly", p)
            conn.execute(
                text(
                    f"""
                    CREATE TABLE IF NOT EXISTS {t_price} (
                        record_date DATE NOT NULL,
                        hour TINYINT NOT NULL,
                        price_da DOUBLE NULL,
                        price_rt DOUBLE NULL,
                        price_diff DOUBLE NULL,
                        updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
                            ON UPDATE CURRENT_TIMESTAMP,
                        PRIMARY KEY (record_date, hour)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                    """
                )
            )
            t_forecast = _strategy_table("strategy_forecast_hourly", p)
            conn.execute(
                text(
                    f"""
                    CREATE TABLE IF NOT EXISTS {t_forecast} (
                        record_date DATE NOT NULL,
                        hour TINYINT NOT NULL,
                        forecast_energy DOUBLE NULL,
                        source VARCHAR(32) NULL,
                        updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
                            ON UPDATE CURRENT_TIMESTAMP,
                        PRIMARY KEY (record_date, hour)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                    """
                )
            )
            t_declared = _strategy_table("strategy_declared_hourly", p)
            conn.execute(
                text(
                    f"""
                    CREATE TABLE IF NOT EXISTS {t_declared} (
                        record_date DATE NOT NULL,
                        hour TINYINT NOT NULL,
                        declared_energy DOUBLE NULL,
                        updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
                            ON UPDATE CURRENT_TIMESTAMP,
                        PRIMARY KEY (record_date, hour)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                    """
                )
            )
            t_profit = _strategy_table("strategy_daily_profit", p)
            conn.execute(
                text(
                    f"""
                    CREATE TABLE IF NOT EXISTS {t_profit} (
                        record_date DATE NOT NULL,
                        profit_real DOUBLE NULL,
                        profit_expected DOUBLE NULL,
                        updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
                            ON UPDATE CURRENT_TIMESTAMP,
                        PRIMARY KEY (record_date)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                    """
                )
            )
            t_metrics = _strategy_table("strategy_daily_metrics", p)
            conn.execute(
                text(
                    f"""
                    CREATE TABLE IF NOT EXISTS {t_metrics} (
                        record_date DATE NOT NULL,
                        actual_total DOUBLE NULL,
                        forecast_total DOUBLE NULL,
                        declared_total DOUBLE NULL,
                        forecast_accuracy DOUBLE NULL,
                        declared_accuracy DOUBLE NULL,
                        forecast_bias DOUBLE NULL,
                        declared_bias DOUBLE NULL,
                        strategy_correct INT NULL,
                        strategy_total INT NULL,
                        strategy_correct_active INT NULL,
                        strategy_total_active INT NULL,
                        forecast_correct INT NULL,
                        forecast_total_hours INT NULL,
                        assessment_recovery DOUBLE NULL,
                        profit_real DOUBLE NULL,
                        profit_expected DOUBLE NULL,
                        coeff_total DOUBLE NULL,
                        coeff_avg DOUBLE NULL,
                        coeff_min DOUBLE NULL,
                        coeff_max DOUBLE NULL,
                        updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP
                            ON UPDATE CURRENT_TIMESTAMP,
                        PRIMARY KEY (record_date)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
                    """
                )
            )

    with _STRATEGY_TABLES_LOCK:
        _STRATEGY_TABLES_READY_PLATFORMS.add(p)


def _upsert_actual_hourly(records: List[dict], platform: Optional[str] = None) -> int:
    if not records:
        return 0
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_actual_hourly", p)
    sql = text(
        f"""
        INSERT INTO {table} (record_date, hour, actual_energy)
        VALUES (:record_date, :hour, :actual_energy)
        ON DUPLICATE KEY UPDATE actual_energy=VALUES(actual_energy)
        """
    )
    with db_manager.engine.connect() as conn:
        with conn.begin():
            conn.execute(sql, records)

    # Metrics may depend on this actual date as well as future dates that use it as D-7/14/21.
    try:
        touched_dates = sorted({r["record_date"] for r in records if r.get("record_date")})
        refresh = set()
        for d in touched_dates:
            refresh.add(d)
            refresh.add(d + timedelta(days=7))
            refresh.add(d + timedelta(days=14))
            refresh.add(d + timedelta(days=21))
        _refresh_daily_metrics_for_dates(list(refresh), platform=p)
    except Exception:
        pass
    return len(records)


def _upsert_settlement_actual_hourly(records: List[dict], platform: Optional[str] = None) -> int:
    if not records:
        return 0
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_settlement_actual_hourly", p)
    sql = text(
        f"""
        INSERT INTO {table} (record_date, hour, actual_energy)
        VALUES (:record_date, :hour, :actual_energy)
        ON DUPLICATE KEY UPDATE actual_energy=VALUES(actual_energy)
        """
    )
    with db_manager.engine.connect() as conn:
        with conn.begin():
            conn.execute(sql, records)

    # Metrics are evaluated against settlement actuals when available.
    try:
        touched_dates = sorted({r["record_date"] for r in records if r.get("record_date")})
        _refresh_daily_metrics_for_dates(touched_dates, platform=p)
    except Exception:
        pass
    return len(records)


def _read_settlement_actual_hourly(d: Date, platform: Optional[str] = None) -> dict:
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_settlement_actual_hourly", p)
    with db_manager.engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT hour, actual_energy FROM {table} WHERE record_date=:d"),
            {"d": d},
        ).fetchall()
    out = {}
    for h, v in rows:
        try:
            hh = int(h)
        except Exception:
            continue
        if 0 <= hh <= 23:
            out[hh] = None if v is None else float(v)
    return out


def _upsert_day_settings(
    d: Date,
    strategy_coeff: Optional[float] = None,
    revenue_transfer: float = 0.0,
    note: Optional[str] = None,
    platform: Optional[str] = None,
):
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_day_settings", p)
    sql = text(
        f"""
        INSERT INTO {table} (record_date, strategy_coeff, revenue_transfer, note)
        VALUES (:record_date, :strategy_coeff, :revenue_transfer, :note)
        ON DUPLICATE KEY UPDATE
            strategy_coeff=VALUES(strategy_coeff),
            revenue_transfer=VALUES(revenue_transfer),
            note=VALUES(note)
        """
    )
    with db_manager.engine.connect() as conn:
        with conn.begin():
            conn.execute(
                sql,
                {
                    "record_date": d,
                    "strategy_coeff": None if strategy_coeff is None else float(strategy_coeff),
                    "revenue_transfer": float(revenue_transfer or 0.0),
                    "note": note,
                },
            )

    # Keep cached daily metrics in sync.
    try:
        _refresh_daily_metrics_for_dates([d], platform=p)
    except Exception:
        pass


def _upsert_hourly_coeff(d: Date, coeff_hourly: List[float], platform: Optional[str] = None) -> None:
    if coeff_hourly is None:
        return
    if len(coeff_hourly) != 24:
        raise HTTPException(status_code=400, detail="strategy_coeff_hourly 长度必须为24")
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_hourly_coeff", p)
    rows = []
    for h in range(24):
        v = coeff_hourly[h]
        if v is None or (isinstance(v, float) and np.isnan(v)):
            rows.append({"record_date": d, "hour": h, "coeff": None})
        else:
            rows.append({"record_date": d, "hour": h, "coeff": float(v)})
    sql = text(
        f"""
        INSERT INTO {table} (record_date, hour, coeff)
        VALUES (:record_date, :hour, :coeff)
        ON DUPLICATE KEY UPDATE coeff=VALUES(coeff)
        """
    )
    with db_manager.engine.connect() as conn:
        with conn.begin():
            conn.execute(sql, rows)

    # Keep cached daily metrics in sync.
    try:
        _refresh_daily_metrics_for_dates([d], platform=p)
    except Exception:
        pass


def _read_hourly_coeff(d: Date, platform: Optional[str] = None) -> dict:
    """Return hour->coeff (only existing rows)."""
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_hourly_coeff", p)
    with db_manager.engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT hour, coeff FROM {table} WHERE record_date=:d"),
            {"d": d},
        ).fetchall()
    out = {}
    for h, v in rows:
        try:
            hh = int(h)
        except Exception:
            continue
        if 0 <= hh <= 23:
            out[hh] = None if v is None else float(v)
    return out


def _upsert_price_hourly(records: List[dict], platform: Optional[str] = None) -> int:
    if not records:
        return 0
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_price_hourly", p)
    sql = text(
        f"""
        INSERT INTO {table} (record_date, hour, price_da, price_rt, price_diff)
        VALUES (:record_date, :hour, :price_da, :price_rt, :price_diff)
        ON DUPLICATE KEY UPDATE
            price_da=VALUES(price_da),
            price_rt=VALUES(price_rt),
            price_diff=VALUES(price_diff)
        """
    )
    with db_manager.engine.connect() as conn:
        with conn.begin():
            conn.execute(sql, records)

    # Keep cached daily metrics in sync (win rates & assessment depend on price_diff).
    try:
        touched_dates = sorted({r["record_date"] for r in records if r.get("record_date")})
        _refresh_daily_metrics_for_dates(touched_dates, platform=p)
    except Exception:
        pass
    return len(records)


def _upsert_forecast_hourly(records: List[dict], platform: Optional[str] = None) -> int:
    """
    Upsert forecast hourly energy.
    This is used to decouple monthly forecasts from the D-7/14/21 weighted forecast when needed.
    """
    if not records:
        return 0
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_forecast_hourly", p)
    sql = text(
        f"""
        INSERT INTO {table} (record_date, hour, forecast_energy, source)
        VALUES (:record_date, :hour, :forecast_energy, :source)
        ON DUPLICATE KEY UPDATE
            forecast_energy=VALUES(forecast_energy),
            source=COALESCE(VALUES(source), source)
        """
    )
    with db_manager.engine.connect() as conn:
        with conn.begin():
            conn.execute(sql, records)
    # Forecast impacts daily metrics for the target date.
    try:
        touched_dates = sorted({r["record_date"] for r in records if r.get("record_date")})
        _refresh_daily_metrics_for_dates(touched_dates, platform=p)
    except Exception:
        pass
    return len(records)


def _read_forecast_hourly(d: Date, platform: Optional[str] = None) -> dict:
    """Return hour->forecast_energy (only existing rows)."""
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_forecast_hourly", p)
    with db_manager.engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT hour, forecast_energy FROM {table} WHERE record_date=:d"),
            {"d": d},
        ).fetchall()
    out = {}
    for h, v in rows:
        try:
            hh = int(h)
        except Exception:
            continue
        if 0 <= hh <= 23:
            out[hh] = None if v is None else float(v)
    return out


def _upsert_declared_hourly(records: List[dict], platform: Optional[str] = None) -> int:
    """
    Upsert day-ahead declared hourly energy (actual submitted).
    This is used for settlement-based assessment recovery.
    """
    if not records:
        return 0
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_declared_hourly", p)
    sql = text(
        f"""
        INSERT INTO {table} (record_date, hour, declared_energy)
        VALUES (:record_date, :hour, :declared_energy)
        ON DUPLICATE KEY UPDATE declared_energy=VALUES(declared_energy)
        """
    )
    with db_manager.engine.connect() as conn:
        with conn.begin():
            conn.execute(sql, records)

    try:
        touched_dates = sorted({r["record_date"] for r in records if r.get("record_date")})
        _refresh_daily_metrics_for_dates(touched_dates, platform=p)
    except Exception:
        pass
    return len(records)


def _read_declared_hourly(d: Date, platform: Optional[str] = None) -> dict:
    """Return hour->declared_energy (only existing rows)."""
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_declared_hourly", p)
    with db_manager.engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT hour, declared_energy FROM {table} WHERE record_date=:d"),
            {"d": d},
        ).fetchall()
    out = {}
    for h, v in rows:
        try:
            hh = int(h)
        except Exception:
            continue
        if 0 <= hh <= 23:
            out[hh] = None if v is None else float(v)
    return out


def _read_price_diff(d: Date, platform: Optional[str] = None) -> dict:
    """
    Return hour -> price_diff (DA-RT).
    Prefer cache_daily_hourly (system cache), then fallback to strategy_price_hourly (workbook import).
    """
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    out = {}
    with db_manager.engine.connect() as conn:
        rows = conn.execute(
            text("SELECT hour, price_diff, price_da, price_rt FROM cache_daily_hourly WHERE record_date=:d"),
            {"d": d},
        ).fetchall()
    for h, diff, pda, prt in rows:
        try:
            hh = int(h)
        except Exception:
            continue
        if 0 <= hh <= 23:
            if diff is not None:
                out[hh] = float(diff)
            elif pda is not None and prt is not None:
                out[hh] = float(pda) - float(prt)

    # Fallback if cache is empty for this date.
    if out:
        return out

    table = _strategy_table("strategy_price_hourly", p)
    with db_manager.engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT hour, price_diff, price_da, price_rt FROM {table} WHERE record_date=:d"),
            {"d": d},
        ).fetchall()
    for h, diff, pda, prt in rows:
        try:
            hh = int(h)
        except Exception:
            continue
        if 0 <= hh <= 23:
            if diff is not None:
                out[hh] = float(diff)
            elif pda is not None and prt is not None:
                out[hh] = float(pda) - float(prt)
    return out


def _upsert_daily_profit(
    d: Date, profit_real: Optional[float] = None, profit_expected: Optional[float] = None, platform: Optional[str] = None
):
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_daily_profit", p)
    sql = text(
        f"""
        INSERT INTO {table} (record_date, profit_real, profit_expected)
        VALUES (:record_date, :profit_real, :profit_expected)
        ON DUPLICATE KEY UPDATE
            profit_real=COALESCE(VALUES(profit_real), profit_real),
            profit_expected=COALESCE(VALUES(profit_expected), profit_expected)
        """
    )
    with db_manager.engine.connect() as conn:
        with conn.begin():
            conn.execute(
                sql,
                {
                    "record_date": d,
                    "profit_real": None if profit_real is None else float(profit_real),
                    "profit_expected": None if profit_expected is None else float(profit_expected),
                },
            )

    # Keep cached daily metrics in sync.
    try:
        _refresh_daily_metrics_for_dates([d], platform=p)
    except Exception:
        pass


def _read_daily_profit(d: Date, platform: Optional[str] = None) -> dict:
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_daily_profit", p)
    with db_manager.engine.connect() as conn:
        row = conn.execute(
            text(f"SELECT profit_real, profit_expected FROM {table} WHERE record_date=:d"),
            {"d": d},
        ).fetchone()
    if not row:
        return {"profit_real": None, "profit_expected": None}
    return {
        "profit_real": None if row[0] is None else float(row[0]),
        "profit_expected": None if row[1] is None else float(row[1]),
    }


def _read_actual_hourly(d: Date, platform: Optional[str] = None) -> dict:
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_actual_hourly", p)
    with db_manager.engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT hour, actual_energy FROM {table} WHERE record_date=:d"),
            {"d": d},
        ).fetchall()
    out = {}
    for h, v in rows:
        try:
            hh = int(h)
        except Exception:
            continue
        if 0 <= hh <= 23:
            out[hh] = None if v is None else float(v)
    return out


def _read_actual_hourly_multi(dates: List[Date], platform: Optional[str] = None) -> dict:
    """
    Read multiple days' actual hourly in a single query to reduce round-trips.
    Return: { record_date (Date) -> {hour -> actual_energy} }
    """
    if not dates:
        return {}
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_actual_hourly", p)
    params = {f"d{i}": d for i, d in enumerate(dates)}
    in_list = ", ".join([f":d{i}" for i in range(len(dates))])
    q = text(
        f"""
        SELECT record_date, hour, actual_energy
        FROM {table}
        WHERE record_date IN ({in_list})
        """
    )
    out = {d: {} for d in dates}
    with db_manager.engine.connect() as conn:
        rows = conn.execute(q, params).fetchall()
    for dd, h, v in rows:
        try:
            hh = int(h)
        except Exception:
            continue
        if 0 <= hh <= 23:
            try:
                d = dd if isinstance(dd, Date) else Date.fromisoformat(str(dd))
            except Exception:
                continue
            if d not in out:
                out[d] = {}
            out[d][hh] = None if v is None else float(v)
    return out


def _count_actual_hourly(d: Date, platform: Optional[str] = None) -> int:
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_actual_hourly", p)
    with db_manager.engine.connect() as conn:
        return int(
            conn.execute(
                text(f"SELECT COUNT(*) FROM {table} WHERE record_date=:d"),
                {"d": d},
            ).scalar()
            or 0
        )


def _read_day_settings(d: Date, platform: Optional[str] = None) -> dict:
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_day_settings", p)
    with db_manager.engine.connect() as conn:
        row = conn.execute(
            text(f"SELECT strategy_coeff, revenue_transfer, note FROM {table} WHERE record_date=:d"),
            {"d": d},
        ).fetchone()
    if not row:
        return {"strategy_coeff": None, "revenue_transfer": 0.0, "note": None}
    return {
        "strategy_coeff": None if row[0] is None else float(row[0]),
        "revenue_transfer": 0.0 if row[1] is None else float(row[1]),
        "note": row[2],
    }


def _read_prices(d: Date) -> dict:
    """Return hour -> (price_da, price_rt)."""
    with db_manager.engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT hour, price_da, price_rt
                FROM cache_daily_hourly
                WHERE record_date=:d
                """
            ),
            {"d": d},
        ).fetchall()
    out = {}
    for h, pda, prt in rows:
        try:
            hh = int(h)
        except Exception:
            continue
        if 0 <= hh <= 23:
            out[hh] = (
                None if pda is None else float(pda),
                None if prt is None else float(prt),
            )
    return out


def _compute_weighted_forecast(target: Date, platform: Optional[str] = None) -> dict:
    """
    月度预测电量（分时）：
    forecast(h) = 0.5 * actual(D-7,h) + 0.3 * actual(D-14,h) + 0.2 * actual(D-21,h)
    """
    # If an explicit forecast exists (e.g., a month-specific forecast after portfolio/user changes),
    # prefer it to avoid mixing with historical series that no longer represent the same load.
    try:
        explicit = _read_forecast_hourly(target, platform=platform)
        if len(explicit) >= 24 and all(explicit.get(h) is not None for h in range(24)):
            hourly = [float(explicit[h]) for h in range(24)]
            return {
                "target_date": target.isoformat(),
                "source": "explicit",
                "source_dates": [{"date": "explicit_forecast", "weight": 1.0}],
                "source_status": [{"date": "explicit_forecast", "weight": 1.0, "count": 24, "has_24": True}],
                "hourly": hourly,
                "total": float(sum(hourly)),
            }
    except Exception:
        pass

    sources = [
        (target - timedelta(days=7), 0.5),
        (target - timedelta(days=14), 0.3),
        (target - timedelta(days=21), 0.2),
    ]
    src_maps = []
    sources_status = []
    missing = []
    actual_maps = _read_actual_hourly_multi([d for d, _w in sources], platform=platform)
    for d, w in sources:
        m = actual_maps.get(d) or {}
        if len(m) < 24:
            missing.append(d.isoformat())
        sources_status.append({"date": d.isoformat(), "weight": w, "count": len(m), "has_24": len(m) >= 24})
        src_maps.append((m, w, d))
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少用于预测的历史实际分时电量: {', '.join(missing)}")

    hourly = []
    for h in range(24):
        v = 0.0
        for m, w, _d in src_maps:
            v += float(m.get(h) or 0.0) * float(w)
        hourly.append(v)

    return {
        "target_date": target.isoformat(),
        "source": "weighted",
        "source_dates": [{"date": d.isoformat(), "weight": w} for d, w in sources],
        "source_status": sources_status,
        "hourly": hourly,
        "total": float(sum(hourly)),
    }


def _hourly_win_stats_for_date(d: Date, platform: Optional[str] = None) -> dict:
    """
    Strategy win-rate computed per-hour:
    - based on coeff vs 1 and price_diff sign
      * coeff < 1 and price_diff > 0 => correct
      * coeff >= 1 and price_diff < 0 => correct

    price_diff == 0 is excluded.
    """
    price_diff = _read_price_diff(d, platform=platform)
    coeff = _read_hourly_coeff(d, platform=platform)

    s_total = 0
    s_correct = 0
    s_total_active = 0
    s_correct_active = 0

    for h in range(24):
        diff = price_diff.get(h)
        if diff is None:
            continue
        diff = float(diff)
        if diff == 0:
            continue

        ch = coeff.get(h)
        if ch is not None:
            s_total += 1
            c = float(ch)
            # Correct rule:
            # - coeff < 1  => bias to more real-time, correct when DA>RT (diff>0)
            # - coeff >= 1 => bias to more day-ahead, correct when DA<RT (diff<0)
            ok = (c < 1 and diff > 0) or (c >= 1 and diff < 0)
            if ok:
                s_correct += 1
            # "Active strategy" excludes neutral coeff==1
            if abs(c - 1.0) > 1e-9:
                s_total_active += 1
                if ok:
                    s_correct_active += 1

    return {
        "strategy_correct": s_correct,
        "strategy_total": s_total,
        "strategy_win_rate": (s_correct / s_total) if s_total else None,
        "strategy_correct_active": s_correct_active,
        "strategy_total_active": s_total_active,
        "strategy_win_rate_active": (s_correct_active / s_total_active) if s_total_active else None,
    }


def _ratio_accuracy(a: float, b: float) -> Optional[float]:
    """Return symmetric accuracy in [0,1] based on two totals."""
    try:
        a = float(a)
        b = float(b)
    except Exception:
        return None
    denom = max(a, b)
    if denom <= 0:
        return None
    return min(a, b) / denom


def _compute_declared_from_forecast(
    d: Date, forecast_hourly: List[float], coeff_fallback: Optional[float], platform: Optional[str] = None
) -> dict:
    """
    Compute declared hourly using stored hourly coeff if available.
    Fallback order per hour:
      - hourly coeff
      - coeff_fallback (legacy scalar from day_settings)
      - 1.0
    """
    coeff_hourly_map = _read_hourly_coeff(d, platform=platform)
    out_coeff = []
    declared = []
    for h in range(24):
        base = float(forecast_hourly[h] or 0.0)
        ch = coeff_hourly_map.get(h)
        if ch is None:
            ch = 1.0 if coeff_fallback is None else float(coeff_fallback)
        out_coeff.append(float(ch))
        # Treat declared energy as a submitted value: keep it consistent with typical sheet rounding (0.001).
        declared.append(round(base * float(ch), 3))
    return {"coeff_hourly": out_coeff, "declared_hourly": declared, "declared_total": float(sum(declared))}


def _compute_assessment_recovery(actual_hourly: dict, declared_hourly: List[float], price_diff: dict) -> dict:
    """
    Compute deviation assessment recovery under rule:
      - declared must be within [0.8*actual, 1.2*actual]
      - profit from exceeding that band is recovered if positive; losses are not recovered.
    Uses incremental-profit representation: (A - D) * (DA-RT).
    """
    recovered = 0.0
    used_hours = 0
    for h in range(24):
        a = actual_hourly.get(h)
        diff = price_diff.get(h)
        if a is None or diff is None:
            continue
        # Align with reference sheet rounding:
        # - actual: 0.0001
        # - declared (submitted): 0.001
        a = round(float(a), 4)
        d = round(float(declared_hourly[h]), 3)
        diff = float(diff)
        if a < 0:
            continue
        lo = 0.8 * a
        hi = 1.2 * a
        d_clamp = d
        if d < lo:
            d_clamp = lo
        elif d > hi:
            d_clamp = hi

        inc = (a - d) * diff
        inc_clamp = (a - d_clamp) * diff
        recovered += max(0.0, inc - inc_clamp)
        used_hours += 1
    return {"assessment_recovery": float(recovered), "hours": used_hours}


def _compute_profit_raw_expected(
    actual_hourly: dict, declared_hourly: List[float], coeff_hourly: List[float], price_diff: dict
) -> dict:
    """
    Compute per-day profit using incremental-profit representation:
      profit_raw(h)      = (A - D) * (DA-RT)
      profit_expected(h) = (A - clamp(A*coeff, [0.8A, 1.2A])) * (DA-RT)

    These align with the reference workbook sheets:
      - “每日收益测算” -> profit_raw, and net profit after recovery
      - “预期收益（负荷无偏差）” -> profit_expected
    """
    raw = 0.0
    expected = 0.0
    used_hours = 0
    for h in range(24):
        a = actual_hourly.get(h)
        diff = price_diff.get(h)
        if a is None or diff is None:
            continue
        try:
            a = round(float(a), 4)
            d = round(float(declared_hourly[h]), 3)
            c = float(coeff_hourly[h])
            diff = float(diff)
        except Exception:
            continue
        if a < 0:
            continue
        lo = 0.8 * a
        hi = 1.2 * a
        # Expected declared removes monthly forecast error: use actual * coeff, then clamp to penalty band.
        # Expected declared removes monthly forecast error: use actual * coeff (and keep sheet-like rounding).
        d_exp = round(a * c, 5)
        d_exp_clamp = d_exp
        if d_exp < lo:
            d_exp_clamp = lo
        elif d_exp > hi:
            d_exp_clamp = hi
        d_exp_clamp = round(float(d_exp_clamp), 5)
        raw += (a - d) * diff
        expected += (a - d_exp_clamp) * diff
        used_hours += 1
    if used_hours <= 0:
        return {"profit_raw": None, "profit_expected": None, "hours": 0}
    return {"profit_raw": float(raw), "profit_expected": float(expected), "hours": used_hours}


def _compute_daily_metrics(d: Date, platform: Optional[str] = None) -> dict:
    """
    Compute and return a single-row daily metrics dict for strategy review.
    This is used to persist a fixed daily table and avoid recomputing on every page view.
    """
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)

    settings = _read_day_settings(d, platform=p)
    coeff_total = settings.get("strategy_coeff")

    coeff_hourly_map = _read_hourly_coeff(d, platform=p)
    coeff_vals = [coeff_hourly_map.get(h) for h in range(24)]
    coeff_present = [v for v in coeff_vals if v is not None]
    coeff_avg = (sum(coeff_present) / len(coeff_present)) if coeff_present else None
    coeff_min = min(coeff_present) if coeff_present else None
    coeff_max = max(coeff_present) if coeff_present else None

    # Forecast inputs use strategy_actual_hourly; settlement evaluation prefers settlement actuals if available.
    actual_map_forecast = _read_actual_hourly(d, platform=p)
    actual_map_settlement = _read_settlement_actual_hourly(d, platform=p)
    actual_map = actual_map_settlement if len(actual_map_settlement) >= 24 else actual_map_forecast

    actual_total = None
    if len(actual_map) >= 24:
        actual_total = float(sum(float(actual_map.get(h) or 0.0) for h in range(24)))

    forecast_total = None
    declared_total = None
    forecast_accuracy = None
    declared_accuracy = None
    forecast_bias = None
    declared_bias = None
    assessment_recovery = None
    profit_real = None
    profit_expected = None

    declared_hourly = None
    if actual_total is not None:
        try:
            forecast = _compute_weighted_forecast(d, platform=p)
            forecast_total = float(forecast["total"])
            forecast_accuracy = _ratio_accuracy(actual_total, forecast_total)
            forecast_bias = float(forecast_total) - float(actual_total)

            # Declared energy: prefer imported actual-submitted (from settlement sheet) if complete; otherwise compute.
            computed = _compute_declared_from_forecast(d, forecast["hourly"], coeff_total, platform=p)
            coeff_hourly = computed["coeff_hourly"]
            computed_declared_hourly = computed["declared_hourly"]

            declared_override = _read_declared_hourly(d, platform=p)
            override_complete = sum(1 for h in range(24) if declared_override.get(h) is not None) >= 24
            if override_complete:
                declared_hourly = [round(float(declared_override[h]), 3) for h in range(24)]
            else:
                # Fill any provided overrides, fallback to computed values.
                declared_hourly = []
                for h in range(24):
                    v = declared_override.get(h)
                    declared_hourly.append(round(float(v), 3) if v is not None else float(computed_declared_hourly[h]))

            declared_total = float(sum(float(v or 0.0) for v in declared_hourly))
            declared_accuracy = _ratio_accuracy(actual_total, declared_total)
            declared_bias = float(declared_total) - float(actual_total)

            diff = _read_price_diff(d, platform=p)
            rec = _compute_assessment_recovery(actual_map, declared_hourly, diff)
            assessment_recovery = rec["assessment_recovery"]

            ps = _compute_profit_raw_expected(actual_map, declared_hourly, coeff_hourly, diff)
            transfer = float(settings.get("revenue_transfer") or 0.0)
            if ps["profit_raw"] is not None:
                profit_real = float(ps["profit_raw"]) - float(assessment_recovery or 0.0) + transfer
            if ps["profit_expected"] is not None:
                profit_expected = float(ps["profit_expected"]) + transfer
        except Exception:
            pass

    ws = _hourly_win_stats_for_date(d, platform=p)

    return {
        "record_date": d,
        "actual_total": actual_total,
        "forecast_total": forecast_total,
        "declared_total": declared_total,
        "forecast_accuracy": forecast_accuracy,
        "declared_accuracy": declared_accuracy,
        "forecast_bias": forecast_bias,
        "declared_bias": declared_bias,
        "strategy_correct": ws["strategy_correct"],
        "strategy_total": ws["strategy_total"],
        "strategy_correct_active": ws["strategy_correct_active"],
        "strategy_total_active": ws["strategy_total_active"],
        "forecast_correct": None,
        "forecast_total_hours": None,
        "assessment_recovery": assessment_recovery,
        "profit_real": profit_real,
        "profit_expected": profit_expected,
        "coeff_total": coeff_total,
        "coeff_avg": coeff_avg,
        "coeff_min": coeff_min,
        "coeff_max": coeff_max,
    }


def _upsert_daily_metrics(rows: List[dict], platform: Optional[str] = None) -> int:
    if not rows:
        return 0
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_daily_metrics", p)
    sql = text(
        f"""
        INSERT INTO {table} (
            record_date,
            actual_total, forecast_total, declared_total,
            forecast_accuracy, declared_accuracy,
            forecast_bias, declared_bias,
            strategy_correct, strategy_total,
            strategy_correct_active, strategy_total_active,
            forecast_correct, forecast_total_hours,
            assessment_recovery,
            profit_real, profit_expected,
            coeff_total, coeff_avg, coeff_min, coeff_max
        )
        VALUES (
            :record_date,
            :actual_total, :forecast_total, :declared_total,
            :forecast_accuracy, :declared_accuracy,
            :forecast_bias, :declared_bias,
            :strategy_correct, :strategy_total,
            :strategy_correct_active, :strategy_total_active,
            :forecast_correct, :forecast_total_hours,
            :assessment_recovery,
            :profit_real, :profit_expected,
            :coeff_total, :coeff_avg, :coeff_min, :coeff_max
        )
        ON DUPLICATE KEY UPDATE
            actual_total=VALUES(actual_total),
            forecast_total=VALUES(forecast_total),
            declared_total=VALUES(declared_total),
            forecast_accuracy=VALUES(forecast_accuracy),
            declared_accuracy=VALUES(declared_accuracy),
            forecast_bias=VALUES(forecast_bias),
            declared_bias=VALUES(declared_bias),
            strategy_correct=VALUES(strategy_correct),
            strategy_total=VALUES(strategy_total),
            strategy_correct_active=VALUES(strategy_correct_active),
            strategy_total_active=VALUES(strategy_total_active),
            forecast_correct=VALUES(forecast_correct),
            forecast_total_hours=VALUES(forecast_total_hours),
            assessment_recovery=VALUES(assessment_recovery),
            profit_real=VALUES(profit_real),
            profit_expected=VALUES(profit_expected),
            coeff_total=VALUES(coeff_total),
            coeff_avg=VALUES(coeff_avg),
            coeff_min=VALUES(coeff_min),
            coeff_max=VALUES(coeff_max)
        """
    )
    with db_manager.engine.connect() as conn:
        with conn.begin():
            conn.execute(sql, rows)
    return len(rows)


def _refresh_daily_metrics_for_dates(dates: List[Date], platform: Optional[str] = None) -> int:
    p = _normalize_platform(platform)
    uniq = sorted({d for d in dates if d is not None})
    rows = []
    for d in uniq:
        rows.append(_compute_daily_metrics(d, platform=p))
    return _upsert_daily_metrics(rows, platform=p)


def _clamp_declared(actual: float, declared: float) -> float:
    lo = 0.8 * actual
    hi = 1.2 * actual
    if declared < lo:
        return lo
    if declared > hi:
        return hi
    return declared


def _compute_incremental_profit(
    actual_hourly: dict,
    declared_hourly: List[float],
    prices: dict,
) -> dict:
    """
    增量收益（相对基准：日前申报=实际电量）：
      inc = (A - D) * (P_DA - P_RT)
    偏差考核：D 不能超出 [0.8A, 1.2A]，超出部分带来的“正收益”回收（负收益不回收）。
    """
    gross = 0.0
    gross_clamped = 0.0
    used_hours = 0
    for h in range(24):
        a = actual_hourly.get(h)
        if a is None:
            continue
        pda, prt = prices.get(h, (None, None))
        if pda is None or prt is None:
            continue
        d = float(declared_hourly[h])
        diff_p = float(pda) - float(prt)
        inc = (float(a) - d) * diff_p
        d_clamp = _clamp_declared(float(a), d)
        inc_clamp = (float(a) - float(d_clamp)) * diff_p
        gross += inc
        gross_clamped += inc_clamp
        used_hours += 1

    assessment_fee = max(0.0, gross - gross_clamped)
    net = gross - assessment_fee
    return {
        "gross": float(gross),
        "assessment_fee": float(assessment_fee),
        "net": float(net),
        "used_hours": used_hours,
    }


def _parse_actual_sheet_like_reference(df: pd.DataFrame) -> List[dict]:
    """Parse a sheet shaped like '代理实际分时电量'."""
    # Drop fully empty columns
    df = df.loc[:, ~df.isna().all(axis=0)]

    # Find date column
    date_col = None
    for c in df.columns:
        if str(c).strip() == "日期":
            date_col = c
            break
    if date_col is None:
        # fallback: first datetime-like column
        for c in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[c]):
                date_col = c
                break
    if date_col is None:
        raise HTTPException(status_code=400, detail="未找到“日期”列")

    # Build hour columns map
    hour_col = {}
    for c in df.columns:
        h = _hour_label_to_hour(c)
        if h is None:
            continue
        if 0 <= h <= 23:
            hour_col[h] = c
    missing_hours = [h for h in range(24) if h not in hour_col]
    if missing_hours:
        raise HTTPException(status_code=400, detail=f"缺少分时列: {missing_hours}")

    records = []
    for _, row in df.iterrows():
        raw_date = row.get(date_col)
        if pd.isna(raw_date):
            continue
        try:
            d = pd.to_datetime(raw_date).date()
        except Exception:
            # Some templates include summary rows like "均值" in the date column.
            continue
        for h in range(24):
            v = row.get(hour_col[h])
            if v is None or (isinstance(v, float) and np.isnan(v)):
                continue
            try:
                fv = float(v)
            except Exception:
                continue
            records.append({"record_date": d, "hour": int(h), "actual_energy": fv})
    return records


def _parse_strategy_sheet_coeff(df: pd.DataFrame) -> List[dict]:
    """
    Parse a sheet shaped like '日前报价策略' and extract hourly strategy coefficients.

    Returns list of:
      {
        "date": date,
        "strategy_coeff": <optional total coeff from 总量列>,
        "strategy_coeff_hourly": [24 floats],
        "note": <报价策略文本(可选)>
      }
    """
    df = df.loc[:, ~df.isna().all(axis=0)]
    if df.shape[1] < 6:
        return []

    date_col = df.columns[0]
    strategy_text_col = df.columns[1] if df.shape[1] >= 2 else None
    label_col = df.columns[2] if df.shape[1] >= 3 else None

    total_col = None
    for c in df.columns:
        if str(c).strip() == "总量":
            total_col = c
            break

    hour_cols = {}
    for c in df.columns:
        h = _hour_label_to_hour(c)
        if h is None:
            continue
        hour_cols[h] = c

    if label_col is None or len(hour_cols) < 24:
        return []

    df = df.copy()
    df[date_col] = df[date_col].ffill()
    if strategy_text_col is not None:
        df[strategy_text_col] = df[strategy_text_col].ffill()

    out = []
    for _, r in df.iterrows():
        if str(r.get(label_col)).strip() != "策略系数":
            continue
        raw_date = r.get(date_col)
        if pd.isna(raw_date):
            continue
        try:
            d = pd.to_datetime(raw_date).date()
        except Exception:
            continue

        coeff_total = None
        if total_col is not None:
            v = r.get(total_col)
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                try:
                    coeff_total = float(v)
                except Exception:
                    coeff_total = None

        coeff_hourly = []
        ok = True
        for h in range(24):
            v = r.get(hour_cols.get(h))
            if v is None or (isinstance(v, float) and np.isnan(v)):
                ok = False
                coeff_hourly.append(None)
                continue
            try:
                coeff_hourly.append(float(v))
            except Exception:
                ok = False
                coeff_hourly.append(None)
        if not ok:
            # still keep, but UI/logic can decide; for import we accept and store None where missing.
            pass

        note = None
        if strategy_text_col is not None:
            v = r.get(strategy_text_col)
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                note = str(v)

        out.append(
            {
                "date": d,
                "strategy_coeff": coeff_total,
                "strategy_coeff_hourly": coeff_hourly,
                "note": note,
            }
        )
    return out


def _parse_profit_total_sheet(df: pd.DataFrame) -> List[dict]:
    """
    Parse sheets like:
      - 每日收益测算 (profit_real)
      - 预期收益（负荷无偏差） (profit_expected)
    We only extract per-day total profit from column 1 (盈利（元）) on the date row.
    """
    # Read as-is; template often includes blank rows and multi-row blocks.
    df = df.copy()
    if df.shape[1] < 2:
        return []
    out = []
    for _, row in df.iterrows():
        raw_date = row.iloc[0]
        if pd.isna(raw_date):
            continue
        try:
            d = pd.to_datetime(raw_date).date()
        except Exception:
            continue
        v = row.iloc[1]
        if v is None or (isinstance(v, float) and np.isnan(v)):
            continue
        try:
            out.append({"date": d, "profit": float(v)})
        except Exception:
            continue
    return out


def _parse_price_hourly_from_profit_sheet(df: pd.DataFrame) -> List[dict]:
    """
    Parse hourly prices from '每日收益测算' style sheet.

    Expected columns (header=0):
      - 日期
      - 量价（元/MWh、万kWh）: contains '电价'
      - Unnamed: 3: contains '日前'/'实时'/'差值'
      - Hour columns 00:00..23:00
    """
    df = df.copy()
    if df.empty:
        return []

    # Identify date column
    date_col = None
    for c in df.columns:
        if str(c).strip() == "日期":
            date_col = c
            break
    if date_col is None:
        date_col = df.columns[0]

    # Identify type/side columns by position (stable in your template)
    type_col = df.columns[2] if len(df.columns) > 2 else None
    side_col = df.columns[3] if len(df.columns) > 3 else None

    hour_cols = {}
    for c in df.columns:
        h = _hour_label_to_hour(c)
        if h is not None and 0 <= h <= 23:
            hour_cols[h] = c
    if len(hour_cols) < 24 or type_col is None or side_col is None:
        return []

    df[date_col] = df[date_col].ffill()
    # Merged-cell templates only set "电价" on the first row of the block.
    if type_col is not None:
        df[type_col] = df[type_col].ffill()

    # Collect rows per date
    by_date = {}
    for _, r in df.iterrows():
        if str(r.get(type_col)).strip() != "电价":
            continue
        side = str(r.get(side_col)).strip()
        if side not in ("日前", "实时", "差值"):
            continue
        raw_date = r.get(date_col)
        if pd.isna(raw_date):
            continue
        try:
            d = pd.to_datetime(raw_date).date()
        except Exception:
            continue

        m = by_date.setdefault(d, {})
        m[side] = {h: r.get(hour_cols[h]) for h in range(24)}

    records = []
    for d, m in by_date.items():
        da = m.get("日前", {})
        rt = m.get("实时", {})
        diff = m.get("差值", {})
        for h in range(24):
            pda = da.get(h)
            prt = rt.get(h)
            pdiff = diff.get(h)
            def _to_float(x):
                if x is None or (isinstance(x, float) and np.isnan(x)):
                    return None
                try:
                    return float(x)
                except Exception:
                    return None
            f_da = _to_float(pda)
            f_rt = _to_float(prt)
            f_diff = _to_float(pdiff)
            if f_diff is None and f_da is not None and f_rt is not None:
                f_diff = f_da - f_rt
            # Only insert rows where at least something exists (diff preferred)
            if f_da is None and f_rt is None and f_diff is None:
                continue
            records.append(
                {
                    "record_date": d,
                    "hour": int(h),
                    "price_da": f_da,
                    "price_rt": f_rt,
                    "price_diff": f_diff,
                }
            )
    return records


def _parse_declared_hourly_from_profit_sheet(df: pd.DataFrame) -> List[dict]:
    """
    Parse declared (day-ahead submitted) hourly energy from '每日收益测算' style sheet.

    Expected columns (header=0):
      - 日期
      - 量价（元/MWh、万kWh）: contains '电量'
      - Unnamed: 3: contains '日前'
      - Hour columns 00:00..23:00
    """
    df = df.copy()
    if df.empty:
        return []

    date_col = None
    for c in df.columns:
        if str(c).strip() == "日期":
            date_col = c
            break
    if date_col is None:
        date_col = df.columns[0]

    type_col = df.columns[2] if len(df.columns) > 2 else None
    side_col = df.columns[3] if len(df.columns) > 3 else None

    hour_cols = {}
    for c in df.columns:
        h = _hour_label_to_hour(c)
        if h is not None and 0 <= h <= 23:
            hour_cols[h] = c
    if len(hour_cols) < 24 or type_col is None or side_col is None:
        return []

    df[date_col] = df[date_col].ffill()
    if type_col is not None:
        df[type_col] = df[type_col].ffill()

    records = []
    for _, r in df.iterrows():
        if str(r.get(type_col)).strip() != "电量":
            continue
        if str(r.get(side_col)).strip() != "日前":
            continue
        raw_date = r.get(date_col)
        if pd.isna(raw_date):
            continue
        try:
            d = pd.to_datetime(raw_date).date()
        except Exception:
            continue

        for h in range(24):
            v = r.get(hour_cols[h])
            if v is None or (isinstance(v, float) and np.isnan(v)):
                continue
            try:
                fv = float(v)
            except Exception:
                continue
            records.append({"record_date": d, "hour": int(h), "declared_energy": round(float(fv), 3)})
    return records


def _parse_realtime_actual_hourly_from_profit_sheet(df: pd.DataFrame) -> List[dict]:
    """
    Parse settlement (real-time) actual hourly energy from '每日收益测算' style sheet.

    Expected columns (header=0):
      - 日期
      - 量价（元/MWh、万kWh）: contains '电量'
      - Unnamed: 3: contains '实时'
      - Hour columns 00:00..23:00
    """
    df = df.copy()
    if df.empty:
        return []

    date_col = None
    for c in df.columns:
        if str(c).strip() == "日期":
            date_col = c
            break
    if date_col is None:
        date_col = df.columns[0]

    type_col = df.columns[2] if len(df.columns) > 2 else None
    side_col = df.columns[3] if len(df.columns) > 3 else None

    hour_cols = {}
    for c in df.columns:
        h = _hour_label_to_hour(c)
        if h is not None and 0 <= h <= 23:
            hour_cols[h] = c
    if len(hour_cols) < 24 or type_col is None or side_col is None:
        return []

    df[date_col] = df[date_col].ffill()
    if type_col is not None:
        df[type_col] = df[type_col].ffill()

    records = []
    for _, r in df.iterrows():
        if str(r.get(type_col)).strip() != "电量":
            continue
        if str(r.get(side_col)).strip() != "实时":
            continue
        raw_date = r.get(date_col)
        if pd.isna(raw_date):
            continue
        try:
            d = pd.to_datetime(raw_date).date()
        except Exception:
            continue

        for h in range(24):
            v = r.get(hour_cols[h])
            if v is None or (isinstance(v, float) and np.isnan(v)):
                continue
            try:
                fv = float(v)
            except Exception:
                continue
            records.append({"record_date": d, "hour": int(h), "actual_energy": round(float(fv), 4)})
    return records


@app.post("/api/strategy/import-workbook")
async def import_strategy_workbook(file: UploadFile = File(...), platform: Optional[str] = Form(None)):
    """
    导入复盘/报价Excel（参考“天朗25年12月复盘.xlsx”）：
    - 实际分时电量：写入 strategy_actual_hourly
    - 策略系数：写入 strategy_day_settings + strategy_hourly_coeff
    - 小时电价/价差：从“每日收益测算”导入，写入 strategy_price_hourly
    - 日前申报电量（历史实际提交）：从“每日收益测算”导入，写入 strategy_declared_hourly
    - 实时实际电量（用于清算/考核口径）：从“每日收益测算”导入，写入 strategy_settlement_actual_hourly

    注意：每日收益（真实/预期）不再从Excel导入，改为按规则在后端计算并缓存到 strategy_daily_metrics。
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="只支持 .xlsx")
    p = _normalize_platform(platform)
    content = await file.read()
    xl = pd.ExcelFile(BytesIO(content))

    # Heuristics: pick first matching sheets, fallback to first/second sheet.
    actual_sheet = None
    strategy_sheet = None
    for name in xl.sheet_names:
        if actual_sheet is None and ("实际" in name and "分时" in name and "电量" in name):
            actual_sheet = name
        if strategy_sheet is None and ("报价" in name or "策略" in name):
            strategy_sheet = name
    if actual_sheet is None:
        actual_sheet = xl.sheet_names[0]
    if strategy_sheet is None and len(xl.sheet_names) >= 2:
        strategy_sheet = xl.sheet_names[1]

    profit_sheet_real = None
    profit_sheet_expected = None
    for name in xl.sheet_names:
        if profit_sheet_real is None and ("收益测算" in name):
            profit_sheet_real = name
        if profit_sheet_expected is None and ("预期收益" in name or "无偏差" in name):
            profit_sheet_expected = name

    inserted_actual = 0
    inserted_settings = 0
    inserted_price_hourly = 0
    inserted_declared_hourly = 0
    inserted_settlement_actual_hourly = 0

    # Actual hourly
    try:
        df_actual = pd.read_excel(BytesIO(content), sheet_name=actual_sheet, header=0, engine="openpyxl")
        actual_records = _parse_actual_sheet_like_reference(df_actual)
        inserted_actual = _upsert_actual_hourly(actual_records, platform=p)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析实际分时电量失败: {e}") from e

    # Strategy coeff
    if strategy_sheet is not None:
        try:
            df_strategy = pd.read_excel(BytesIO(content), sheet_name=strategy_sheet, header=0, engine="openpyxl")
            settings = _parse_strategy_sheet_coeff(df_strategy)
            for s in settings:
                _upsert_day_settings(s["date"], s.get("strategy_coeff"), 0.0, s.get("note"), platform=p)
                if s.get("strategy_coeff_hourly") is not None:
                    _upsert_hourly_coeff(s["date"], s["strategy_coeff_hourly"], platform=p)
            inserted_settings = len(settings)
        except Exception:
            # Not fatal: allow import actuals only.
            inserted_settings = 0

    # Hourly prices/diffs (from profit sheet)
    if profit_sheet_real is not None:
        try:
            df_profit_price = pd.read_excel(BytesIO(content), sheet_name=profit_sheet_real, header=0, engine="openpyxl")
            price_rows = _parse_price_hourly_from_profit_sheet(df_profit_price)
            inserted_price_hourly = _upsert_price_hourly(price_rows, platform=p)
        except Exception:
            inserted_price_hourly = 0

        # Also import declared day-ahead energy (actual submitted) from the same sheet.
        try:
            declared_rows = _parse_declared_hourly_from_profit_sheet(df_profit_price)
            inserted_declared_hourly = _upsert_declared_hourly(declared_rows, platform=p)
        except Exception:
            inserted_declared_hourly = 0

        # Also import settlement actual hourly energy (real-time).
        try:
            rt_actual_rows = _parse_realtime_actual_hourly_from_profit_sheet(df_profit_price)
            inserted_settlement_actual_hourly = _upsert_settlement_actual_hourly(rt_actual_rows, platform=p)
        except Exception:
            inserted_settlement_actual_hourly = 0

    return {
        "status": "success",
        "platform": p,
        "sheets": xl.sheet_names,
        "used": {"actual_sheet": actual_sheet, "strategy_sheet": strategy_sheet},
        "inserted": {
            "actual_hourly": inserted_actual,
            "day_settings": inserted_settings,
            "price_hourly": inserted_price_hourly,
            "declared_hourly": inserted_declared_hourly,
            "settlement_actual_hourly": inserted_settlement_actual_hourly,
        },
    }


@app.post("/api/strategy/actual-hourly/upload")
async def upload_strategy_actual_hourly(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    target_date: Optional[str] = Form(None),
    record_date: Optional[str] = Form(None),
    platform: Optional[str] = Form(None),
):
    """
    上传实际分时电量（单独上传一张表也可以）。

    业务约束：日常只需要更新目标日 D 的 D-5 实际分时电量。
    - 若传入 target_date(=D)，则仅写入 D-5 当天的 24 点数据（其余日期忽略），避免误导入。
    - 若传入 record_date(=YYYY-MM-DD)，则仅写入该日期（优先于 target_date）。
    - 不传 target_date 则按文件内容全量写入（适合首次导入/补历史）。
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="只支持 .xlsx")
    p = _normalize_platform(platform)
    content = await file.read()
    xl = pd.ExcelFile(BytesIO(content))
    target_sheet = sheet_name or xl.sheet_names[0]
    if target_sheet not in xl.sheet_names:
        raise HTTPException(status_code=400, detail=f"sheet 不存在: {target_sheet}")
    df = pd.read_excel(BytesIO(content), sheet_name=target_sheet, header=0, engine="openpyxl")
    records = _parse_actual_sheet_like_reference(df)

    expected_date = None
    fixed_date = None
    if record_date:
        fixed_date = _parse_iso_date(record_date)
        records = [r for r in records if r.get("record_date") == fixed_date]
    if target_date:
        d = _parse_iso_date(target_date)
        expected_date = (d - timedelta(days=5))
        # If record_date is provided, it wins (fixed import for maintenance).
        if fixed_date is None:
            records = [r for r in records if r.get("record_date") == expected_date]

    inserted = _upsert_actual_hourly(records, platform=p)
    return {
        "status": "success",
        "platform": p,
        "sheet": target_sheet,
        "inserted": inserted,
        "expected_date": None if expected_date is None else expected_date.isoformat(),
        "record_date": None if fixed_date is None else fixed_date.isoformat(),
    }


def _list_actual_hourly_summary(start: Date, end: Date, platform: Optional[str] = None) -> List[dict]:
    """Return a per-day summary for strategy_actual_hourly + corresponding hourly coeff coverage."""
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_actual_hourly", p)
    coeff_table = _strategy_table("strategy_hourly_coeff", p)
    if end < start:
        raise HTTPException(status_code=400, detail="start 不能大于 end")

    q = text(
        f"""
        SELECT
          record_date,
          SUM(CASE WHEN actual_energy IS NOT NULL THEN 1 ELSE 0 END) AS cnt,
          SUM(actual_energy) AS total_energy,
          MAX(updated_at) AS updated_at
        FROM {table}
        WHERE record_date >= :start AND record_date <= :end
        GROUP BY record_date
        """
    )
    q_coeff = text(
        f"""
        SELECT
          record_date,
          SUM(CASE WHEN coeff IS NOT NULL THEN 1 ELSE 0 END) AS cnt,
          MAX(updated_at) AS updated_at
        FROM {coeff_table}
        WHERE record_date >= :start AND record_date <= :end
        GROUP BY record_date
        """
    )
    stats = {}
    coeff_stats = {}
    with db_manager.engine.connect() as conn:
        for d, cnt, total_energy, updated_at in conn.execute(q, {"start": start, "end": end}).fetchall():
            try:
                dd = d if isinstance(d, Date) else Date.fromisoformat(str(d))
            except Exception:
                continue
            stats[dd] = {
                "count": int(cnt or 0),
                "total_energy": None if total_energy is None else float(total_energy),
                "updated_at": None if updated_at is None else _normalize_dt(updated_at),
            }
        for d, cnt, updated_at in conn.execute(q_coeff, {"start": start, "end": end}).fetchall():
            try:
                dd = d if isinstance(d, Date) else Date.fromisoformat(str(d))
            except Exception:
                continue
            coeff_stats[dd] = {
                "count": int(cnt or 0),
                "updated_at": None if updated_at is None else _normalize_dt(updated_at),
            }

    days = []
    cur = start
    while cur <= end:
        s = stats.get(cur)
        cnt = int(s["count"]) if s else 0
        total_energy = None
        if s and cnt >= 24:
            total_energy = s.get("total_energy")
        cs = coeff_stats.get(cur)
        ccnt = int(cs["count"]) if cs else 0
        days.append(
            {
                "date": cur.isoformat(),
                "count": cnt,
                "has_24": bool(cnt >= 24),
                "updated_at": None if not s or s["updated_at"] is None else s["updated_at"].isoformat(sep=" "),
                "total_energy": None if total_energy is None else float(total_energy),
                "coeff_count": ccnt,
                "coeff_has_24": bool(ccnt >= 24),
                "coeff_updated_at": None if not cs or cs["updated_at"] is None else cs["updated_at"].isoformat(sep=" "),
            }
        )
        cur = cur + timedelta(days=1)
    # newest first for UI
    days.reverse()
    return days


@app.get("/api/strategy/actual-hourly/summary")
async def strategy_actual_hourly_summary(start: Optional[str] = None, end: Optional[str] = None, platform: Optional[str] = None):
    """
    按日期范围汇总“实际分时电量”入库情况（用于维护台账/补数）。
    返回包含缺失日期（count=0），便于一眼看出哪天没数据。
    """
    today = Date.today()
    end_d = _parse_iso_date(end) if end else today
    start_d = _parse_iso_date(start) if start else (end_d - timedelta(days=30))
    p = _normalize_platform(platform)
    return {
        "status": "success",
        "platform": p,
        "start": start_d.isoformat(),
        "end": end_d.isoformat(),
        "days": _list_actual_hourly_summary(start_d, end_d, platform=p),
    }


def _normalize_dt(value):
    # Accept datetime/date strings, and strip tz info to keep UI stable.
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, datetime.date) and not isinstance(value, datetime.datetime):
        return datetime.datetime.combine(value, datetime.datetime.min.time())
    if isinstance(value, str):
        try:
            return datetime.datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
            return None
    return None


@app.get("/api/strategy/actual-hourly")
async def get_strategy_actual_hourly(
    date: str,
    platform: Optional[str] = None,
    source: Literal["actual", "settlement"] = "actual",
):
    """读取某天 24 点“实际分时电量”（用于维护面板回填/检查）。"""
    d = _parse_iso_date(date)
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    table = _strategy_table("strategy_actual_hourly", p) if source == "actual" else _strategy_table("strategy_settlement_actual_hourly", p)

    with db_manager.engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT hour, actual_energy, updated_at FROM {table} WHERE record_date=:d"),
            {"d": d},
        ).fetchall()

    hourly = [None] * 24
    updated_max = None
    for h, v, updated_at in rows:
        try:
            hh = int(h)
        except Exception:
            continue
        if 0 <= hh <= 23:
            hourly[hh] = None if v is None else float(v)
        u = _normalize_dt(updated_at)
        if u is not None and (updated_max is None or u > updated_max):
            updated_max = u

    cnt = sum(1 for v in hourly if v is not None)
    return {
        "status": "success",
        "platform": p,
        "date": d.isoformat(),
        "source": source,
        "count": int(cnt),
        "has_24": bool(cnt >= 24),
        "updated_at": None if updated_max is None else updated_max.isoformat(sep=" "),
        "hourly": hourly,
    }


@app.post("/api/strategy/actual-hourly")
async def upsert_strategy_actual_hourly(payload: StrategyActualHourlyIn):
    """
    直接写入“实际分时电量”到数据库（不依赖Excel）。

    - source=actual: 写入 strategy_actual_hourly
    - source=settlement: 写入 strategy_settlement_actual_hourly
    - source=both: 两张表都写（默认）
    """
    d = _parse_iso_date(payload.date)
    p = _normalize_platform(payload.platform)
    if len(payload.hourly) != 24:
        raise HTTPException(status_code=400, detail="hourly 长度必须为24")

    rows = []
    for h, v in enumerate(payload.hourly):
        if v is None:
            continue
        try:
            fv = float(v)
        except Exception:
            continue
        # Store raw float; profit/assessment computations apply their own rounding rules.
        rows.append({"record_date": d, "hour": int(h), "actual_energy": float(fv)})

    inserted_actual = 0
    inserted_settlement = 0
    if payload.source in ("actual", "both"):
        inserted_actual = _upsert_actual_hourly(rows, platform=p)
    if payload.source in ("settlement", "both"):
        inserted_settlement = _upsert_settlement_actual_hourly(rows, platform=p)

    return {
        "status": "success",
        "platform": p,
        "date": d.isoformat(),
        "inserted": {
            "actual_hourly": int(inserted_actual),
            "settlement_actual_hourly": int(inserted_settlement),
        },
    }


@app.post("/api/strategy/actual-hourly/batch")
async def upsert_strategy_actual_hourly_batch(payload: List[StrategyActualHourlyIn]):
    """批量写入多天的实际分时电量到数据库（不依赖Excel）。"""
    inserted_actual = 0
    inserted_settlement = 0
    dates = []
    for item in payload:
        d = _parse_iso_date(item.date)
        p = _normalize_platform(item.platform)
        dates.append(d.isoformat())
        if len(item.hourly) != 24:
            raise HTTPException(status_code=400, detail=f"{item.date} hourly 长度必须为24")
        rows = []
        for h, v in enumerate(item.hourly):
            if v is None:
                continue
            try:
                fv = float(v)
            except Exception:
                continue
            # Store raw float; profit/assessment computations apply their own rounding rules.
            rows.append({"record_date": d, "hour": int(h), "actual_energy": float(fv)})
        if item.source in ("actual", "both"):
            inserted_actual += _upsert_actual_hourly(rows, platform=p)
        if item.source in ("settlement", "both"):
            inserted_settlement += _upsert_settlement_actual_hourly(rows, platform=p)

    return {
        "status": "success",
        "dates": dates,
        "inserted": {
            "actual_hourly": int(inserted_actual),
            "settlement_actual_hourly": int(inserted_settlement),
        },
    }


@app.post("/api/strategy/day-settings")
async def set_strategy_day_settings(payload: StrategyDaySettingsIn):
    d = _parse_iso_date(payload.date)
    p = _normalize_platform(payload.platform)
    if payload.strategy_coeff_hourly is not None:
        if len(payload.strategy_coeff_hourly) != 24:
            raise HTTPException(status_code=400, detail="strategy_coeff_hourly 长度必须为24")
        _upsert_hourly_coeff(d, payload.strategy_coeff_hourly, platform=p)

        # Also store a derived "total coeff" for quick display if not provided.
        coeff_total = payload.strategy_coeff
        if coeff_total is None:
            vals = [v for v in payload.strategy_coeff_hourly if v is not None]
            coeff_total = (sum(vals) / len(vals)) if vals else None
        _upsert_day_settings(d, coeff_total, payload.revenue_transfer, payload.note, platform=p)
    else:
        # Back-compat: only a scalar coeff
        _upsert_day_settings(d, payload.strategy_coeff, payload.revenue_transfer, payload.note, platform=p)
    return {"status": "success", "platform": p, "date": d.isoformat()}


@app.get("/api/strategy/day-settings")
async def get_strategy_day_settings(date: str, platform: Optional[str] = None):
    """读取某天策略系数（逐时24个 + 总系数/备注），用于台账维护回填。"""
    d = _parse_iso_date(date)
    p = _normalize_platform(platform)
    settings = _read_day_settings(d, platform=p)
    coeff_map = _read_hourly_coeff(d, platform=p)
    hourly = [None] * 24
    for h in range(24):
        hourly[h] = coeff_map.get(h)
    cnt = sum(1 for v in hourly if v is not None)
    return {
        "status": "success",
        "platform": p,
        "date": d.isoformat(),
        "settings": settings,
        "coeff_count": int(cnt),
        "coeff_has_24": bool(cnt >= 24),
        "strategy_coeff_hourly": hourly,
    }


@app.get("/api/strategy/quote")
async def strategy_quote(
    date: str,
    platform: Optional[str] = Query(None, description="平台：天朗/辉华（默认天朗）"),
    strategy_coeff: Optional[float] = Query(None, description="可选：临时覆盖策略系数"),
):
    """返回某天的预测电量 + 申报电量（默认使用已保存策略系数，可用参数覆盖）。"""
    d = _parse_iso_date(date)
    p = _normalize_platform(platform)
    # If called as a plain function (not via FastAPI injection), `strategy_coeff` may still be a Query() object.
    if strategy_coeff is not None and not isinstance(strategy_coeff, (int, float)):
        strategy_coeff = None
    settings = _read_day_settings(d, platform=p)
    coeff_hourly_map = _read_hourly_coeff(d, platform=p)
    coeff_hourly = [None] * 24
    coeff_source = None
    for h in range(24):
        coeff_hourly[h] = coeff_hourly_map.get(h)

    # Prefer hourly coefficients if we have (most/all) of them.
    if any(v is not None for v in coeff_hourly):
        coeff_source = "hourly"
    coeff = strategy_coeff if strategy_coeff is not None else settings["strategy_coeff"]
    forecast = _compute_weighted_forecast(d, platform=p)

    declared = []
    used_coeff_hourly = []
    for h in range(24):
        base = float(forecast["hourly"][h])
        if coeff_source == "hourly":
            ch = coeff_hourly[h]
            if ch is None:
                # fallback to scalar if hour missing
                ch = 1.0 if coeff is None else float(coeff)
            used_coeff_hourly.append(float(ch))
            declared.append(round(base * float(ch), 3))
        else:
            c = 1.0 if coeff is None else float(coeff)
            used_coeff_hourly.append(float(c))
            declared.append(round(base * float(c), 3))
    if coeff_source is None:
        coeff_source = "scalar" if coeff is not None else "default_1.0"

    # Daily update reminder: only D-5 actual hourly is updated day-by-day.
    update_d = d - timedelta(days=5)
    update_cnt = _count_actual_hourly(update_d, platform=p)

    return {
        "status": "success",
        "platform": p,
        "platform_label": _STRATEGY_PLATFORM_LABELS.get(p),
        "date": d.isoformat(),
        "strategy_coeff": None if coeff is None else float(coeff),
        "strategy_coeff_hourly": used_coeff_hourly,
        "strategy_coeff_source": coeff_source,
        "forecast": forecast,
        "declared": {"hourly": declared, "total": float(sum(declared))},
        "settings": settings,
        "data_reminder": {
            "update_actual_date": update_d.isoformat(),
            "update_actual_count": int(update_cnt),
            "update_actual_has_24": bool(update_cnt >= 24),
        },
    }


@app.get("/api/strategy/review")
async def strategy_review(month: str, platform: Optional[str] = None):
    """
    月度复盘：
    - 策略胜率（逐时）：按策略系数与小时价差的方向规则统计
    - 收益（真实/预期）：按“每日收益测算/预期收益（负荷无偏差）”同口径公式计算并落库缓存
    - 电量：申报/月度预测 的总量、准确率、偏差；考核回收按±20%规则计算
    """
    # month: YYYY-MM
    try:
        y, m = month.split("-")
        start = Date(int(y), int(m), 1)
    except Exception as e:
        raise HTTPException(status_code=400, detail="month 格式应为 YYYY-MM") from e
    # next month
    if start.month == 12:
        end = Date(start.year + 1, 1, 1)
    else:
        end = Date(start.year, start.month + 1, 1)

    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    t_settings = _strategy_table("strategy_day_settings", p)
    t_coeff = _strategy_table("strategy_hourly_coeff", p)
    t_metrics = _strategy_table("strategy_daily_metrics", p)
    t_price = _strategy_table("strategy_price_hourly", p)

    def _list_strategy_dates(range_start: Date, range_end: Date, inclusive_end: bool) -> List[Date]:
        op = "<=" if inclusive_end else "<"
        q = f"""
            SELECT record_date
            FROM (
                SELECT DISTINCT record_date
                FROM {t_settings}
                WHERE record_date >= :start AND record_date {op} :end
                UNION
                SELECT DISTINCT record_date
                FROM {t_coeff}
                WHERE record_date >= :start AND record_date {op} :end
            ) t
            ORDER BY record_date
        """
        with db_manager.engine.connect() as conn:
            return [r[0] for r in conn.execute(text(q), {"start": range_start, "end": range_end}).fetchall()]

    dates = _list_strategy_dates(start, end, inclusive_end=False)

    # Ensure daily metrics rows exist (fixed daily table; compute missing only).
    try:
        with db_manager.engine.connect() as conn:
            existing = {
                r[0]
                for r in conn.execute(
                    text(
                        f"""
                        SELECT record_date
                        FROM {t_metrics}
                        WHERE record_date >= :start AND record_date < :end
                        """
                    ),
                    {"start": start, "end": end},
                ).fetchall()
            }
        missing = [d for d in dates if d not in existing]
        if missing:
            _refresh_daily_metrics_for_dates(missing, platform=p)
    except Exception:
        pass

    daily = []
    # Track cumulative (from earliest strategy date)
    with db_manager.engine.connect() as conn:
        first_setting = conn.execute(
            text(
                f"""
                SELECT MIN(record_date) AS d
                FROM (
                    SELECT record_date FROM {t_settings}
                    UNION
                    SELECT record_date FROM {t_coeff}
                ) t
                """
            )
        ).scalar()

    cum_start = first_setting if first_setting else start
    today = Date.today()

    def compute_range_summary(range_start: Date, range_end: Date) -> dict:
        # Read cached metrics (fixed daily table). If missing, compute once.
        strategy_dates = _list_strategy_dates(range_start, range_end, inclusive_end=True)

        try:
            with db_manager.engine.connect() as conn:
                existing = {
                    r[0]
                    for r in conn.execute(
                        text(
                            f"""
                            SELECT record_date
                            FROM {t_metrics}
                            WHERE record_date >= :start AND record_date <= :end
                            """
                        ),
                        {"start": range_start, "end": range_end},
                    ).fetchall()
                }
            missing = [d for d in strategy_dates if d not in existing]
            if missing:
                _refresh_daily_metrics_for_dates(missing, platform=p)
        except Exception:
            pass

        with db_manager.engine.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    SELECT
                        record_date,
                        profit_expected, profit_real,
                        strategy_correct, strategy_total,
                        declared_accuracy, forecast_accuracy,
                        declared_bias, forecast_bias,
                        assessment_recovery,
                        coeff_total, coeff_avg
                    FROM {t_metrics}
                    WHERE record_date >= :start AND record_date <= :end
                    ORDER BY record_date
                    """
                ),
                {"start": range_start, "end": range_end},
            ).fetchall()

        wins = 0
        total_days = 0
        sum_expected = 0.0
        sum_real = 0.0

        s_correct = 0
        s_total = 0
        f_correct = 0
        f_total = 0
        declared_acc_vals = []
        forecast_acc_vals = []
        declared_bias_sum = 0.0
        forecast_bias_sum = 0.0
        assessment_recovery_sum = 0.0

        for (
            _d,
            p_exp,
            p_real,
            sc,
            st,
            d_acc,
            f_acc,
            d_bias,
            f_bias,
            rec,
            coeff_total,
            coeff_avg,
        ) in rows:
            if p_exp is None:
                continue
            if coeff_total is None and coeff_avg is None:
                continue
            total_days += 1
            sum_expected += float(p_exp)
            if float(p_exp) > 0:
                wins += 1
            if p_real is not None:
                sum_real += float(p_real)

            if st:
                s_correct += int(sc or 0)
                s_total += int(st)

            if d_acc is not None:
                declared_acc_vals.append(float(d_acc))
            if f_acc is not None:
                forecast_acc_vals.append(float(f_acc))
            if d_bias is not None:
                declared_bias_sum += float(d_bias)
            if f_bias is not None:
                forecast_bias_sum += float(f_bias)
            if rec is not None:
                assessment_recovery_sum += float(rec)

        win_rate = (wins / total_days) if total_days else None
        return {
            "start": range_start.isoformat(),
            "end": range_end.isoformat(),
            "days": total_days,
            "wins": wins,
            "win_rate": win_rate,
            "expected_profit_sum": float(sum_expected),
            "real_profit_sum": float(sum_real),
            "strategy_win_rate": (s_correct / s_total) if s_total else None,
            "strategy_hours": s_total,
            "declared_accuracy_avg": (sum(declared_acc_vals) / len(declared_acc_vals)) if declared_acc_vals else None,
            "forecast_accuracy_avg": (sum(forecast_acc_vals) / len(forecast_acc_vals)) if forecast_acc_vals else None,
            "declared_bias_sum": float(declared_bias_sum) if total_days else None,
            "forecast_bias_sum": float(forecast_bias_sum) if total_days else None,
            "assessment_recovery_sum": float(assessment_recovery_sum),
        }

    # Month daily rows (read fixed daily table)
    if dates:
        with db_manager.engine.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    SELECT
                        record_date,
                        coeff_avg, coeff_min, coeff_max, coeff_total,
                        actual_total,
                        declared_total,
                        forecast_total,
                        declared_accuracy,
                        forecast_accuracy,
                        declared_bias,
                        forecast_bias,
                        strategy_correct, strategy_total,
                        assessment_recovery,
                        profit_real, profit_expected
                    FROM {t_metrics}
                    WHERE record_date >= :start AND record_date < :end
                      AND (coeff_total IS NOT NULL OR coeff_avg IS NOT NULL)
                    ORDER BY record_date
                    """
                ),
                {"start": start, "end": end},
            ).fetchall()

        # Precompute per-day DA-RT diagnostics in one query (avoid per-row round trips).
        diff_stats = {}
        try:
            with db_manager.engine.connect() as conn:
                ds = conn.execute(
                    text(
                        """
                        SELECT
                            record_date,
                            AVG(CASE
                                  WHEN price_diff IS NOT NULL THEN price_diff
                                  WHEN price_da IS NOT NULL AND price_rt IS NOT NULL THEN (price_da - price_rt)
                                  ELSE NULL
                                END) AS avg_diff,
                            SUM(CASE
                                  WHEN (CASE
                                          WHEN price_diff IS NOT NULL THEN price_diff
                                          WHEN price_da IS NOT NULL AND price_rt IS NOT NULL THEN (price_da - price_rt)
                                          ELSE NULL
                                        END) > 0 THEN 1 ELSE 0
                                END) AS pos_cnt,
                            SUM(CASE
                                  WHEN (CASE
                                          WHEN price_diff IS NOT NULL THEN price_diff
                                          WHEN price_da IS NOT NULL AND price_rt IS NOT NULL THEN (price_da - price_rt)
                                          ELSE NULL
                                        END) < 0 THEN 1 ELSE 0
                                END) AS neg_cnt,
                            COUNT(CASE
                                    WHEN price_diff IS NOT NULL THEN 1
                                    WHEN price_da IS NOT NULL AND price_rt IS NOT NULL THEN 1
                                    ELSE NULL
                                  END) AS cnt
                        FROM cache_daily_hourly
                        WHERE record_date >= :start AND record_date < :end
                        GROUP BY record_date
                        """
                    ),
                    {"start": start, "end": end},
                ).fetchall()
            for d_rec, avg_diff, pos_cnt, neg_cnt, cnt in ds:
                if not d_rec:
                    continue
                cnt_i = int(cnt or 0)
                if cnt_i < 24:
                    diff_stats[d_rec] = (None, None, None)
                else:
                    diff_stats[d_rec] = (
                        None if avg_diff is None else float(avg_diff),
                        float(pos_cnt or 0) / float(cnt_i),
                        float(neg_cnt or 0) / float(cnt_i),
                    )
        except Exception:
            diff_stats = {}

        # If cache table is missing/unavailable, fallback to the smaller strategy_price_hourly.
        if not diff_stats:
            try:
                with db_manager.engine.connect() as conn:
                    ds = conn.execute(
                        text(
                            f"""
                            SELECT
                                record_date,
                                AVG(CASE
                                      WHEN price_diff IS NOT NULL THEN price_diff
                                      WHEN price_da IS NOT NULL AND price_rt IS NOT NULL THEN (price_da - price_rt)
                                      ELSE NULL
                                    END) AS avg_diff,
                                SUM(CASE
                                      WHEN (CASE
                                              WHEN price_diff IS NOT NULL THEN price_diff
                                              WHEN price_da IS NOT NULL AND price_rt IS NOT NULL THEN (price_da - price_rt)
                                              ELSE NULL
                                            END) > 0 THEN 1 ELSE 0
                                    END) AS pos_cnt,
                                SUM(CASE
                                      WHEN (CASE
                                              WHEN price_diff IS NOT NULL THEN price_diff
                                              WHEN price_da IS NOT NULL AND price_rt IS NOT NULL THEN (price_da - price_rt)
                                              ELSE NULL
                                            END) < 0 THEN 1 ELSE 0
                                    END) AS neg_cnt,
                                COUNT(CASE
                                        WHEN price_diff IS NOT NULL THEN 1
                                        WHEN price_da IS NOT NULL AND price_rt IS NOT NULL THEN 1
                                        ELSE NULL
                                      END) AS cnt
                            FROM {t_price}
                            WHERE record_date >= :start AND record_date < :end
                            GROUP BY record_date
                            """
                        ),
                        {"start": start, "end": end},
                    ).fetchall()
                for d_rec, avg_diff, pos_cnt, neg_cnt, cnt in ds:
                    if not d_rec:
                        continue
                    cnt_i = int(cnt or 0)
                    if cnt_i < 24:
                        diff_stats[d_rec] = (None, None, None)
                    else:
                        diff_stats[d_rec] = (
                            None if avg_diff is None else float(avg_diff),
                            float(pos_cnt or 0) / float(cnt_i),
                            float(neg_cnt or 0) / float(cnt_i),
                        )
            except Exception:
                pass

        for (
            dd,
            coeff_avg,
            coeff_min,
            coeff_max,
            coeff_total,
            actual_total,
            declared_total,
            forecast_total,
            declared_accuracy,
            forecast_accuracy,
            declared_bias,
            forecast_bias,
            s_correct,
            s_total,
            assessment_recovery,
            profit_real,
            profit_expected,
        ) in rows:
            price_diff_avg, price_diff_pos_share, price_diff_neg_share = diff_stats.get(dd, (None, None, None))

            daily.append(
                {
                    "date": dd.isoformat() if hasattr(dd, "isoformat") else str(dd),
                    "strategy_coeff_total": None if coeff_total is None else float(coeff_total),
                    "strategy_coeff_avg": None if coeff_avg is None else float(coeff_avg),
                    "strategy_coeff_min": None if coeff_min is None else float(coeff_min),
                    "strategy_coeff_max": None if coeff_max is None else float(coeff_max),
                    "actual_total": None if actual_total is None else float(actual_total),
                    "declared_total": None if declared_total is None else float(declared_total),
                    "forecast_total": None if forecast_total is None else float(forecast_total),
                    "declared_accuracy": None if declared_accuracy is None else float(declared_accuracy),
                    "forecast_accuracy": None if forecast_accuracy is None else float(forecast_accuracy),
                    "declared_bias": None if declared_bias is None else float(declared_bias),
                    "forecast_bias": None if forecast_bias is None else float(forecast_bias),
                    "strategy_win_rate": (float(s_correct) / float(s_total)) if s_total else None,
                    "strategy_hours": int(s_total or 0),
                    "assessment_recovery": None if assessment_recovery is None else float(assessment_recovery),
                    "profit_real": None if profit_real is None else float(profit_real),
                    "profit_expected": None if profit_expected is None else float(profit_expected),
                    "price_diff_avg": None if price_diff_avg is None else float(price_diff_avg),
                    "price_diff_pos_share": None if price_diff_pos_share is None else float(price_diff_pos_share),
                    "price_diff_neg_share": None if price_diff_neg_share is None else float(price_diff_neg_share),
                }
            )

    # Month summary
    month_days = 0
    month_wins = 0
    month_expected_sum = 0.0
    month_real_sum = 0.0
    month_forecast_acc_vals = []
    month_forecast_bias_vals = []
    month_declared_acc_vals = []
    month_declared_bias_vals = []
    month_assessment_recovery_sum = 0.0
    month_s_correct = 0
    month_s_total = 0
    month_f_correct = 0
    month_f_total = 0
    for r in daily:
        if r["profit_expected"] is not None and (r.get("strategy_coeff_avg") is not None or r.get("strategy_coeff_total") is not None):
            month_days += 1
            month_expected_sum += float(r["profit_expected"])
            if float(r["profit_expected"]) > 0:
                month_wins += 1
        if r["profit_real"] is not None and (r.get("strategy_coeff_avg") is not None or r.get("strategy_coeff_total") is not None):
            month_real_sum += float(r["profit_real"])
        if r.get("forecast_accuracy") is not None:
            month_forecast_acc_vals.append(float(r["forecast_accuracy"]))
        if r.get("forecast_bias") is not None:
            month_forecast_bias_vals.append(float(r["forecast_bias"]))
        if r.get("declared_accuracy") is not None:
            month_declared_acc_vals.append(float(r["declared_accuracy"]))
        if r.get("declared_bias") is not None:
            month_declared_bias_vals.append(float(r["declared_bias"]))
        if r.get("assessment_recovery") is not None:
            month_assessment_recovery_sum += float(r["assessment_recovery"] or 0.0)
        if r.get("strategy_hours"):
            month_s_total += int(r["strategy_hours"])
            # reconstruct correct from rate is lossy; just recompute per date when needed
        if r.get("forecast_hours"):
            month_f_total += int(r["forecast_hours"])

    month_summary = {
        "month": month,
        "days": month_days,
        "wins": month_wins,
        "win_rate": (month_wins / month_days) if month_days else None,
        "expected_profit_sum": float(month_expected_sum),
        "real_profit_sum": float(month_real_sum),
        "declared_accuracy_avg": (sum(month_declared_acc_vals) / len(month_declared_acc_vals)) if month_declared_acc_vals else None,
        "forecast_accuracy_avg": (sum(month_forecast_acc_vals) / len(month_forecast_acc_vals)) if month_forecast_acc_vals else None,
        "declared_bias_sum": float(sum(month_declared_bias_vals)) if month_declared_bias_vals else None,
        "forecast_bias_sum": float(sum(month_forecast_bias_vals)) if month_forecast_bias_vals else None,
        "assessment_recovery_sum": float(month_assessment_recovery_sum) if month_assessment_recovery_sum else 0.0,
        "strategy_win_rate": None,
        "strategy_hours": None,
    }

    # Compute month hourly win rates from cached counts.
    s_correct = 0
    s_total = 0
    f_correct = 0
    f_total = 0
    for r in daily:
        if r.get("strategy_hours"):
            s_total += int(r["strategy_hours"])
            # win_rate already derived; use underlying totals by recomputing from counts if present
        if r.get("forecast_hours"):
            f_total += int(r["forecast_hours"])
    # Better: sum the original counts from DB for the month
    with db_manager.engine.connect() as conn:
        row = conn.execute(
            text(
                f"""
                SELECT
                    SUM(strategy_correct) AS sc,
                    SUM(strategy_total) AS st
                FROM {t_metrics}
                WHERE record_date >= :start AND record_date < :end
                  AND (coeff_total IS NOT NULL OR coeff_avg IS NOT NULL)
                """
            ),
            {"start": start, "end": end},
        ).fetchone()
    if row:
        sc, st = row[0], row[1]
        if st:
            month_summary["strategy_win_rate"] = float(sc or 0) / float(st)
            month_summary["strategy_hours"] = int(st)
        else:
            month_summary["strategy_win_rate"] = None
            month_summary["strategy_hours"] = None

    cum_summary = compute_range_summary(cum_start, today)

    return {
        "status": "success",
        "platform": p,
        "platform_label": _STRATEGY_PLATFORM_LABELS.get(p),
        "month": month,
        "summary": month_summary,
        "cumulative": cum_summary,
        "daily": daily,
    }


@app.get("/api/strategy/review/latest-month")
async def strategy_review_latest_month(platform: Optional[str] = None):
    """
    Return the most recent month (YYYY-MM) that has any strategy settings/coeff data.
    Used by the strategy review page to default to the latest available month instead of "current month".
    """
    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    t_settings = _strategy_table("strategy_day_settings", p)
    t_coeff = _strategy_table("strategy_hourly_coeff", p)
    with db_manager.engine.connect() as conn:
        ym = conn.execute(
            text(
                f"""
                SELECT DATE_FORMAT(MAX(record_date), '%Y-%m') AS ym
                FROM (
                    SELECT record_date FROM {t_settings}
                    UNION ALL
                    SELECT record_date FROM {t_coeff}
                ) t
                """
            )
        ).scalar()
    return {"status": "success", "platform": p, "platform_label": _STRATEGY_PLATFORM_LABELS.get(p), "month": ym}


@app.get("/api/strategy/review/diagnose")
async def strategy_review_diagnose(date: str, platform: Optional[str] = None):
    """
    Per-day diagnosis for strategy review:
    - Returns hourly series for the selected date (actual/declared/coeff/DA-RT and derived PnL contributions)
    - Returns month baseline (per-hour averages/stddev from cache_daily_hourly) so the UI can compare where the anomaly comes from.
    """
    d = _parse_iso_date(date)
    start = Date(d.year, d.month, 1)
    end = Date(d.year + 1, 1, 1) if d.month == 12 else Date(d.year, d.month + 1, 1)

    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    t_coeff = _strategy_table("strategy_hourly_coeff", p)
    t_declared = _strategy_table("strategy_declared_hourly", p)
    t_actual = _strategy_table("strategy_actual_hourly", p)
    t_settle = _strategy_table("strategy_settlement_actual_hourly", p)

    def _zscore(v, mu, sd):
        try:
            if v is None or mu is None or sd is None:
                return None
            sdv = float(sd)
            if sdv <= 1e-9:
                return None
            return (float(v) - float(mu)) / sdv
        except Exception:
            return None

    # Baseline (month averages per hour) from DB (fast, 4 aggregate queries).
    baseline = {h: {"hour": h} for h in range(24)}
    try:
        with db_manager.engine.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    SELECT hour, AVG(coeff) AS v, COUNT(*) AS n
                    FROM {t_coeff}
                    WHERE record_date >= :start AND record_date < :end AND coeff IS NOT NULL
                    GROUP BY hour
                    """
                ),
                {"start": start, "end": end},
            ).fetchall()
        for h, v, n in rows:
            hh = int(h)
            baseline[hh]["coeff_avg"] = None if v is None else float(v)
            baseline[hh]["coeff_n"] = int(n or 0)
    except Exception:
        pass

    try:
        with db_manager.engine.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    SELECT hour, AVG(declared_energy) AS v, COUNT(*) AS n
                    FROM {t_declared}
                    WHERE record_date >= :start AND record_date < :end AND declared_energy IS NOT NULL
                    GROUP BY hour
                    """
                ),
                {"start": start, "end": end},
            ).fetchall()
        for h, v, n in rows:
            hh = int(h)
            baseline[hh]["declared_avg"] = None if v is None else float(v)
            baseline[hh]["declared_n"] = int(n or 0)
    except Exception:
        pass

    # Prefer settlement actuals if we have coverage; otherwise fallback to strategy_actual_hourly.
    actual_table = t_settle
    try:
        with db_manager.engine.connect() as conn:
            total = conn.execute(
                text(
                    f"""
                    SELECT COUNT(*) FROM {t_settle}
                    WHERE record_date >= :start AND record_date < :end AND actual_energy IS NOT NULL
                    """
                ),
                {"start": start, "end": end},
            ).scalar()
        if not total or int(total) < 24:
            actual_table = t_actual
    except Exception:
        actual_table = t_actual

    try:
        with db_manager.engine.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    SELECT hour, AVG(actual_energy) AS v, COUNT(*) AS n
                    FROM {actual_table}
                    WHERE record_date >= :start AND record_date < :end AND actual_energy IS NOT NULL
                    GROUP BY hour
                    """
                ),
                {"start": start, "end": end},
            ).fetchall()
        for h, v, n in rows:
            hh = int(h)
            baseline[hh]["actual_avg"] = None if v is None else float(v)
            baseline[hh]["actual_n"] = int(n or 0)
    except Exception:
        pass

    try:
        with db_manager.engine.connect() as conn:
            rows = conn.execute(
                text(
                    """
                    SELECT
                      hour,
                      AVG(COALESCE(price_diff, price_da - price_rt)) AS diff_avg,
                      STDDEV_SAMP(COALESCE(price_diff, price_da - price_rt)) AS diff_std,
                      AVG(price_da) AS da_avg,
                      AVG(price_rt) AS rt_avg,
                      AVG(load_forecast) AS load_avg,
                      STDDEV_SAMP(load_forecast) AS load_std,
                      AVG(temperature) AS temp_avg,
                      STDDEV_SAMP(temperature) AS temp_std,
                      COUNT(*) AS n
                    FROM cache_daily_hourly
                    WHERE record_date >= :start AND record_date < :end
                      AND (price_diff IS NOT NULL OR (price_da IS NOT NULL AND price_rt IS NOT NULL))
                    GROUP BY hour
                    """
                ),
                {"start": start, "end": end},
            ).fetchall()
        for h, diff_avg, diff_std, da_avg, rt_avg, load_avg, load_std, temp_avg, temp_std, n in rows:
            hh = int(h)
            baseline[hh]["price_diff_avg"] = None if diff_avg is None else float(diff_avg)
            baseline[hh]["price_diff_std"] = None if diff_std is None else float(diff_std)
            baseline[hh]["price_da_avg"] = None if da_avg is None else float(da_avg)
            baseline[hh]["price_rt_avg"] = None if rt_avg is None else float(rt_avg)
            baseline[hh]["load_forecast_avg"] = None if load_avg is None else float(load_avg)
            baseline[hh]["load_forecast_std"] = None if load_std is None else float(load_std)
            baseline[hh]["temperature_avg"] = None if temp_avg is None else float(temp_avg)
            baseline[hh]["temperature_std"] = None if temp_std is None else float(temp_std)
            baseline[hh]["price_n"] = int(n or 0)
    except Exception:
        pass

    # Day series.
    try:
        actual_map_settlement = _read_settlement_actual_hourly(d, platform=p)
        actual_map_forecast = _read_actual_hourly(d, platform=p)
        actual_map = actual_map_settlement if len(actual_map_settlement) >= 24 else actual_map_forecast
    except Exception:
        actual_map = {}

    # Forecast (explicit preferred) and derived coeff_hourly.
    forecast = None
    coeff_hourly = [None] * 24
    try:
        forecast = _compute_weighted_forecast(d, platform=p)
        computed = _compute_declared_from_forecast(
            d, forecast["hourly"], _read_day_settings(d, platform=p).get("strategy_coeff"), platform=p
        )
        coeff_hourly = computed["coeff_hourly"]
        computed_declared_hourly = computed["declared_hourly"]
    except Exception:
        computed_declared_hourly = [0.0] * 24

    # Declared: prefer imported actual-submitted if complete; otherwise computed.
    declared_hourly = None
    try:
        declared_override = _read_declared_hourly(d, platform=p)
        override_complete = sum(1 for h in range(24) if declared_override.get(h) is not None) >= 24
        if override_complete:
            declared_hourly = [round(float(declared_override[h]), 3) for h in range(24)]
        else:
            declared_hourly = []
            for h in range(24):
                v = declared_override.get(h)
                declared_hourly.append(round(float(v), 3) if v is not None else float(computed_declared_hourly[h]))
    except Exception:
        declared_hourly = [float(v) for v in computed_declared_hourly]

    # Day cache data (prices/temperature/load forecast).
    cache = {}
    try:
        with db_manager.engine.connect() as conn:
            rows = conn.execute(
                text(
                    """
                    SELECT hour, price_da, price_rt, price_diff, load_forecast, temperature, day_type
                    FROM cache_daily_hourly
                    WHERE record_date=:d
                    """
                ),
                {"d": d},
            ).fetchall()
        for h, pda, prt, diff, lf, temp, day_type in rows:
            hh = int(h)
            diff_val = diff
            if diff_val is None and pda is not None and prt is not None:
                try:
                    diff_val = float(pda) - float(prt)
                except Exception:
                    diff_val = None
            cache[hh] = {
                "price_da": None if pda is None else float(pda),
                "price_rt": None if prt is None else float(prt),
                "price_diff": None if diff_val is None else float(diff_val),
                "load_forecast": None if lf is None else float(lf),
                "temperature": None if temp is None else float(temp),
                "day_type": day_type,
            }
    except Exception:
        cache = {}

    # Use the same diff mapping as the rest of the strategy logic (cache first).
    diff_map = _read_price_diff(d, platform=p)

    hours = []
    gap_hours = []
    for h in range(24):
        a = actual_map.get(h)
        dcl = None if declared_hourly is None else declared_hourly[h]
        c = coeff_hourly[h] if coeff_hourly else None
        diff = diff_map.get(h)

        raw = None
        expected = None
        recovered = None
        d_exp = None
        d_exp_clamp = None

        if a is not None and dcl is not None and diff is not None:
            try:
                a4 = round(float(a), 4)
                d3 = round(float(dcl), 3)
                df = float(diff)
                raw = (a4 - d3) * df

                # Expected declared removes forecast error: A*c then clamp into penalty band.
                if c is not None:
                    cc = float(c)
                    lo = 0.8 * a4
                    hi = 1.2 * a4
                    d_exp = round(a4 * cc, 5)
                    d_exp_clamp = min(max(d_exp, lo), hi)
                    expected = (a4 - float(d_exp_clamp)) * df

                    # Recovery: only positive extra profit outside band is recovered.
                    d_clamp = min(max(d3, lo), hi)
                    inc = (a4 - d3) * df
                    inc_clamp = (a4 - float(d_clamp)) * df
                    recovered = max(0.0, inc - inc_clamp)

                    gap_hours.append((abs((expected or 0.0) - (raw or 0.0)), h, raw, expected))
            except Exception:
                pass

        row = {
            "hour": h,
            "actual": None if a is None else float(a),
            "declared": None if dcl is None else float(dcl),
            "coeff": None if c is None else float(c),
            "price_diff": None if diff is None else float(diff),
            "profit_raw": None if raw is None else float(raw),
            "profit_expected": None if expected is None else float(expected),
            "recovered": None if recovered is None else float(recovered),
            "expected_declared": None if d_exp is None else float(d_exp),
            "expected_declared_clamped": None if d_exp_clamp is None else float(d_exp_clamp),
        }
        if h in cache:
            row.update(cache[h])

        # Z-score vs month baseline (cache_daily_hourly) for anomaly spotting.
        b = baseline.get(h, {})
        row["z_price_diff"] = _zscore(row.get("price_diff"), b.get("price_diff_avg"), b.get("price_diff_std"))
        row["z_load_forecast"] = _zscore(row.get("load_forecast"), b.get("load_forecast_avg"), b.get("load_forecast_std"))
        row["z_temperature"] = _zscore(row.get("temperature"), b.get("temperature_avg"), b.get("temperature_std"))
        hours.append(row)

    gap_hours.sort(reverse=True)
    top_hours = []
    for _abs_gap, h, raw, exp in gap_hours[:8]:
        top_hours.append(
            {
                "hour": h,
                "profit_gap": float((exp or 0.0) - (raw or 0.0)),
                "profit_raw": None if raw is None else float(raw),
                "profit_expected": None if exp is None else float(exp),
            }
        )

    # Top anomaly hours by cache Z-score (combined).
    z_rank = []
    for r in hours:
        try:
            score = 0.0
            cnt = 0
            for k in ("z_price_diff", "z_load_forecast", "z_temperature"):
                z = r.get(k)
                if z is None:
                    continue
                score += abs(float(z))
                cnt += 1
            if cnt:
                z_rank.append((score, int(r["hour"]), r))
        except Exception:
            continue
    z_rank.sort(reverse=True, key=lambda x: x[0])
    top_z_hours = []
    for _score, h, r in z_rank[:8]:
        top_z_hours.append(
            {
                "hour": h,
                "z_price_diff": r.get("z_price_diff"),
                "z_load_forecast": r.get("z_load_forecast"),
                "z_temperature": r.get("z_temperature"),
            }
        )

    # Use cached daily metrics for consistent headline numbers if available.
    try:
        dm = _compute_daily_metrics(d, platform=p)
    except Exception:
        dm = {"record_date": d}

    return {
        "status": "success",
        "platform": p,
        "platform_label": _STRATEGY_PLATFORM_LABELS.get(p),
        "date": d.isoformat(),
        "month": start.isoformat()[:7],
        "baseline": {"actual_table": actual_table, "hours": [baseline[h] for h in range(24)]},
        "day": {"summary": dm, "hours": hours, "top_hours": top_hours, "top_z_hours": top_z_hours},
        "forecast_source": None if not forecast else forecast.get("source"),
    }


@app.post("/api/strategy/review/refresh")
async def strategy_review_refresh(month: str = Form(...), platform: Optional[str] = Form(None)):
    """
    强制刷新某月的 strategy_daily_metrics（用于“页面复盘数据更新”）。

    说明：/api/strategy/review 默认只补缺失，不会覆盖已有缓存；当你更新了实际分时数据但缓存已存在时，
    用这个接口强制重算并写回 strategy_daily_metrics。
    """
    try:
        y, m = month.split("-")
        start = Date(int(y), int(m), 1)
    except Exception as e:
        raise HTTPException(status_code=400, detail="month 格式应为 YYYY-MM") from e
    end = Date(start.year + 1, 1, 1) if start.month == 12 else Date(start.year, start.month + 1, 1)

    p = _normalize_platform(platform)
    _ensure_strategy_tables(p)
    t_settings = _strategy_table("strategy_day_settings", p)
    t_coeff = _strategy_table("strategy_hourly_coeff", p)
    with db_manager.engine.connect() as conn:
        dates = [
            r[0]
            for r in conn.execute(
                text(
                    f"""
                    SELECT record_date
                    FROM (
                        SELECT DISTINCT record_date
                        FROM {t_settings}
                        WHERE record_date >= :start AND record_date < :end
                        UNION
                        SELECT DISTINCT record_date
                        FROM {t_coeff}
                        WHERE record_date >= :start AND record_date < :end
                    ) t
                    ORDER BY record_date
                    """
                ),
                {"start": start, "end": end},
            ).fetchall()
        ]

    _refresh_daily_metrics_for_dates(dates, platform=p)
    return {"status": "success", "platform": p, "platform_label": _STRATEGY_PLATFORM_LABELS.get(p), "month": month, "refreshed_days": len(dates)}


@app.post("/api/update-weather")
async def update_weather(background_tasks: BackgroundTasks):
    """手动触发天气数据更新"""
    try:
        import calendar_weather
        today = datetime.date.today()
        # 更新最近30天和未来15天的数据
        start_date = today - datetime.timedelta(days=30)
        end_date = today + datetime.timedelta(days=15)
        
        # 使用后台任务执行，避免阻塞
        def run_update():
            print(f"🌦️ 开始更新天气数据: {start_date} -> {end_date}")
            # update_calendar 内部现在会自动调用 update_price_cache_for_date(..., only_weather=True)
            # 从而实现“只更新天气表，并存入缓存表，不更新价差数据”
            calendar_weather.update_calendar(start_date, end_date)
            print("✅ 天气数据及缓存更新完成")
            
        background_tasks.add_task(run_update)
        
        return {"status": "success", "message": f"天气更新任务已启动 ({start_date} 至 {end_date})"}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"启动天气更新任务失败: {str(e)}")

@app.post("/import")
async def import_file(filename: str = Form(...), background_tasks: BackgroundTasks = BackgroundTasks()):
    """导入指定的Excel文件到数据库"""
    data_folder = "data"
    file_path = os.path.join(data_folder, filename)
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"文件 {filename} 不存在")
    
    # 定义正则表达式模式
    dated_realtime_pattern = r'\d{4}-\d{2}-\d{2}实时节点电价查询'
    dated_dayahead_pattern = r'\d{4}-\d{2}-\d{2}日前节点电价查询'

    if "负荷实际信息" in filename or "负荷预测信息" in filename:
        method = importer.import_power_data
    # elif "信息披露(区域)查询实际信息" in filename:
    #     method = importer.import_custom_excel
    # elif "信息披露(区域)查询预测信息" in filename:
    #     method = importer.import_custom_excel_pred
    elif "信息披露查询预测信息" in filename:
        method = importer.import_imformation_pred
    elif "信息披露查询实际信息" in filename:
        method = importer.import_imformation_true    
    # 先处理带日期的特殊版本
    elif re.search(dated_realtime_pattern, filename) or re.search(dated_dayahead_pattern, filename):
        method = importer.import_point_data_new
    # 然后处理不带日期的通用版本
    elif "实时节点电价查询" in filename or "日前节点电价查询" in filename:
        method = importer.import_point_data
    else:
        raise HTTPException(status_code=400, detail=f"无匹配的导入规则: {filename}")

    # 执行同步导入
    result = method(file_path)
    
    # 检查结果是否为 False (表示导入失败)
    if result is False:
        raise HTTPException(status_code=500, detail=f"导入失败: {filename}，请检查文件格式或日志")

    # [新增逻辑] 自动触发缓存更新
    try:
        # 尝试从文件名提取日期
        date_match = re.search(r"(\d{4}-\d{2}-\d{2})", filename)
        if not date_match:
             date_match = re.search(r"(\d{8})", filename)
        
        target_date = None
        if date_match:
            d_str = date_match.group(1)
            if len(d_str) == 8:
                target_date = f"{d_str[:4]}-{d_str[4:6]}-{d_str[6:]}"
            else:
                target_date = d_str

        # 触发缓存更新：电价或负荷导入都会影响 cache_daily_hourly
        should_update_cache = False
        if target_date:
            if method in (importer.import_power_data, importer.import_point_data, importer.import_point_data_new):
                should_update_cache = True
            # 保持原有行为：信息披露类文件也会尝试更新缓存
            elif "信息披露" in filename:
                should_update_cache = True
        
        if should_update_cache:
            print(f"🚀 自动触发缓存更新任务: {target_date}")
            background_tasks.add_task(update_price_cache_for_date, target_date)
            
    except Exception as e:
        print(f"⚠️ 自动触发缓存更新失败: {e}")

    if method == importer.import_imformation_pred:
        # 结果可能是单个四元组 (success, table, count, preview)
        # 也可能是多个四元组的元组 ((s1,t1,c1,p1), (s2,t2,c2,p2))
        
        # 情况1: 单个结果 (4个元素)
        if isinstance(result, tuple) and len(result) == 4 and not isinstance(result[0], tuple):
             success, table_name, record_count, preview_data = result
             
        # 情况2: 多个结果 (元组的元组)
        elif isinstance(result, tuple) and len(result) > 0 and isinstance(result[0], tuple):
             # 合并所有结果
             success = all(r[0] for r in result)
             table_name = ", ".join([str(r[1]) for r in result])
             record_count = sum(r[2] for r in result)
             # 合并预览数据 (取前几个)
             preview_data = []
             for r in result:
                 if r[3]:
                     preview_data.extend(r[3])
             preview_data = preview_data[:5] # 只保留前5条作为总预览
             
        else:
             raise HTTPException(status_code=500, detail=f"导入返回格式错误: {result}")
    
    elif method == importer.import_imformation_true:
         if isinstance(result, tuple) and len(result) == 4:
             success, table_name, record_count, preview_data = result
         # 处理可能返回None的情况（例如导入过程报错了）
         elif result is None:
             raise HTTPException(status_code=500, detail="导入失败: 内部错误")
         # 处理返回多表结果的情况 (tuple of tuples)
         elif isinstance(result, tuple) and len(result) > 0 and isinstance(result[0], tuple):
             # 合并所有结果
             success = all(r[0] for r in result)
             table_name = ", ".join([str(r[1]) for r in result])
             record_count = sum(r[2] for r in result)
             # 合并预览数据 (取前几个)
             preview_data = []
             for r in result:
                 if r[3]:
                     preview_data.extend(r[3])
             preview_data = preview_data[:5]
         else:
             # 如果是其他格式，尝试打印一下看看
             print(f"DEBUG: import_imformation_true returned: {type(result)} - {result}")
             raise HTTPException(status_code=500, detail=f"导入返回格式错误: {result}")

    elif method == importer.import_custom_excel:
        if isinstance(result, tuple) and len(result) == 3:
            # 解包三个结果元组
            (success1, table_name1, record_count1, preview_data1), (success2, table_name2, record_count2, preview_data2),(success3,table_name3,record_count3,preview_data3) = result
            # 合并结果，这里我们使用三个结果的组合
            success = success1 and success2 and success3
            table_name = f"{table_name1}, {table_name2}, {table_name3}"
            record_count = record_count1 + record_count2 + record_count3
            preview_data = preview_data1 + preview_data2 + preview_data3
        else:
             raise HTTPException(status_code=500, detail=f"导入返回格式错误: {result}")

    elif method == importer.import_custom_excel_pred:
        if isinstance(result, tuple) and len(result) == 4:
            (success1, table_name1, record_count1, preview_data1), (success2, table_name2, record_count2, preview_data2), (success4, table_name4, record_count4, preview_data4), (success5, table_name5, record_count5, preview_data5) = result
            # 合并结果，这里我们使用四个结果的组合
            success = success1 and success2 and success4 and success5
            table_name = f"{table_name1}, {table_name2}, {table_name4}, {table_name5}"
            record_count = record_count1 + record_count2 + record_count4 + record_count5 
            preview_data = preview_data1 + preview_data2 + preview_data4 + preview_data5 
        else:
             raise HTTPException(status_code=500, detail=f"导入返回格式错误: {result}")
    else:
        # 其他导入方法的常规处理
        if isinstance(result, tuple) and len(result) == 4:
            success, table_name, record_count, preview_data = result
        else:
            raise HTTPException(status_code=500, detail=f"导入返回格式错误: {result}")
        
    if success:
        return {
            "filename": filename, 
            "status": "imported", 
            "table_name": table_name, 
            "record_count": record_count,
            "preview_data": preview_data
        }
    else:
        raise HTTPException(status_code=500, detail=f"导入失败: {filename}")

@app.get("/tables")
async def get_tables():
    """获取所有数据表"""
    tables = db_manager.get_tables()
    return {"tables": tables}

@app.get("/tables/{table_name}")
async def get_table_data(table_name: str, limit: int = 5):
    """获取指定表的数据"""
    result = db_manager.get_table_data(table_name, limit)
    return result

# 新增：获取表结构信息
@app.get("/tables/{table_name}/schema")
async def get_table_schema(table_name: str):
    """获取指定表的结构信息"""
    try:
        with db_manager.engine.connect() as conn:
            result = conn.execute(text(f"DESCRIBE {table_name}"))
            schema = []
            for row in result:
                schema.append({
                    "field": row[0],
                    "type": row[1],
                    "null": row[2],
                    "key": row[3],
                    "default": row[4],
                    "extra": row[5]
                })
            return {"schema": schema}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取表结构失败: {str(e)}")

# 新增：查询表数据接口
@app.get("/tables/{table_name}/query")
async def query_table_data(table_name: str, 
                          offset: int = 0, 
                          limit: int = 20,
                          conditions: str = None):
    """查询指定表的数据，支持多条件查询
    conditions: JSON字符串，格式如 [{"column": "col1", "operator": "=", "value": "val1"}, 
                                   {"column": "col2", "operator": ">", "value": "val2"}]
    """
    try:
        with db_manager.engine.connect() as conn:
            # 构建查询条件
            where_clauses = []
            params = {}
            
            if conditions:
                import json
                try:
                    condition_list = json.loads(conditions)
                    if isinstance(condition_list, list):
                        for i, cond in enumerate(condition_list):
                            column = cond.get("column")
                            operator = cond.get("operator")
                            value = cond.get("value")
                            
                            if column and operator and value is not None:
                                # 简单的SQL注入防护
                                allowed_operators = ['=', '!=', '>', '<', '>=', '<=', 'LIKE']
                                if operator not in allowed_operators:
                                    raise HTTPException(status_code=400, detail=f"不支持的操作符: {operator}")
                                
                                param_name = f"value_{i}"
                                if operator == 'LIKE':
                                    where_clauses.append(f"{column} LIKE :{param_name}")
                                    params[param_name] = f"%{value}%"
                                else:
                                    where_clauses.append(f"{column} {operator} :{param_name}")
                                    # 尝试转换数值类型
                                    try:
                                        params[param_name] = int(value)
                                    except ValueError:
                                        try:
                                            params[param_name] = float(value)
                                        except ValueError:
                                            params[param_name] = value
                except json.JSONDecodeError:
                    raise HTTPException(status_code=400, detail="条件格式错误")
            
            # 构建WHERE子句
            where_clause = ""
            if where_clauses:
                where_clause = "WHERE " + " AND ".join(where_clauses)
            
            # 获取总记录数
            count_query = f"SELECT COUNT(*) FROM {table_name} {where_clause}"
            count_result = conn.execute(text(count_query), params)
            total_count = count_result.scalar()
            
            # 获取分页数据
            # 默认添加排序：优先按record_date倒序，其次按id倒序
            order_clause = ""
            # 简单检查表结构中是否有record_date列（可以通过查询一行数据或describe，这里简化处理，假设大部分表都有id）
            # 更稳妥的方式是直接尝试ORDER BY id DESC，如果报错则忽略
            # 但由于我们要执行SQL，这里最好直接拼接到SQL中。
            # 为了兼容性，我们先不强制加ORDER BY，除非用户没有指定排序（当前接口不支持指定排序）
            # 我们可以默认加 ORDER BY id DESC，因为大部分表都有id主键
            
            # 检查是否有id列或record_date列比较耗时，这里直接尝试按id倒序，因为我们的建表语句都包含id
            data_query = f"SELECT * FROM {table_name} {where_clause} ORDER BY id DESC LIMIT :limit OFFSET :offset"
            
            params.update({"limit": limit, "offset": offset})
            try:
                data_result = conn.execute(text(data_query), params)
            except Exception:
                # 如果失败（例如没有id列），回退到无排序
                data_query = f"SELECT * FROM {table_name} {where_clause} LIMIT :limit OFFSET :offset"
                data_result = conn.execute(text(data_query), params)
            
            data = []
            for row in data_result:
                row_dict = dict(row._mapping)
                data.append(row_dict)
            
            return {
                "data": data,
                "total": total_count,
                "offset": offset,
                "limit": limit
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"查询数据失败: {str(e)}")

@app.get("/tables/{table_name}/export")
async def export_table_data(table_name: str,
                           conditions: str = None):
    """导出指定表的数据为Excel格式，支持多条件查询
    conditions: JSON字符串，格式如 [{"column": "col1", "operator": "=", "value": "val1"}, 
                                   {"column": "col2", "operator": ">", "value": "val2"}]
    """
    try:
        print(f"导出请求开始: table_name={table_name}, conditions={conditions}")
        
        with db_manager.engine.connect() as conn:
            # 构建查询条件
            where_clauses = []
            params = {}
            
            if conditions:
                import json
                try:
                    condition_list = json.loads(conditions)
                    if isinstance(condition_list, list):
                        for i, cond in enumerate(condition_list):
                            column = cond.get("column")
                            operator = cond.get("operator")
                            value = cond.get("value")
                            
                            if column and operator and value is not None:
                                # 简单的SQL注入防护
                                allowed_operators = ['=', '!=', '>', '<', '>=', '<=', 'LIKE']
                                if operator not in allowed_operators:
                                    raise HTTPException(status_code=400, detail=f"不支持的操作符: {operator}")
                                
                                param_name = f"value_{i}"
                                if operator == 'LIKE':
                                    where_clauses.append(f"{column} LIKE :{param_name}")
                                    params[param_name] = f"%{value}%"
                                else:
                                    where_clauses.append(f"{column} {operator} :{param_name}")
                                    # 尝试转换数值类型
                                    try:
                                        params[param_name] = int(value)
                                    except ValueError:
                                        try:
                                            params[param_name] = float(value)
                                        except ValueError:
                                            params[param_name] = value
                except json.JSONDecodeError:
                    raise HTTPException(status_code=400, detail="条件格式错误")
            
            # 构建WHERE子句
            where_clause = ""
            if where_clauses:
                where_clause = "WHERE " + " AND ".join(where_clauses)
            
            # 获取所有数据
            data_query = f"SELECT * FROM {table_name} {where_clause}"
            print(f"执行查询: {data_query}, 参数: {params}")
            data_result = conn.execute(text(data_query), params)
            
            data = []
            for row in data_result:
                row_dict = dict(row._mapping)
                data.append(row_dict)
            
            print(f"查询结果数量: {len(data)}")
            if len(data) > 0:
                print(f"前几条数据示例: {data[:2]}")
            
            # 如果没有数据，返回空Excel
            if not data:
                import pandas as pd
                import numpy as np
                from io import BytesIO
                df = pd.DataFrame()
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                output.seek(0)
                
                from fastapi.responses import StreamingResponse
                return StreamingResponse(
                    iter([output.getvalue()]),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={table_name}.xlsx"}
                )
            
            # 转换为DataFrame进行处理
            import pandas as pd
            import numpy as np
            from io import BytesIO
            import os
            from datetime import datetime
            
            df = pd.DataFrame(data)
            print(f"DataFrame列: {df.columns.tolist()}")
            print(f"DataFrame形状: {df.shape}")
            if len(df) > 0:
                print(f"DataFrame前几行:\n{df.head(2)}")
            
            # 删除id列（如果存在）
            if 'id' in df.columns:
                df = df.drop(columns=['id'])
                print("已删除id列")
            
            # 检查是否包含必要的列
            required_columns = ['channel_name', 'record_date', 'record_time', 'value', 'sheet_name']
            if not all(col in df.columns for col in required_columns):
                print(f"缺少必要列，当前列: {df.columns.tolist()}")
                print("使用原始导出方式")
                # 如果不包含必要列，使用原始导出方式
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                output.seek(0)
                
                # 生成文件名
                record_date = df['record_date'].iloc[0] if 'record_date' in df.columns and len(df) > 0 else 'unknown'
                data_type = df['type'].iloc[0] if 'type' in df.columns and len(df) > 0 else 'unknown'
                
                # 格式化record_date为字符串
                if hasattr(record_date, 'strftime'):
                    record_date_str = record_date.strftime('%Y-%m-%d')
                else:
                    record_date_str = str(record_date)
                
                filename = f"{record_date_str}_{data_type}.xlsx"
                
                import urllib.parse
                encoded_filename = urllib.parse.quote(filename)
                from fastapi.responses import StreamingResponse
                return StreamingResponse(
                    iter([output.getvalue()]),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
                    }
                )
            
            # 类似preHandle.py的处理方式
            # 提取唯一的sheet_name（假设数据中sheet_name唯一）
            sheet_name = df['sheet_name'].unique()[0] if len(df['sheet_name'].unique()) > 0 else 'Sheet1'
            # 提取唯一的日期（假设数据中日期唯一）
            record_date = df['record_date'].unique()[0] if len(df['record_date'].unique()) > 0 else pd.Timestamp.now().date()
            
            # 格式化日期为YYYY-MM-DD
            if hasattr(record_date, 'strftime'):
                record_date_str = record_date.strftime('%Y-%m-%d')
            else:
                record_date_str = str(record_date)
            
            # 处理文件名特殊字符（避免斜杠、空格等导致保存失败）
            sheet_name_clean = str(sheet_name).replace('/', '_').replace('\\', '_').replace(' ', '')
            # 构造文件名：{sheet_name}({日期})_小时.xlsx
            filename = f"{sheet_name_clean}({record_date_str})_小时.xlsx"
            print(f"生成文件名: {filename}")
            
            # 检查record_time格式并处理
            print(f"record_time示例值: {df['record_time'].head()}")
            
            # 转换record_time为小时（处理各种可能的格式）
            def extract_hour(time_value):
                if pd.isna(time_value):
                    return None
                if isinstance(time_value, str):
                    if ':' in time_value:
                        # 格式如 "01:00", "1:00"
                        return int(time_value.split(':')[0])
                    else:
                        # 可能是数字字符串如 "100" 表示 01:00
                        try:
                            time_int = int(time_value)
                            return time_int // 100
                        except:
                            return None
                elif isinstance(time_value, (int, float)):
                    # 数字格式如 100 表示 01:00
                    return int(time_value) // 100
                else:
                    # timedelta或其他格式
                    try:
                        # 如果是timedelta对象
                        hours = time_value.seconds // 3600
                        return hours
                    except:
                        return None
            
            # 应用小时提取函数
            df['hour'] = df['record_time'].apply(extract_hour)
            print(f"提取的小时列示例: {df['hour'].head()}")
            
            # 删除hour为NaN的行
            df = df.dropna(subset=['hour'])
            print(f"删除无效小时后DataFrame形状: {df.shape}")
            
            # 生成电站级透视表
            if len(df) > 0:
                print("开始创建透视表")
                pivot_df = pd.pivot_table(
                    df,
                    index=['channel_name', 'record_date'],
                    columns='hour',
                    values='value',
                    aggfunc='mean'
                )
                print(f"透视表创建完成，形状: {pivot_df.shape}")
                print(f"透视表列: {pivot_df.columns.tolist()}")
                
                # 重新索引确保有24小时列
                pivot_df = pivot_df.reindex(columns=range(24), fill_value=np.nan)
                pivot_df.columns = [f'{int(h)}:00' for h in pivot_df.columns]
                pivot_df = pivot_df.reset_index()
                
                # 修改前两列名称
                pivot_df = pivot_df.rename(columns={
                    'channel_name': '节点名称',
                    'record_date': '日期'
                })
                
                # 插入单位列
                pivot_df.insert(
                    loc=2,
                    column='单位',
                    value='电价(元/MWh)'
                )
                
                # 添加发电侧全省统一均价行
                hour_columns = [f'{h}:00' for h in range(24)]
                # 确保所有小时列都存在
                for col in hour_columns:
                    if col not in pivot_df.columns:
                        pivot_df[col] = np.nan
                
                # 在计算平均值前，确保所有列为数值类型
                for col in hour_columns:
                    pivot_df[col] = pd.to_numeric(pivot_df[col], errors='coerce')
                
                final_df = pivot_df
                print(f"最终DataFrame形状: {final_df.shape}")
                print(f"最终DataFrame列: {final_df.columns.tolist()}")
                if len(final_df) > 0:
                    print(f"最终DataFrame前几行:\n{final_df.head()}")
            else:
                # 如果处理后没有数据，创建空的DataFrame
                print("处理后没有有效数据，创建空DataFrame")
                columns = ['节点名称', '日期', '单位'] + [f'{h}:00' for h in range(24)]
                final_df = pd.DataFrame(columns=columns)
            
            # 确保created文件夹存在
            created_folder = "created"
            if not os.path.exists(created_folder):
                os.makedirs(created_folder)
                print(f"创建文件夹: {created_folder}")
            
            # 生成文件名（带时间戳避免重复）
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name_with_timestamp = f"{sheet_name_clean}_{timestamp}.xlsx"
            file_path = os.path.join(created_folder, file_name_with_timestamp)
            print(f"生成文件路径: {file_path}")
            
            # 将处理后的final_df保存到服务器文件夹
            print("开始生成Excel文件到服务器")
            try:
                # 使用openpyxl引擎直接导出
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    final_df.to_excel(writer, index=False, sheet_name=sheet_name_clean[:31])
                print(f"Excel文件生成完成: {file_path}")
                
            except Exception as e:
                print(f"Excel文件生成失败: {e}")
                import traceback
                traceback.print_exc()
                
                # 回退到CSV格式
                file_name_with_timestamp = file_name_with_timestamp.replace('.xlsx', '.csv')
                file_path = os.path.join(created_folder, file_name_with_timestamp)
                final_df.to_csv(file_path, index=False)
                print(f"CSV文件生成完成: {file_path}")
            
            # 返回文件下载链接
            from fastapi.responses import JSONResponse
            download_url = f"/download/{file_name_with_timestamp}"
            return JSONResponse({
                "status": "success",
                "message": "文件生成成功",
                "download_url": download_url,
                "filename": file_name_with_timestamp
            })
            
    except HTTPException:
        raise
    except Exception as e:
        print(f"导出数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出数据失败: {str(e)}")

@app.get("/download/{filename}")
async def download_file(filename: str):
    """下载生成的文件"""
    import os
    from fastapi.responses import FileResponse
    from fastapi import HTTPException
    
    file_path = os.path.join("created", filename)
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="文件不存在")
    
    # 根据文件扩展名设置正确的媒体类型
    if filename.endswith('.xlsx'):
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif filename.endswith('.csv'):
        media_type = "text/csv"
    else:
        media_type = "application/octet-stream"
    
    return FileResponse(
        path=file_path,
        media_type=media_type,
        filename=filename
    )

@app.delete("/tables/{table_name}")
async def delete_table(table_name: str):
    """删除指定表"""
    success = db_manager.delete_table(table_name)
    if success:
        return {"status": "success", "message": f"表 {table_name} 已删除"}
    else:
        raise HTTPException(status_code=500, detail=f"删除表 {table_name} 失败")

@app.post("/import-all")
async def import_all_files(background_tasks: BackgroundTasks):
    """导入data目录中的所有Excel文件"""
    data_folder = "data"
    excel_files = glob.glob(os.path.join(data_folder, "*.xlsx"))
    
    if not excel_files:
        raise HTTPException(status_code=404, detail=f"在 {data_folder} 文件夹中未找到任何Excel文件")
    
    # 添加所有文件到后台任务
    for excel_file in excel_files:
        filename = os.path.basename(excel_file)
        # 修复：正确传递参数
        background_tasks.add_task(import_file, filename=filename)
    
    return {
        "total": len(excel_files),
        "files": [os.path.basename(file) for file in excel_files],
        "status": "importing"
    }

@app.delete("/files/{filename}")
async def delete_file(filename: str):
    """删除指定的Excel文件"""
    data_folder = "data"
    file_path = os.path.join(data_folder, filename)
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"文件 {filename} 不存在")
    
    # 删除文件
    try:
        os.remove(file_path)
        return {"filename": filename, "status": "deleted"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除文件失败: {str(e)}")

@app.delete("/files")
async def delete_all_files():
    """删除所有Excel文件"""
    data_folder = "data"
    if not os.path.exists(data_folder):
        raise HTTPException(status_code=404, detail="数据目录不存在")
    
    deleted_files = []
    for filename in os.listdir(data_folder):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(data_folder, filename)
            try:
                os.remove(file_path)
                deleted_files.append(filename)
            except Exception as e:
                logger.error(f"删除文件 {filename} 失败: {e}")
    
    return {
        "message": f"成功删除 {len(deleted_files)} 个文件",
        "deleted_files": deleted_files
    }

@app.delete("/tables")
async def delete_all_tables():
    """删除所有数据库表"""
    try:
        # 获取所有表名
        tables = db_manager.get_tables()
        
        deleted_tables = []
        for table in tables:
            try:
                # 删除表
                db_manager.delete_table(table)
                deleted_tables.append(table)
            except Exception as e:
                print(f"删除表 {table} 失败: {e}")
        
        return {
            "message": f"成功删除 {len(deleted_tables)} 个表",
            "deleted_tables": deleted_tables
        }
    except Exception as e:
        print(f"删除所有表时出错: {e}")
        raise HTTPException(status_code=500, detail="删除所有表失败")

        return {
            "message": f"成功删除 {len(deleted_tables)} 个表",
            "deleted_tables": deleted_tables
        }
    except Exception as e:
        print(f"删除所有表时出错: {e}")
        raise HTTPException(status_code=500, detail="删除所有表失败")

@app.post("/api/generate-daily-hourly-cache")
async def generate_daily_hourly_cache():
    """
    生成所有日期的分时数据缓存
    (修改为：仅执行 init_weather 逻辑，即全量更新日历和天气，并同步缓存中的天气数据)
    """
    from sql_config import SQL_RULES
    from fastapi.concurrency import run_in_threadpool
    import calendar_weather
    
    try:
        # 1. 确定表结构 (保留建表逻辑，防止表不存在导致后续更新缓存失败)
        table_name = "cache_daily_hourly"
        
        # 构建字段列表
        # 基础字段
        columns_def = [
            "`record_date` DATE NOT NULL",
            "`hour` TINYINT NOT NULL",
            "`updated_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP"
        ]
        
        # 从 SQL_RULES 动态生成字段
        # 加上计算字段
        calc_fields = {
            "price_diff": "FLOAT COMMENT '价差'",
            "load_deviation": "FLOAT COMMENT '负荷偏差'",
            "new_energy_forecast": "FLOAT COMMENT '新能源预测总和'"
        }
        
        # 合并所有字段
        all_fields = {}
        
        # 添加规则中的字段
        for key, rule in SQL_RULES.items():
            field_name = key
            # 默认都是 FLOAT，除了日期/字符串类型
            if key in ['date', 'day_type', 'week_day', 'weather', 'wind_direction']:
                col_type = "VARCHAR(50)"
            else:
                col_type = "FLOAT"
            
            all_fields[field_name] = f"`{field_name}` {col_type} COMMENT '{rule.get('name', '')}'"
            
        # 添加计算字段
        for k, v in calc_fields.items():
            all_fields[k] = f"`{k}` {v}"
            
        # 组装 CREATE TABLE 语句
        cols_sql = ",\n".join(list(all_fields.values()) + columns_def)
        
        with db_manager.engine.begin() as conn:
            create_sql = f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                {cols_sql},
                PRIMARY KEY (`record_date`, `hour`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """
            conn.execute(text(create_sql))
            print(f"✅ 缓存表 {table_name} 已就绪")

        # 2. 执行 init_weather 逻辑 (全量更新日历和天气)
        # 参考 init_calendar.py 的范围，或者覆盖较长的时间段
        start_date = datetime.date(2023, 1, 1)
        end_date = datetime.date(2027, 12, 31)
        
        print(f"🚀 开始执行全量天气初始化: {start_date} -> {end_date}")
        
        # 在线程池中运行，避免阻塞主线程
        await run_in_threadpool(calendar_weather.update_calendar, start_date, end_date)
        
        return {"status": "success", "message": f"全量天气及缓存更新完成 ({start_date} 至 {end_date})"}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})

async def calculate_daily_hourly_data(date: str):
    """
    计算指定日期的分时数据（核心逻辑提取）
    返回: List[Dict] (24小时数据)
    """
    from sql_config import SQL_RULES, TABLE_SOURCE_POWER, TABLE_SOURCE_WEATHER
    try:
        target_date = pd.to_datetime(date).date()
        date_str = target_date.strftime("%Y%m%d")
        target_date_str = target_date.strftime("%Y-%m-%d")
        table_name_power = f"power_data_{date_str}"
        table_name_weather = "calendar_weather"
        
        tables = db_manager.get_tables()
        if table_name_power not in tables:
            return None
            
        hourly_data_lists = {h: {} for h in range(24)}
        daily_weather_data = {}
        
        with db_manager.engine.connect() as conn:
            # 1. 查电力数据
            for key, rule in SQL_RULES.items():
                if rule.get("source") == TABLE_SOURCE_POWER:
                    where_clause = rule["where"]
                    sql = text(f"SELECT record_time, value FROM {table_name_power} WHERE {where_clause}")
                    result = conn.execute(sql).fetchall()
                    
                    for row in result:
                        r_time = row[0]
                        val = float(row[1]) if row[1] is not None else 0
                        
                        if hasattr(r_time, 'total_seconds'):
                            hour = int(r_time.total_seconds() // 3600)
                        else:
                            continue
                            
                        if 0 <= hour <= 23:
                            hourly_data_lists[hour].setdefault(key, []).append(val)

        # 2. 查天气数据
        if table_name_weather in tables:
            with db_manager.engine.connect() as conn:
                sql = text(f"SELECT * FROM {table_name_weather} WHERE date = :d")
                row = conn.execute(sql, {"d": target_date_str}).fetchone()
                
                if row:
                    row_dict = dict(row._mapping)
                    weather_json = row_dict.get("weather_json")
                    if isinstance(weather_json, str):
                        try:
                            import json
                            weather_json = json.loads(weather_json)
                        except:
                            weather_json = {}
                    elif weather_json is None:
                        weather_json = {}
                    
                    for key, rule in SQL_RULES.items():
                        if rule.get("source") == TABLE_SOURCE_WEATHER:
                            col = rule.get("column")
                            json_key = rule.get("json_key")
                            
                            val = None
                            if col == "weather_json" and json_key:
                                val = weather_json.get(json_key)
                                if isinstance(val, list) and len(val) == 24:
                                    for h in range(24):
                                        hourly_data_lists[h][key] = val[h]
                                    continue
                            elif col in row_dict:
                                val = row_dict[col]
                            
                            daily_weather_data[key] = val

        # 3. 聚合与计算
        result_list = []
        for h in range(24):
            lists = hourly_data_lists[h]
            row = {"hour": h}
            
            # 均值聚合
            for key, rule in SQL_RULES.items():
                if rule.get("source") == TABLE_SOURCE_POWER:
                    vals = lists.get(key, [])
                    if vals:
                        row[key] = sum(vals) / len(vals)
                elif key in lists:
                    row[key] = lists[key]
            
            # 填充单日天气
            for k, v in daily_weather_data.items():
                row[k] = v
            
            # 计算衍生字段
            if "price_da" in row and "price_rt" in row:
                row["price_diff"] = row["price_da"] - row["price_rt"]
            
            if "load_forecast" in row and "load_actual" in row:
                row["load_deviation"] = row["load_forecast"] - row["load_actual"]
            
            if "new_energy_forecast" not in row:
                pv = row.get("ne_pv_forecast", 0) or 0
                wind = row.get("ne_wind_forecast", 0) or 0
                if pv > 0 or wind > 0:
                    row["new_energy_forecast"] = pv + wind
            
            result_list.append(row)
            
        return result_list
    except Exception as e:
        print(f"Calculation error for {date}: {e}")
        return None

@app.post("/daily-averages")
async def query_daily_averages(
    dates: str = Form(..., description="日期列表，JSON格式，例如: [\"2023-09-18\", \"2023-09-19\"]"),
    data_type_keyword: str = Form("日前节点电价", description="数据类型关键字"),
    station_name: str = Form(None, description="站点名称（可选）"),
    city: str = Form(None, description="城市名称（可选）")
):
    """
    查询多天的均值数据

    参数:
    - dates: 日期列表，JSON格式
    - data_type_keyword: 数据类型关键字
    - station_name: 站点名称（可选）

    返回:
    - 查询结果
    """
    try:
        import json
        date_list = json.loads(dates)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"日期格式错误: {str(e)}")

    # 如果有城市或具体站点名称，查询数据库
    if city and str(city).strip():
        result = importer.query_daily_averages(date_list, data_type_keyword, station_name=None, city=city)
        if result["total"] == 0:
            return {"total": 0, "data": []}
        return result

    if station_name and station_name.strip():
        result = importer.query_daily_averages(date_list, data_type_keyword, station_name)
        if result["total"] == 0:
            return {"total": 0, "data": []}
        return result

    # 没有指定站点名称时，从缓存表查询
    try:
        table_name = "cache_daily_hourly"
        tables = db_manager.get_tables()

        if table_name not in tables:
            return JSONResponse(status_code=404, content={
                "error": "缓存表不存在，请先生成缓存数据",
                "table": table_name
            })

        # 根据 data_type_keyword 确定查询哪个字段
        if "日前" in data_type_keyword:
            price_field = "price_da"
        elif "实时" in data_type_keyword:
            price_field = "price_rt"
        else:
            # 默认查日前
            price_field = "price_da"

        with db_manager.engine.connect() as conn:
            # 构建查询 SQL
            placeholders = ", ".join([f":d{i}" for i in range(len(date_list))])
            params = {f"d{i}": d for i, d in enumerate(date_list)}

            # 查询缓存表
            sql = text(f"""
                SELECT record_date, hour, {price_field} as value
                FROM {table_name}
                WHERE record_date IN ({placeholders})
                  AND {price_field} IS NOT NULL
                ORDER BY record_date DESC, hour
            """)

            result = conn.execute(sql, params).fetchall()

            if not result:
                return {"total": 0, "data": [], "message": "未找到符合条件的数据"}

            # 转换为与原接口兼容的格式
            data = []
            for row in result:
                d = dict(row._mapping)
                hour = d['hour']
                # 将小时数格式化为 "HH:00" 字符串
                record_time_str = f"{hour:02d}:00"

                # 格式化为与原接口一致
                formatted_row = {
                    "record_date": str(d['record_date']),
                    "record_time": record_time_str,  # "HH:00" 格式
                    "value": d['value'],
                    "channel_name": "均值",  # 缓存表中的是均值数据
                    "type": data_type_keyword,
                    "sheet_name": data_type_keyword
                }
                data.append(formatted_row)

            return {
                "total": len(data),
                "data": data,
                "source": "cache"
            }

    except Exception as e:
        import traceback
        traceback.print_exc()
        # 如果缓存表查询失败，回退到原有方法
        print(f"缓存表查询失败，回退到原有方法: {e}")
        result = importer.query_daily_averages(date_list, data_type_keyword, station_name)
        if result["total"] == 0:
            return {"total": 0, "data": []}
        return result

@app.get("/daily-averages/export")
async def export_daily_averages(
    dates: str = Query(..., description="日期列表，JSON格式"),
    data_type_keyword: str = Query("日前节点电价", description="数据类型关键字")
):
    """
    导出多天的均值数据为Excel文件
    
    参数:
    - dates: 日期列表，JSON格式
    - data_type_keyword: 数据类型关键字
    
    返回:
    - Excel文件下载
    """
    try:
        date_list = json.loads(dates)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"日期格式错误: {str(e)}")
    
    # 查询数据
    result = importer.query_daily_averages(date_list, data_type_keyword)
    
    # 生成文件名：多天均值查询_时间戳.xlsx
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"多天均值查询_{timestamp}.xlsx"
    
    if not result["data"]:
        # 如果没有数据，返回空Excel
        df = pd.DataFrame()
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        output.seek(0)
        
        from fastapi.responses import StreamingResponse
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    
    # 转换为DataFrame
    df = pd.DataFrame(result["data"])
    
    # 检查是否包含必要的列
    required_columns = ['channel_name', 'record_date', 'record_time', 'value', 'sheet_name']
    if all(col in df.columns for col in required_columns):
        # 类似preHandle.py的处理方式，生成透视表格式
        try:
            # 提取唯一的sheet_name（假设数据中sheet_name唯一）
            sheet_name = df['sheet_name'].unique()[0] if len(df['sheet_name'].unique()) > 0 else 'Sheet1'
            # 提取唯一的日期（假设数据中日期唯一）
            record_date = df['record_date'].unique()[0] if len(df['record_date'].unique()) > 0 else pd.Timestamp.now().date()
            
            # 格式化日期为YYYY-MM-DD
            if hasattr(record_date, 'strftime'):
                record_date_str = record_date.strftime('%Y-%m-%d')
            else:
                record_date_str = str(record_date)
            
            # 处理文件名特殊字符（避免斜杠、空格等导致保存失败）
            sheet_name_clean = str(sheet_name).replace('/', '_').replace('\\', '_').replace(' ', '')
            
            # 转换record_time为小时（处理各种可能的格式）
            def extract_hour(time_value):
                if pd.isna(time_value):
                    return None
                if isinstance(time_value, str):
                    if ':' in time_value:
                        # 格式如 "01:00", "1:00"
                        return int(time_value.split(':')[0])
                    else:
                        # 可能是数字字符串如 "100" 表示 01:00
                        try:
                            time_int = int(time_value)
                            return time_int // 100
                        except:
                            return None
                elif isinstance(time_value, (int, float)):
                    # 数字格式如 100 表示 01:00
                    return int(time_value) // 100
                else:
                    # timedelta或其他格式
                    try:
                        # 如果是timedelta对象
                        hours = time_value.seconds // 3600
                        return hours
                    except:
                        return None
            
            # 应用小时提取函数
            df['hour'] = df['record_time'].apply(extract_hour)
            
            # 删除hour为NaN的行
            df = df.dropna(subset=['hour'])
            
            # 生成电站级透视表
            if len(df) > 0:
                pivot_df = pd.pivot_table(
                    df,
                    index=['channel_name', 'record_date'],
                    columns='hour',
                    values='value',
                    aggfunc='mean'
                )
                
                # 重新索引确保有24小时列，并正确格式化列名
                pivot_df = pivot_df.reindex(columns=range(24), fill_value=np.nan)
                # 确保列名格式为 HH:00
                pivot_df.columns = [f'{int(h):02d}:00' for h in pivot_df.columns]
                pivot_df = pivot_df.reset_index()
                
                # 修改前两列名称
                pivot_df = pivot_df.rename(columns={
                    'channel_name': '节点名称',
                    'record_date': '日期'
                })
                
                # 插入单位列
                pivot_df.insert(
                    loc=2,
                    column='单位',
                    value='电价(元/MWh)'
                )
                
                # 添加发电侧全省统一均价行
                hour_columns = [f'{h:02d}:00' for h in range(24)]
                # 确保所有小时列都存在
                for col in hour_columns:
                    if col not in pivot_df.columns:
                        pivot_df[col] = np.nan
                
                # 在计算平均值前，确保所有列为数值类型
                for col in hour_columns:
                    pivot_df[col] = pd.to_numeric(pivot_df[col], errors='coerce')
                
                # 计算全省统一均价行
                province_avg = {}
                for col in hour_columns:
                    if col in pivot_df.columns:
                        province_avg[col] = pivot_df[col].mean(skipna=True)
                              
                final_df = pivot_df
            else:
                # 如果处理后没有数据，创建空的DataFrame
                columns = ['节点名称', '日期', '单位'] + [f'{h:02d}:00' for h in range(24)]
                final_df = pd.DataFrame(columns=columns)
            
            # 直接返回Excel文件流
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name=sheet_name_clean[:31])
            output.seek(0)
            
            import urllib.parse
            encoded_filename = urllib.parse.quote(filename)
            from fastapi.responses import StreamingResponse
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
                }
            )
        except Exception as e:
            print(f"处理透视表格式时出错: {e}")
            import traceback
            traceback.print_exc()
    
    # 如果不包含必要列或处理透视表失败，使用原始导出方式
    # 直接返回Excel文件流
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='多天均值数据')
    output.seek(0)
    
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )

@app.post("/daily-averages/export-from-result")
async def export_daily_averages_from_result(
    query_result: str = Form(..., description="查询结果数据"),
    data_type_keyword: str = Form("日前节点电价", description="数据类型关键字")
):
    """
    根据当前查询结果导出多天的均值数据为Excel文件
    
    参数:
    - query_result: 当前查询结果，JSON格式
    - data_type_keyword: 数据类型关键字
    
    返回:
    - Excel文件下载
    """
    try:
        import json
        query_result_data = json.loads(query_result)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"查询结果格式错误: {str(e)}")
    
    # 生成文件名：多天均值查询_时间戳.xlsx
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"多天均值查询_{timestamp}.xlsx"
    
    if not query_result_data:
        # 如果没有数据，返回空Excel
        df = pd.DataFrame()
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        output.seek(0)
        
        from fastapi.responses import StreamingResponse
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    
    # 转换为DataFrame
    df = pd.DataFrame(query_result_data)
    
    # 检查是否包含必要的列
    required_columns = ['channel_name', 'record_date', 'record_time', 'value', 'sheet_name']
    if all(col in df.columns for col in required_columns):
        # 类似preHandle.py的处理方式，生成透视表格式
        try:
            # 提取唯一的sheet_name（假设数据中sheet_name唯一）
            sheet_name = df['sheet_name'].unique()[0] if len(df['sheet_name'].unique()) > 0 else 'Sheet1'
            # 提取唯一的日期（假设数据中日期唯一）
            record_date = df['record_date'].unique()[0] if len(df['record_date'].unique()) > 0 else pd.Timestamp.now().date()
            
            # 格式化日期为YYYY-MM-DD
            if hasattr(record_date, 'strftime'):
                record_date_str = record_date.strftime('%Y-%m-%d')
            else:
                record_date_str = str(record_date)
            
            # 处理文件名特殊字符（避免斜杠、空格等导致保存失败）
            sheet_name_clean = str(sheet_name).replace('/', '_').replace('\\', '_').replace(' ', '')
            
            # 转换record_time为小时（处理各种可能的格式）
            # 转换record_time为小时
            def extract_hour(time_value):
                if pd.isna(time_value):
                    return None
                
                try:
                    # 1. 优先处理整数/浮点数
                    if isinstance(time_value, (int, float, np.number)):
                        val = int(time_value)
                        
                        # 【核心修复逻辑】
                        # 如果数值很大（超过2400），说明肯定是秒数，不是HHMM
                        # 例如 3600(秒) / 3600 = 1点
                        if val >= 3600: 
                             return val // 3600
                        
                        # 如果数值在 0-23 之间，直接是小时
                        if 0 <= val < 24:
                            return val
                            
                        # 如果是 HHMM 格式 (例如 100 代表 01:00, 2300 代表 23:00)
                        if 100 <= val <= 2400:
                            return val // 100
                            
                        # 兜底：如果是 0，既可能是0点也可能是0秒，返回0
                        if val == 0:
                            return 0

                    # 2. 处理字符串
                    time_str = str(time_value).strip()
                    if ':' in time_str:
                        return int(time_str.split(':')[0])
                    
                    # 3. 处理 Timedelta 对象
                    if hasattr(time_value, 'total_seconds'):
                        return int(time_value.total_seconds() // 3600)
                    if hasattr(time_value, 'seconds'):
                        return int(time_value.seconds // 3600)

                    # 再次尝试转数字处理（防止字符串类型的数字 "3600"）
                    try:
                        val = int(float(time_str))
                        if val >= 3600: return val // 3600
                        if val < 24: return val
                        return val // 100
                    except:
                        pass

                    return None
                except Exception as e:
                    return None
            
            # 应用小时提取函数
            df['hour'] = df['record_time'].apply(extract_hour)
            print("转换后的前10行数据:")
            print(df[['record_time', 'hour']].head(10))
            print("Hour列的唯一值:", df['hour'].unique())
            # 删除hour为NaN的行
            df = df.dropna(subset=['hour'])
            
            # 生成电站级透视表
            if len(df) > 0:
                pivot_df = pd.pivot_table(
                    df,
                    index=['channel_name', 'record_date'],
                    columns='hour',
                    values='value',
                    aggfunc='mean'
                )
                
                # 重新索引确保有24小时列，并正确格式化列名
                pivot_df = pivot_df.reindex(columns=range(24), fill_value=np.nan)
                # 确保列名格式为 HH:00
                pivot_df.columns = [f'{int(h):02d}:00' for h in pivot_df.columns]
                pivot_df = pivot_df.reset_index()
                
                # 修改前两列名称
                pivot_df = pivot_df.rename(columns={
                    'channel_name': '节点名称',
                    'record_date': '日期'
                })
                
                # 插入单位列
                pivot_df.insert(
                    loc=2,
                    column='单位',
                    value='电价(元/MWh)'
                )
                
                # 添加发电侧全省统一均价行
                hour_columns = [f'{h:02d}:00' for h in range(24)]
                # 确保所有小时列都存在
                for col in hour_columns:
                    if col not in pivot_df.columns:
                        pivot_df[col] = np.nan
                
                # 在计算平均值前，确保所有列为数值类型
                for col in hour_columns:
                    pivot_df[col] = pd.to_numeric(pivot_df[col], errors='coerce')
                
                # 计算全省统一均价行
                province_avg = {}
                for col in hour_columns:
                    if col in pivot_df.columns:
                        province_avg[col] = pivot_df[col].mean(skipna=True)
                              
                final_df = pivot_df
            else:
                # 如果处理后没有数据，创建空的DataFrame
                columns = ['节点名称', '日期', '单位'] + [f'{h:02d}:00' for h in range(24)]
                final_df = pd.DataFrame(columns=columns)
            
            # 直接返回Excel文件流
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name=sheet_name_clean[:31])
            output.seek(0)
            
            import urllib.parse
            encoded_filename = urllib.parse.quote(filename)
            from fastapi.responses import StreamingResponse
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
                }
            )
        except Exception as e:
            print(f"处理透视表格式时出错: {e}")
            import traceback
            traceback.print_exc()
    
    # 如果不包含必要列或处理透视表失败，使用原始导出方式
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='多天均值数据')
    output.seek(0)
    
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )

@app.post("/price-difference")
async def query_price_difference(
    dates: str = Form(..., description="日期列表，JSON格式，例如: [\"2023-09-18\", \"2023-09-19\"]"),
    region: str = Form("", description="地区前缀，如'云南_'，默认为空"),
    station_name: str = Form(None, description="站点名称（可选）"),
    city: str = Form(None, description="城市名称（可选）")
):
    """
    查询价差数据（日前节点电价 - 实时节点电价）

    参数:
    - dates: 日期列表，JSON格式
    - region: 地区前缀，如"云南_"，默认为空
    - station_name: 站点名称（可选）

    返回:
    - 价差查询结果
    """
    try:
        import json
        date_list = json.loads(dates)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"日期格式错误: {str(e)}")

    # 如果有城市或具体站点名称，查询数据库
    if city and str(city).strip():
        result = importer.query_price_difference(date_list, region, station_name=None, city=city)
        return result

    if station_name and station_name.strip():
        result = importer.query_price_difference(date_list, region, station_name)
        return result

    # 没有指定站点名称时，从缓存表查询
    try:
        table_name = "cache_daily_hourly"
        tables = db_manager.get_tables()

        if table_name not in tables:
            return JSONResponse(status_code=404, content={
                "error": "缓存表不存在，请先生成缓存数据",
                "table": table_name
            })

        with db_manager.engine.connect() as conn:
            # 构建查询 SQL
            placeholders = ", ".join([f":d{i}" for i in range(len(date_list))])
            params = {f"d{i}": d for i, d in enumerate(date_list)}

            # 查询缓存表中的价差数据
            sql = text(f"""
                SELECT record_date, hour, price_diff as value, price_da, price_rt
                FROM {table_name}
                WHERE record_date IN ({placeholders})
                  AND price_diff IS NOT NULL
                ORDER BY record_date DESC, hour
            """)

            result = conn.execute(sql, params).fetchall()

            if not result:
                return {
                    "total": 0,
                    "data": [],
                    "message": "未找到符合条件的价差数据"
                }

            # 转换为与原接口兼容的格式
            data = []
            for row in result:
                d = dict(row._mapping)
                hour = d['hour']
                # 将小时数格式化为 "HH:00" 字符串
                record_time_str = f"{hour:02d}:00"

                # 格式化为与原接口一致
                formatted_row = {
                    "record_date": str(d['record_date']),
                    "record_time": record_time_str,  # "HH:00" 格式
                    "value": d['value'],  # 价差
                    "price_da": d.get('price_da'),
                    "price_rt": d.get('price_rt'),
                    "channel_name": "均值",  # 缓存表中的是均值数据
                    "type": "价差",
                    "sheet_name": "价差"
                }
                data.append(formatted_row)

            return {
                "total": len(data),
                "data": data,
                "source": "cache"
            }

    except Exception as e:
        import traceback
        traceback.print_exc()
        # 如果缓存表查询失败，回退到原有方法
        print(f"缓存表查询失败，回退到原有方法: {e}")
        result = importer.query_price_difference(date_list, region, station_name)
        return result

@app.post("/price-difference/export-from-result")
async def export_price_difference_from_result(
    query_result: str = Form(..., description="查询结果数据"),
    region: str = Form("", description="地区前缀")
):
    """
    根据当前查询结果导出价差数据为Excel文件
    
    参数:
    - query_result: 当前查询结果，JSON格式
    - region: 地区前缀
    
    返回:
    - Excel文件下载
    """
    try:
        import json
        query_result_data = json.loads(query_result)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"查询结果格式错误: {str(e)}")
    
    # 生成文件名：价差查询_时间戳.xlsx
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"价差查询_{timestamp}.xlsx"
    
    if not query_result_data:
        # 如果没有数据，返回空Excel
        df = pd.DataFrame()
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        output.seek(0)
        
        from fastapi.responses import StreamingResponse
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    
    # 转换为DataFrame
    df = pd.DataFrame(query_result_data)
    
    # 检查是否包含必要的列
    required_columns = ['channel_name', 'record_date', 'record_time', 'value', 'sheet_name']
    if all(col in df.columns for col in required_columns):
        # 类似preHandle.py的处理方式，生成透视表格式
        try:
            # 提取唯一的sheet_name
            sheet_name = df['sheet_name'].unique()[0] if len(df['sheet_name'].unique()) > 0 else 'Sheet1'
            record_date = df['record_date'].unique()[0] if len(df['record_date'].unique()) > 0 else pd.Timestamp.now().date()
            
            # 格式化日期为YYYY-MM-DD
            if hasattr(record_date, 'strftime'):
                record_date_str = record_date.strftime('%Y-%m-%d')
            else:
                record_date_str = str(record_date)
            
            # 处理文件名特殊字符
            sheet_name_clean = str(sheet_name).replace('/', '_').replace('\\', '_').replace(' ', '')
            
            # 转换record_time为小时
            def extract_hour(time_value):
                if pd.isna(time_value):
                    return None
                try:
                    if isinstance(time_value, (int, float, np.number)):
                        val = int(time_value)
                        if val >= 3600:
                            return val // 3600
                        if 0 <= val < 24:
                            return val
                        if 100 <= val <= 2400:
                            return val // 100
                        if val == 0:
                            return 0
                    time_str = str(time_value).strip()
                    if ':' in time_str:
                        return int(time_str.split(':')[0])
                    if hasattr(time_value, 'total_seconds'):
                        return int(time_value.total_seconds() // 3600)
                    if hasattr(time_value, 'seconds'):
                        return int(time_value.seconds // 3600)
                    try:
                        val = int(float(time_str))
                        if val >= 3600:
                            return val // 3600
                        if val < 24:
                            return val
                        return val // 100
                    except:
                        pass
                    return None
                except Exception as e:
                    return None
            
            # 应用小时提取函数
            df['hour'] = df['record_time'].apply(extract_hour)
            df = df.dropna(subset=['hour'])
            
            # 生成透视表
            if len(df) > 0:
                pivot_df = pd.pivot_table(
                    df,
                    index=['channel_name', 'record_date'],
                    columns='hour',
                    values='value',
                    aggfunc='mean'
                )
                
                # 重新索引确保有24小时列
                pivot_df = pivot_df.reindex(columns=range(24), fill_value=np.nan)
                pivot_df.columns = [f'{int(h):02d}:00' for h in pivot_df.columns]
                pivot_df = pivot_df.reset_index()
                
                # 修改列名称
                pivot_df = pivot_df.rename(columns={
                    'channel_name': '节点名称',
                    'record_date': '日期'
                })
                
                # 插入单位列
                pivot_df.insert(
                    loc=2,
                    column='单位',
                    value='价差(元/MWh)'
                )
                
                # 确保所有小时列都存在
                hour_columns = [f'{h:02d}:00' for h in range(24)]
                for col in hour_columns:
                    if col not in pivot_df.columns:
                        pivot_df[col] = np.nan
                
                # 确保所有列为数值类型
                for col in hour_columns:
                    pivot_df[col] = pd.to_numeric(pivot_df[col], errors='coerce')
                
                final_df = pivot_df
            else:
                columns = ['节点名称', '日期', '单位'] + [f'{h:02d}:00' for h in range(24)]
                final_df = pd.DataFrame(columns=columns)
            
            # 返回Excel文件流
            output = BytesIO()
            # from openpyxl.chart import BarChart, Reference, Series
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name=sheet_name_clean[:31])
                
                # 获取工作表
                worksheet = writer.sheets[sheet_name_clean[:31]]
                
                # 应用条件格式：大于0显示绿色渐变，小于0显示红色渐变
                hour_columns = [f'{h:02d}:00' for h in range(24)]
                
                # 找到所有数值中的最大绝对值，用于确定颜色深度
                max_abs_value = 0
                for col in final_df.columns:
                    if col in hour_columns:
                        max_abs_value = max(max_abs_value, final_df[col].abs().max())
                
                # 如果最大绝对值为0，则设为1避免除零错误
                if max_abs_value == 0:
                    max_abs_value = 1
                
                # 定义颜色填充函数
                def get_fill_color(value):
                    if pd.isna(value):
                        return None
                    
                    # 计算颜色强度，基于绝对值比例
                    intensity = abs(value) / max_abs_value
                    
                    # 确保最小亮度，避免颜色过深
                    min_brightness = 150  # 最亮为255
                    brightness_range = 255 - min_brightness
                    brightness = int(min_brightness + (1 - intensity) * brightness_range)
                    
                    if value > 0:
                        # 正数：绿色系，强度越高颜色越深
                        red = brightness
                        green = 255
                        blue = brightness
                    elif value < 0:
                        # 负数：红色系，强度越高颜色越深
                        red = 255
                        green = brightness
                        blue = brightness
                    else:
                        # 零值：白色
                        return None
                    
                    # 转换为十六进制颜色代码
                    color_code = f"{red:02X}{green:02X}{blue:02X}"
                    return PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
                
                # 找到小时列的列索引并应用条件格式
                for col_idx, col in enumerate(final_df.columns, start=1):
                    if col in hour_columns:
                        # 对每个小时列应用条件格式
                        for row_idx in range(2, len(final_df) + 2):  # 从第2行开始（第1行是表头）
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if cell.value is not None:
                                try:
                                    value = float(cell.value)
                                    fill = get_fill_color(value)
                                    if fill:
                                        cell.fill = fill
                                except (ValueError, TypeError):
                                    pass

            output.seek(0)
            
            import urllib.parse
            encoded_filename = urllib.parse.quote(filename)
            from fastapi.responses import StreamingResponse
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
                }
            )
        except Exception as e:
            print(f"处理透视表格式时出错: {e}")
            import traceback
            traceback.print_exc()
    
    # 如果不包含必要列或处理透视表失败，使用原始导出方式
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='价差数据')
    output.seek(0)
    
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )

@app.get("/daily_hourly", response_class=HTMLResponse)
async def daily_hourly_page(request: Request):
    """返回24小时数据展示页面"""
    return templates.TemplateResponse("daily_hourly.html", {"request": request})

@app.get("/similar_day", response_class=HTMLResponse)
async def similar_day_page(request: Request):
    """返回类比日匹配页面"""
    return templates.TemplateResponse("similar_day.html", {"request": request})

@app.get("/multi_day_compare", response_class=HTMLResponse)
async def multi_day_compare_page(request: Request):
    """返回多日对比看板页面"""
    return templates.TemplateResponse("multi_day_compare.html", {"request": request})

@app.get("/strategy_quote", response_class=HTMLResponse)
async def strategy_quote_page(request: Request, platform: Optional[str] = None):
    """报价/申报页面（策略系数 + 申报电量计算）"""
    p = _normalize_platform(platform)
    return templates.TemplateResponse("strategy_quote.html", {"request": request, "default_platform": p, "default_platform_label": _STRATEGY_PLATFORM_LABELS.get(p)})

@app.get("/strategy_review", response_class=HTMLResponse)
async def strategy_review_page(request: Request, platform: Optional[str] = None):
    """复盘页面（月度胜率/收益/预测偏差）"""
    p = _normalize_platform(platform)
    return templates.TemplateResponse("strategy_review.html", {"request": request, "default_platform": p, "default_platform_label": _STRATEGY_PLATFORM_LABELS.get(p)})


@app.get("/strategy_review_diag", response_class=HTMLResponse)
async def strategy_review_diag_page(request: Request, date: Optional[str] = None, platform: Optional[str] = None):
    """复盘诊断页面（按日期对比 cache_daily_hourly 与当月基线，定位异常小时）。"""
    p = _normalize_platform(platform)
    return templates.TemplateResponse("strategy_review_diag.html", {"request": request, "date": date, "default_platform": p, "default_platform_label": _STRATEGY_PLATFORM_LABELS.get(p)})

@app.get("/api/cache-available-dates")
async def cache_available_dates(
    limit: int = Query(400, ge=1, le=2000),
    require_load: bool = Query(False, description="仅返回负荷预测(load_forecast)有值的日期"),
):
    """返回缓存表 cache_daily_hourly 中可用的日期列表（倒序）"""
    try:
        table_name = "cache_daily_hourly"
        tables = db_manager.get_tables()
        if table_name not in tables:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": "缓存表不存在，请先生成缓存数据", "table": table_name},
            )

        limit_int = max(1, min(int(limit), 2000))
        with db_manager.engine.connect() as conn:
            if require_load:
                sql = text(
                    f"""
                    SELECT record_date
                    FROM {table_name}
                    GROUP BY record_date
                    HAVING SUM(CASE WHEN load_forecast IS NOT NULL AND load_forecast <> 0 THEN 1 ELSE 0 END) > 0
                    ORDER BY record_date DESC
                    LIMIT {limit_int}
                    """
                )
            else:
                sql = text(
                    f"""
                    SELECT DISTINCT record_date
                    FROM {table_name}
                    ORDER BY record_date DESC
                    LIMIT {limit_int}
                    """
                )
            result = conn.execute(sql).fetchall()

        dates = []
        for row in result:
            d = row[0]
            if hasattr(d, "strftime"):
                dates.append(d.strftime("%Y-%m-%d"))
            else:
                dates.append(str(d))

        return {
            "status": "success",
            "dates": dates,
            "total": len(dates),
            "table": table_name,
            "require_load": require_load,
        }
    except Exception as e:
        import traceback

        traceback.print_exc()
        return JSONResponse(status_code=500, content={"status": "error", "error": str(e)})

class MultiDayCompareDatesRequest(BaseModel):
    dates: list[str] = Field(..., description="日期列表(YYYY-MM-DD)，建议按倒序传入：目标日 + 往前N天")

@app.post("/api/multi-day-compare-data-by-dates")
async def get_multi_day_compare_data_by_dates(request: MultiDayCompareDatesRequest):
    """按指定日期列表获取分时对比数据（来自缓存表 cache_daily_hourly）"""
    try:
        table_name = "cache_daily_hourly"
        tables = db_manager.get_tables()
        if table_name not in tables:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": "缓存表不存在，请先生成缓存数据", "table": table_name},
            )

        # 验证日期格式 + 去重
        valid_dates = []
        seen = set()
        for d in request.dates or []:
            try:
                ds = pd.to_datetime(d).date().strftime("%Y-%m-%d")
            except Exception:
                continue
            if ds in seen:
                continue
            seen.add(ds)
            valid_dates.append(ds)

        if not valid_dates:
            return JSONResponse(status_code=400, content={"status": "error", "message": "日期列表为空或格式错误"})

        required_cols = [
            "record_date",
            "hour",
            "temperature",
            "load_forecast",
            "class_b_forecast",
            "spot_ne_d_forecast",
            "ne_pv_forecast",
            "ne_wind_forecast",
            "day_type",
            "weather",
        ]

        placeholders = ", ".join([f":d{i}" for i in range(len(valid_dates))])
        params = {f"d{i}": d for i, d in enumerate(valid_dates)}

        data = []
        with db_manager.engine.connect() as conn:
            try:
                sql = text(
                    f"""
                    SELECT {', '.join(required_cols)}
                    FROM {table_name}
                    WHERE record_date IN ({placeholders})
                    ORDER BY record_date DESC, hour ASC
                    """
                )
                result = conn.execute(sql, params).fetchall()
                for row in result:
                    d = dict(row._mapping)
                    if hasattr(d.get("record_date"), "strftime"):
                        d["record_date"] = d["record_date"].strftime("%Y-%m-%d")
                    data.append(d)
            except Exception:
                sql = text(
                    f"""
                    SELECT *
                    FROM {table_name}
                    WHERE record_date IN ({placeholders})
                    ORDER BY record_date DESC, hour ASC
                    """
                )
                result = conn.execute(sql, params).fetchall()
                for row in result:
                    r = dict(row._mapping)
                    out = {k: r.get(k) for k in required_cols if k in r}
                    for k in required_cols:
                        out.setdefault(k, None)
                    if hasattr(out.get("record_date"), "strftime"):
                        out["record_date"] = out["record_date"].strftime("%Y-%m-%d")
                    data.append(out)

        # 计算 B 类占比（按小时）
        for d in data:
            lf = d.get("load_forecast") or 0.0
            b = d.get("class_b_forecast") or 0.0
            d["b_ratio"] = (float(b) / float(lf)) if lf not in (0, 0.0, None) else None

        unique_dates = sorted({d["record_date"] for d in data if d.get("record_date")})

        segments = [
            {"key": "morning_peak", "name": "早高峰", "start": 7, "end": 10},
            {"key": "pv_midday", "name": "中午光伏大发", "start": 11, "end": 14},
            {"key": "evening_peak", "name": "晚高峰", "start": 18, "end": 21},
        ]

        return {
            "status": "success",
            "table": table_name,
            "total_records": len(data),
            "total_dates": len(unique_dates),
            "dates": unique_dates,
            "segments": segments,
            "data": data,
            "requested_dates": valid_dates,
        }
    except Exception as e:
        import traceback

        traceback.print_exc()
        return JSONResponse(status_code=500, content={"status": "error", "error": str(e)})

@app.get("/api/multi-day-compare-data")
async def get_multi_day_compare_data(start: str = Query(...), end: str = Query(...)):
    """
    获取指定日期范围内的分时对比数据（来自缓存表 cache_daily_hourly）

    返回字段（尽量齐全，缺失则为 null）：
    - temperature, load_forecast, class_b_forecast, b_ratio, spot_ne_d_forecast, ne_pv_forecast, ne_wind_forecast
    - day_type, weather
    """
    try:
        table_name = "cache_daily_hourly"
        tables = db_manager.get_tables()
        if table_name not in tables:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": "缓存表不存在，请先生成缓存数据", "table": table_name},
            )

        start_dt = pd.to_datetime(start).date()
        end_dt = pd.to_datetime(end).date()
        if start_dt > end_dt:
            start_dt, end_dt = end_dt, start_dt

        start_str = start_dt.strftime("%Y-%m-%d")
        end_str = end_dt.strftime("%Y-%m-%d")

        required_cols = [
            "record_date",
            "hour",
            "temperature",
            "load_forecast",
            "class_b_forecast",
            "spot_ne_d_forecast",
            "ne_pv_forecast",
            "ne_wind_forecast",
            "day_type",
            "weather",
        ]

        data = []
        with db_manager.engine.connect() as conn:
            try:
                sql = text(
                    f"""
                    SELECT {', '.join(required_cols)}
                    FROM {table_name}
                    WHERE record_date BETWEEN :start AND :end
                    ORDER BY record_date DESC, hour ASC
                    """
                )
                result = conn.execute(sql, {"start": start_str, "end": end_str}).fetchall()
                for row in result:
                    d = dict(row._mapping)
                    if hasattr(d.get("record_date"), "strftime"):
                        d["record_date"] = d["record_date"].strftime("%Y-%m-%d")
                    data.append(d)
            except Exception:
                # 兼容旧缓存表字段不全：退化为 SELECT *，再挑选需要的字段
                sql = text(
                    f"""
                    SELECT *
                    FROM {table_name}
                    WHERE record_date BETWEEN :start AND :end
                    ORDER BY record_date DESC, hour ASC
                    """
                )
                result = conn.execute(sql, {"start": start_str, "end": end_str}).fetchall()
                for row in result:
                    r = dict(row._mapping)
                    out = {k: r.get(k) for k in required_cols if k in r}
                    # 缺失字段补齐
                    for k in required_cols:
                        out.setdefault(k, None)
                    if hasattr(out.get("record_date"), "strftime"):
                        out["record_date"] = out["record_date"].strftime("%Y-%m-%d")
                    data.append(out)

        # 计算 B 类占比（按小时）
        for d in data:
            lf = d.get("load_forecast") or 0.0
            b = d.get("class_b_forecast") or 0.0
            d["b_ratio"] = (float(b) / float(lf)) if lf not in (0, 0.0, None) else None

        unique_dates = sorted({d["record_date"] for d in data if d.get("record_date")})

        segments = [
            {"key": "morning_peak", "name": "早高峰", "start": 7, "end": 10},
            {"key": "pv_midday", "name": "中午光伏大发", "start": 11, "end": 14},
            {"key": "evening_peak", "name": "晚高峰", "start": 18, "end": 21},
        ]

        return {
            "status": "success",
            "table": table_name,
            "start": start_str,
            "end": end_str,
            "total_records": len(data),
            "total_dates": len(unique_dates),
            "dates": unique_dates,
            "segments": segments,
            "data": data,
        }
    except Exception as e:
        import traceback

        traceback.print_exc()
        return JSONResponse(status_code=500, content={"status": "error", "error": str(e)})

@app.get("/api/daily-hourly-data")
async def get_daily_hourly_data(date: str):
    """获取指定日期的24小时数据 (优先查缓存)"""
    try:
        # 1. 尝试从缓存表查询
        table_name = "cache_daily_hourly"
        target_date = pd.to_datetime(date).date()
        date_str = target_date.strftime("%Y-%m-%d")
        
        tables = db_manager.get_tables()
        if table_name in tables:
            with db_manager.engine.connect() as conn:
                # 获取所有列
                sql = text(f"SELECT * FROM {table_name} WHERE record_date = :d ORDER BY hour ASC")
                result = conn.execute(sql, {"d": date_str}).fetchall()
                
                if result:
                    # 转换回字典列表
                    data_list = []
                    for row in result:
                        d = dict(row._mapping)
                        # 处理日期对象转字符串
                        if 'record_date' in d:
                            d['record_date'] = str(d['record_date'])
                        if 'updated_at' in d:
                            d['updated_at'] = str(d['updated_at'])
                        data_list.append(d)
                    return {"status": "success", "data": data_list, "source": "cache"}

        # 2. 如果缓存没命中，实时计算
        print(f"Cache miss for {date_str}, calculating...")
        data = await calculate_daily_hourly_data(date_str)
        
        if data:
            # 3. 异步写入缓存 (简单起见，这里同步写入，或留给下次批量生成)
            # 为了保证下次查询快，最好这里就写入。
            # 但考虑到表可能还没建，或者 calculate_daily_hourly_data 是独立的
            # 我们可以在 calculate_daily_hourly_data 外部再调一次生成逻辑，或者暂时只返回实时数据
            # 既然用户专门要了缓存表，我们应该尽力去存。
            
            # 尝试自动建表并存入? 
            # 简单起见，直接返回实时计算结果，并建议用户点击"生成缓存"
            # 或者，我们可以调用 generate_daily_hourly_cache 的一部分逻辑来存单日
            # 这里我们选择直接返回实时数据，但在前端提示。
            return {"status": "success", "data": data, "source": "realtime"}
        else:
             return {"status": "error", "message": f"未找到 {date} 的电力数据"}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/generate-price-cache")
async def generate_price_cache(request: Request):
    """
    生成节点电价映射缓存表 -> 合并入 cache_daily_hourly
    """
    try:
        # 1. 获取所有有数据的日期
        all_tables = db_manager.get_tables()
        power_tables = [t for t in all_tables if t.startswith('power_data_')]
        
        dates_to_process = []
        for t in power_tables:
            try:
                d_str = t.replace('power_data_', '')
                dates_to_process.append(d_str) # YYYYMMDD
            except:
                pass
        
        dates_to_process.sort()
        total_days = len(dates_to_process)
        print(f"待处理日期: {total_days} 天")
        
        processed_count = 0
        inserted_count = 0
        
        for date_str in dates_to_process:
            # YYYYMMDD -> YYYY-MM-DD
            target_date_str = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
            
            try:
                count = update_price_cache_for_date(target_date_str)
                inserted_count += count
            except Exception as e:
                print(f"Error processing {date_str}: {e}")
                import traceback
                traceback.print_exc()
                continue
            
            processed_count += 1
            if processed_count % 10 == 0:
                print(f"Price Cache: Processed {processed_count}/{total_days} days")

        return {
            "status": "success", 
            "processed_days": processed_count, 
            "inserted_records": inserted_count,
            "table": "cache_daily_hourly"
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})

def update_price_cache_for_date(target_date_str: str, only_weather: bool = False) -> int:
    """
    更新指定日期的电价缓存 (供 generate_price_cache 和 import_file 调用)
    返回插入/更新的记录数 (最大24)
    
    Args:
        target_date_str: 目标日期 YYYY-MM-DD
        only_weather: 是否只更新天气数据 (保留原有电力数据)
    """
    from sql_config import SQL_RULES, TABLE_SOURCE_POWER, TABLE_SOURCE_WEATHER
    
    table_name = "cache_daily_hourly"

    # 1. 确保表存在
    # (为了性能，这里可以假设表已存在，或者每次都检查，对于单次导入检查一下无妨)
    # 构建字段列表
    columns_def = [
        "`record_date` DATE NOT NULL",
        "`hour` TINYINT NOT NULL",
        "`updated_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP"
    ]
    
    calc_fields = {
        "price_diff": "FLOAT COMMENT '价差'",
        "load_deviation": "FLOAT COMMENT '负荷偏差'",
        "new_energy_forecast": "FLOAT COMMENT '新能源预测总和'"
    }
    
    all_fields = {}
    for key, rule in SQL_RULES.items():
        field_name = key
        if key in ['date', 'day_type', 'week_day', 'weather', 'wind_direction']:
            col_type = "VARCHAR(50)"
        else:
            col_type = "FLOAT"
        all_fields[field_name] = f"`{field_name}` {col_type} COMMENT '{rule.get('name', '')}'"
        
    for k, v in calc_fields.items():
        all_fields[k] = f"`{k}` {v}"
        
    cols_sql = ",\n".join(list(all_fields.values()) + columns_def)
    
    with db_manager.engine.begin() as conn:
        create_sql = f"""
        CREATE TABLE IF NOT EXISTS {table_name} (
            {cols_sql},
            PRIMARY KEY (`record_date`, `hour`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        conn.execute(text(create_sql))
    
    # 2. 获取数据 (使用 sql_config 中的规则动态查询)
    # from sql_config import SQL_RULES, TABLE_SOURCE_POWER, TABLE_SOURCE_WEATHER (Moved to top)
    
    # 2.1 构造小时数据映射 {hour: {field_name: [val1, val2]}}
    # ...
    
    hourly_map = {h: {} for h in range(24)}
    
    # 初始化字段列表 (用于 hourly_map)
    # 包括 price_da, price_rt 以及 SQL_RULES 中定义的所有 POWER 数据
    field_keys = ['price_da', 'price_rt']
    for k, v in SQL_RULES.items():
        if v.get('source') == TABLE_SOURCE_POWER and k not in ['price_da', 'price_rt']:
            field_keys.append(k)
            
    # 如果 only_weather=True，则不需要初始化这些字段的列表，也不需要查询电力数据
    if not only_weather:
        for h in range(24):
            for k in field_keys:
                hourly_map[h][k] = []

        # 2.2 获取日前/实时电价 (保留之前的特定逻辑：区域过滤)
        da_result = importer.query_daily_averages([target_date_str], "日前节点电价")
        da_data = da_result.get("data", [])
        
        rt_result = importer.query_daily_averages([target_date_str], "实时节点电价")
        rt_data = rt_result.get("data", [])
        
        def filter_and_process_price(data_list, type_key):
            filtered = [item for item in data_list if "云南" not in str(item.get('type', ''))]
            has_guangdong = any("广东" in str(item.get('type', '')) for item in filtered)
            if has_guangdong:
                filtered = [item for item in filtered if "广东" in str(item.get('type', ''))]
            
            for item in filtered:
                rt_val = item['record_time']
                norm_time = normalize_record_time(rt_val, target_date_str)
                if norm_time is None:
                    continue
                
                hour = norm_time.hour
                if 0 <= hour <= 23:
                    val = float(item['value']) if item['value'] is not None else 0
                    hourly_map[hour][type_key].append(val)

        filter_and_process_price(da_data, 'price_da')
        filter_and_process_price(rt_data, 'price_rt')

        # 2.3 获取 SQL_RULES 中定义的其他电力数据
        # 构造表名
        d_obj = datetime.datetime.strptime(target_date_str, "%Y-%m-%d")
        table_name_power = f"power_data_{d_obj.strftime('%Y%m%d')}"
        
        # 检查表是否存在
        if table_name_power in db_manager.get_tables():
            with db_manager.engine.connect() as conn:
                for key, rule in SQL_RULES.items():
                    if rule.get('source') == TABLE_SOURCE_POWER and key not in ['price_da', 'price_rt']:
                        where_clause = rule.get('where')
                        if not where_clause:
                            continue
                            
                        try:
                            sql = text(f"SELECT record_time, value FROM {table_name_power} WHERE {where_clause}")
                            result = conn.execute(sql).fetchall()
                            
                            for row in result:
                                rt_val = row[0]
                                norm_time = normalize_record_time(rt_val, target_date_str)
                                if norm_time is None:
                                    continue
                                
                                hour = norm_time.hour
                                if 0 <= hour <= 23:
                                    val = float(row[1]) if row[1] is not None else 0
                                    hourly_map[hour][key].append(val)
                        except Exception as e:
                            print(f"查询规则 {key} 失败: {e}")

    # 2.4 获取 SQL_RULES 中定义的天气数据 (TABLE_SOURCE_WEATHER)
    # 这部分数据需要从 calendar_weather 表中查询，然后拆解 json
    # 查询该日期的天气数据
    weather_row = None
    with db_manager.engine.connect() as conn:
        try:
            sql = text("SELECT * FROM calendar_weather WHERE date = :d")
            weather_row = conn.execute(sql, {"d": target_date_str}).mappings().fetchone()
        except Exception as e:
            print(f"查询天气数据失败: {e}")

    # 无论是否有 weather_row，如果该日期只有天气数据而没有电力数据，我们也希望能入库
    # 所以必须确保遍历到所有可能的来源
    
    if weather_row:
        # 解析 JSON
        weather_json = None
        if weather_row.get('weather_json'):
            try:
                if isinstance(weather_row['weather_json'], str):
                    weather_json = json.loads(weather_row['weather_json'])
                else:
                    weather_json = weather_row['weather_json']
            except:
                pass
        
        # 遍历规则填充数据
        for key, rule in SQL_RULES.items():
            if rule.get('source') == TABLE_SOURCE_WEATHER:
                # 1. 直接映射列
                col_name = rule.get('column')
                json_key = rule.get('json_key')
                
                # 如果有 json_key，则从 JSON 中取值 (通常是数组)
                if json_key and weather_json and json_key in weather_json:
                    values = weather_json[json_key]
                    if isinstance(values, list):
                        # 假设数组长度为 24，对应 0-23 小时
                        # 如果不足 24，则尽力填充
                        for h in range(min(len(values), 24)):
                            val = values[h]
                            if val is not None:
                                try:
                                    hourly_map[h].setdefault(key, []).append(float(val))
                                except (ValueError, TypeError):
                                    hourly_map[h].setdefault(key, []).append(val)
                
                # 2. 如果没有 json_key，则是取列的标量值 (全天相同)
                elif col_name and col_name in weather_row and not json_key:
                    val = weather_row[col_name]
                    # 特殊处理日期字段，将其转换为字符串
                    if isinstance(val, (datetime.date, datetime.datetime)):
                        val = val.strftime("%Y-%m-%d")
                        
                    if val is not None:
                        # 全天 24 小时都用这个值
                        for h in range(24):
                            # 注意：如果是字符串，append 后求均值会报错
                            # 这里需要判断类型
                            if isinstance(val, (int, float)):
                                hourly_map[h].setdefault(key, []).append(float(val))
                            else:
                                hourly_map[h].setdefault(key, []).append(val)
    
    # 即使没有 weather_row，也可能因为有电力数据而继续执行
    # 如果只有天气数据没有电力数据，也会因为 weather_row 存在而有数据
    # 如果两者都没有，下面的 batch_data 为空，返回 0

    # 4. 构造入库数据
    batch_data = []
    
    # 收集所有需要更新的字段
    all_update_fields = set()
    
    if not only_weather:
        all_update_fields.add('price_da')
        all_update_fields.add('price_rt')
        all_update_fields.add('price_diff')
        all_update_fields.add('new_energy_forecast')
        all_update_fields.add('load_deviation')
        for k in field_keys:
            all_update_fields.add(k)
    
    # 添加天气相关字段到更新列表
    for key, rule in SQL_RULES.items():
        if rule.get('source') == TABLE_SOURCE_WEATHER:
            all_update_fields.add(key)

    for h in range(24):
        row_data = {
            "record_date": target_date_str,
            "hour": h
        }
        
        has_data = False
        
        # 处理均值字段
        for k in list(all_update_fields): # 遍历所有可能字段
            if k in ['record_date', 'hour', 'price_diff', 'new_energy_forecast', 'load_deviation']:
                continue
                
            vals = hourly_map[h].get(k, [])
            if vals:
                # 检查是否是数字
                first_val = vals[0]
                # 特殊处理：如果 first_val 是 datetime.date 对象，也转为字符串
                if isinstance(first_val, (datetime.date, datetime.datetime)):
                    first_val = first_val.strftime("%Y-%m-%d")
                    row_data[k] = first_val
                elif isinstance(first_val, (int, float)):
                    avg = sum(vals) / len(vals)
                    row_data[k] = avg
                else:
                    # 非数字，取第一个非空值
                    row_data[k] = first_val
                has_data = True
            else:
                row_data[k] = None
                
            # [新增] 对所有 row_data 的值再次进行类型清洗，确保没有 date 对象
            val = row_data[k]
            if isinstance(val, (datetime.date, datetime.datetime)):
                row_data[k] = val.strftime("%Y-%m-%d")
        
        # 如果整行没有任何数据(连电价都没有)，是否跳过？
        # 如果是增量更新，可能只想更新部分字段。
        # 但如果是 Upsert，None 会覆盖旧值吗？
        # 我们应该只包含有值的字段，或者全部包含。
        # 这里选择：如果没有任何数据，跳过该小时；否则插入/更新所有字段。
        # 修改逻辑：只要有天气数据也算有数据，不能跳过
        if not has_data:
            continue
            
        # 计算衍生字段
        # 1. 价差
        p_da = row_data.get('price_da')
        p_rt = row_data.get('price_rt')
        # 修改逻辑：只要其中一个有值就可以更新，而不是必须两个都有
        # 如果只有一个有值，diff 为 None (因为无法计算价差)，但原有的值应该保留
        if p_da is not None and p_rt is not None:
            row_data['price_diff'] = p_da - p_rt
        else:
            row_data['price_diff'] = None
            
        # 2. 新能源预测总和 (光伏+风电)
        # 假设规则里有 ne_pv_forecast 和 ne_wind_forecast
        pv = row_data.get('ne_pv_forecast', 0) or 0
        wind = row_data.get('ne_wind_forecast', 0) or 0
        if pv or wind:
            row_data['new_energy_forecast'] = pv + wind
        else:
            row_data['new_energy_forecast'] = None

        # 3. 负荷偏差 (预测 - 实际)
        l_fore = row_data.get('load_forecast')
        l_act = row_data.get('load_actual')
        if l_fore is not None and l_act is not None:
            row_data['load_deviation'] = l_fore - l_act
        else:
            row_data['load_deviation'] = None
            
        # [新增] 确保 record_date 和 hour 始终存在 (虽然前面已经定义了)
        row_data['record_date'] = target_date_str
        row_data['hour'] = h
            
        batch_data.append(row_data)
    
    # 5. 入库
    if batch_data:
        # 动态构建 SQL
        # 字段列表: record_date, hour + 其他所有字段
        # 因为 batch_data 里的 keys 可能不完全一致(有些是 None)，最好统一一下
        # 其实 executemany 要求所有字典 keys 一致
        
        # 确保所有字典都有所有字段
        final_keys = list(all_update_fields)
        # 过滤掉不在 batch_data[0] 里的 key (虽然我们在循环里都加了)
        # 为了安全，重新整理 batch_data
        
        # 移除 'record_date' 和 'hour'，因为它们已经单独处理
        if 'record_date' in final_keys:
             final_keys.remove('record_date')
        if 'hour' in final_keys:
             final_keys.remove('hour')
             
        # [DEBUG] 打印一下 final_keys 和 batch_data 的样例，方便调试
        if len(batch_data) > 0:
             print(f"[DEBUG] Cache Update for {target_date_str}: {len(batch_data)} records")
             # print(f"[DEBUG] Keys: {final_keys}")
             # print(f"[DEBUG] Sample Row: {batch_data[0]}")
        else:
             print(f"[DEBUG] Cache Update for {target_date_str}: NO DATA to update.")
             if weather_row:
                 print(f"[DEBUG] Weather Row found but no data mapped? Weather Keys: {weather_row.keys()}")
             else:
                 print(f"[DEBUG] No Weather Row and No Power Data.")
        
        clean_batch = []
        for row in batch_data:
            clean_row = {"record_date": row["record_date"], "hour": row["hour"]}
            for k in final_keys:
                clean_row[k] = row.get(k) # 默认为 None
            clean_batch.append(clean_row)
            
        # 构建 INSERT ... ON DUPLICATE KEY UPDATE 语句
        field_list = [f"`{k}`" for k in final_keys]
        param_list = [f":{k}" for k in final_keys]
        
        # UPDATE 部分
        update_parts = [f"`{k}`=VALUES(`{k}`)" for k in final_keys]
        
        # 注意: 这里的 record_date 和 hour 需要显式加入 VALUES 列表，但不在 UPDATE 列表(主键)
        sql = f"""
            INSERT INTO {table_name} 
            (`record_date`, `hour`, {', '.join(field_list)})
            VALUES (:record_date, :hour, {', '.join(param_list)})
            ON DUPLICATE KEY UPDATE
            {', '.join(update_parts)}
        """
        
        with db_manager.engine.begin() as conn:
             try:
                conn.execute(text(sql), clean_batch)
             except Exception as e:
                 print(f"⚠️ SQL Execution Failed for {target_date_str}: {e}")
                 import traceback
                 traceback.print_exc()
                 raise e # 重新抛出以便上层捕获
            
        return len(clean_batch)
    
    return 0

class PriceDiffCacheRequest(BaseModel):
    dates: list[str] = Field(..., description="日期列表")
    sort_by: Optional[str] = Field("avg_diff", description="排序字段: avg_diff-平均价差, max_diff-最大价差, min_diff-最小价差, total_abs-总绝对值")
    sort_order: Optional[str] = Field("desc", description="排序方向: asc-升序, desc-降序")

@app.post("/api/price-diff-cache", summary="直接从缓存表查询价差数据")
async def query_price_diff_from_cache(request: PriceDiffCacheRequest):
    """
    直接从缓存表 cache_daily_hourly 中查询价差数据

    参数:
    - dates: 日期列表，JSON格式，例如: ["2023-09-18", "2023-09-19"]
    - sort_by: 排序字段 (默认: avg_diff - 平均价差)
    - sort_order: 排序方向 (默认: desc - 降序)

    返回:
    - 缓存表中的价差数据，包含每小时的价差值
    """
    try:
        date_list = request.dates
        sort_by = request.sort_by or "avg_diff"
        sort_order = request.sort_order or "desc"

        table_name = "cache_daily_hourly"

        # 检查表是否存在
        tables = db_manager.get_tables()
        if table_name not in tables:
            return JSONResponse(status_code=404, content={
                "error": "缓存表不存在，请先生成缓存数据",
                "table": table_name
            })

        # 验证日期格式
        valid_dates = []
        for d in date_list:
            try:
                pd.to_datetime(d)
                valid_dates.append(d)
            except:
                pass

        if not valid_dates:
            return JSONResponse(status_code=400, content={"error": "日期格式错误"})

        with db_manager.engine.connect() as conn:
            # 构建查询 SQL
            placeholders = ", ".join([f":d{i}" for i in range(len(valid_dates))])
            params = {f"d{i}": d for i, d in enumerate(valid_dates)}

            # 查询缓存表中的价差数据
            sql = text(f"""
                SELECT record_date, hour, price_diff, price_da, price_rt,
                       load_forecast, temperature, day_type
                FROM {table_name}
                WHERE record_date IN ({placeholders})
                  AND price_diff IS NOT NULL
                ORDER BY record_date, hour
            """)

            result = conn.execute(sql, params).fetchall()

            if not result:
                return {
                    "status": "success",
                    "data": [],
                    "message": "未找到符合条件的价差数据",
                    "dates": valid_dates
                }

            # 转换为列表
            data = []
            for row in result:
                d = dict(row._mapping)
                if hasattr(d.get('record_date'), 'strftime'):
                    d['record_date'] = str(d['record_date'])
                data.append(d)

            # 按日期分组统计
            date_stats = {}
            for row in data:
                date = row['record_date']
                if date not in date_stats:
                    date_stats[date] = {
                        "date": date,
                        "hours": [],
                        "avg_diff": 0,
                        "max_diff": float('-inf'),
                        "min_diff": float('inf'),
                        "total_abs": 0,
                        "positive_count": 0,
                        "negative_count": 0
                    }
                stats = date_stats[date]
                diff = row['price_diff']
                stats["hours"].append({
                    "hour": row["hour"],
                    "price_diff": diff
                })
                if diff is not None:
                    stats["avg_diff"] += diff
                    stats["max_diff"] = max(stats["max_diff"], diff)
                    stats["min_diff"] = min(stats["min_diff"], diff)
                    stats["total_abs"] += abs(diff)
                    if diff > 0:
                        stats["positive_count"] += 1
                    else:
                        stats["negative_count"] += 1

            # 计算平均值
            for date, stats in date_stats.items():
                hour_count = len([h for h in stats["hours"] if h["price_diff"] is not None])
                if hour_count > 0:
                    stats["avg_diff"] = round(stats["avg_diff"] / hour_count, 2)
                    stats["total_abs"] = round(stats["total_abs"], 2)
                stats["hour_count"] = hour_count
                del stats["hours"]  # 移除详细小时数据，只保留统计

            # 转换为列表并排序
            stats_list = list(date_stats.values())

            # 排序
            reverse = sort_order == "desc"
            sort_field = sort_by if sort_by in ["avg_diff", "max_diff", "min_diff", "total_abs"] else "avg_diff"
            stats_list.sort(key=lambda x: x.get(sort_field, 0) or 0, reverse=reverse)

            return {
                "status": "success",
                "data": data,
                "date_stats": stats_list,
                "total_records": len(data),
                "total_dates": len(stats_list),
                "sort_by": sort_by,
                "sort_order": sort_order
            }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/price-diff-cache/export")
async def export_price_diff_from_cache(request: PriceDiffCacheRequest):
    """
    导出缓存表中的价差数据为Excel文件（透视表格式）
    """
    try:
        date_list = request.dates

        table_name = "cache_daily_hourly"
        tables = db_manager.get_tables()

        if table_name not in tables:
            df = pd.DataFrame()
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            output.seek(0)

            from fastapi.responses import StreamingResponse
            import urllib.parse
            filename = f"价差数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            encoded_filename = urllib.parse.quote(filename)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )

        # 验证日期格式
        valid_dates = []
        for d in date_list:
            try:
                pd.to_datetime(d)
                valid_dates.append(d)
            except:
                pass

        with db_manager.engine.connect() as conn:
            placeholders = ", ".join([f":d{i}" for i in range(len(valid_dates))])
            params = {f"d{i}": d for i, d in enumerate(valid_dates)}

            # 构建 FIELD 子句来保持日期顺序（按排名倒序传入的顺序）
            field_order = ", ".join([f":d{i}" for i in range(len(valid_dates))])

            sql = text(f"""
                SELECT record_date, hour, price_diff
                FROM {table_name}
                WHERE record_date IN ({placeholders})
                  AND price_diff IS NOT NULL
                ORDER BY FIELD(record_date, {field_order}), hour
            """)

            result = conn.execute(sql, params).fetchall()

            if not result:
                df = pd.DataFrame()
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                output.seek(0)

                from fastapi.responses import StreamingResponse
                import urllib.parse
                now = datetime.datetime.now()
                filename = f"价差数据_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
                encoded_filename = urllib.parse.quote(filename)
                return StreamingResponse(
                    iter([output.getvalue()]),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
                )

            # 转换为DataFrame
            data = []
            for row in result:
                d = dict(row._mapping)
                if hasattr(d.get('record_date'), 'strftime'):
                    d['record_date'] = str(d['record_date'])
                data.append(d)

            df = pd.DataFrame(data)

            # 透视表格式导出
            if len(df) > 0:
                pivot_df = pd.pivot_table(
                    df,
                    index=['record_date'],
                    columns='hour',
                    values='price_diff',
                    aggfunc='mean'
                )
                pivot_df = pivot_df.reindex(columns=range(24), fill_value=np.nan)
                pivot_df.columns = [f'{int(h):02d}:00' for h in pivot_df.columns]
                pivot_df = pivot_df.reset_index()

                # 按传入的日期顺序（排名倒序）重新排序
                # valid_dates 已经是按排名倒序排列的
                pivot_df['sort_key'] = pivot_df['record_date'].apply(lambda x: valid_dates.index(x) if x in valid_dates else len(valid_dates))
                pivot_df = pivot_df.sort_values('sort_key').drop('sort_key', axis=1)
                pivot_df = pivot_df.reset_index(drop=True)

                # 添加节点名称和单位列在最前面
                pivot_df.insert(0, '节点名称', '价差')
                pivot_df.insert(2, '单位', '价差(元/MWh)')

                # 重命名日期列
                pivot_df = pivot_df.rename(columns={'record_date': '日期'})

                # 重新排列列顺序: 节点名称, 日期, 单位, 00:00...23:00
                cols = ['节点名称', '日期', '单位'] + [f'{h:02d}:00' for h in range(24)]
                for col in cols:
                    if col not in pivot_df.columns:
                        pivot_df[col] = np.nan
                pivot_df = pivot_df[cols]

                final_df = pivot_df
            else:
                columns = ['节点名称', '日期', '单位'] + [f'{h:02d}:00' for h in range(24)]
                final_df = pd.DataFrame(columns=columns)

            # 生成Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name='价差数据')

                # 获取工作表
                worksheet = writer.sheets['价差数据']

                # 应用条件格式：大于0显示绿色渐变，小于0显示红色渐变
                hour_columns = [f'{h:02d}:00' for h in range(24)]

                # 找到所有数值中的最大绝对值，用于确定颜色深度
                max_abs_value = 0
                for col in hour_columns:
                    if col in final_df.columns:
                        max_abs_value = max(max_abs_value, final_df[col].abs().max())

                # 如果最大绝对值为0，则设为1避免除零错误
                if max_abs_value == 0:
                    max_abs_value = 1

                # 定义颜色填充函数
                def get_fill_color(value):
                    if pd.isna(value):
                        return None

                    # 计算颜色强度，基于绝对值比例
                    intensity = abs(value) / max_abs_value

                    # 确保最小亮度，避免颜色过深
                    min_brightness = 150  # 最亮为255
                    brightness_range = 255 - min_brightness
                    brightness = int(min_brightness + (1 - intensity) * brightness_range)

                    if value > 0:
                        # 正数：绿色系，强度越高颜色越深
                        red = brightness
                        green = 255
                        blue = brightness
                    elif value < 0:
                        # 负数：红色系，强度越高颜色越深
                        red = 255
                        green = brightness
                        blue = brightness
                    else:
                        # 零值：白色
                        return None

                    # 转换为十六进制颜色代码
                    color_code = f"{red:02X}{green:02X}{blue:02X}"
                    return PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')

                # 找到小时列的列索引并应用条件格式
                for col_idx, col in enumerate(final_df.columns, start=1):
                    if col in hour_columns:
                        for row_idx in range(2, len(final_df) + 2):  # 从第2行开始（第1行是表头）
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if cell.value is not None:
                                try:
                                    value = float(cell.value)
                                    fill = get_fill_color(value)
                                    if fill:
                                        cell.fill = fill
                                except (ValueError, TypeError):
                                    pass

            output.seek(0)

            from fastapi.responses import StreamingResponse
            import urllib.parse
            now = datetime.datetime.now()
            filename = f"价差数据_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
            encoded_filename = urllib.parse.quote(filename)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})

def normalize_record_time(val, date_str):
    """标准化时间字段，处理 timedelta 和 datetime"""
    try:
        # 1. 已经是 datetime
        if isinstance(val, datetime.datetime):
            return val
            
        # 2. 是 timedelta (Python/Pandas/NumPy)
        # 注意: pd.Timedelta 也是 timedelta 的子类 (在某些版本中)，或者行为类似
        # 分开检查更稳妥
        is_delta = isinstance(val, (datetime.timedelta, pd.Timedelta, np.timedelta64))
        
        if is_delta:
            base_date = pd.to_datetime(date_str)
            return base_date + val
            
        # 3. 尝试 pd.to_datetime (针对字符串或 timestamp)
        # 如果 val 是 timedelta 类型的字符串 (如 "00:15:00")，pd.to_datetime 可能会报错或行为不符合预期
        # 所以先尝试转 timedelta
        try:
            base_date = pd.to_datetime(date_str)
            delta = pd.to_timedelta(val)
            return base_date + delta
        except:
            pass

        return pd.to_datetime(val)
    except:
        # 4. 最后的尝试
        try:
            base_date = pd.to_datetime(date_str)
            # 假设 val 是某种可以转为 timedelta 的东西
            delta = pd.to_timedelta(val)
            return base_date + delta
        except:
            # 打印错误以便调试，但在生产环境中可能太吵
            # print(f"Failed to normalize time: {val} type: {type(val)}")
            return None

if __name__ == "__main__":
    uvicorn.run("api:app", host="0.0.0.0", port=8000, reload=True)


@app.post("/api/daily-hourly-cache/export", summary="导出缓存表的日前/实时/价差数据")
async def export_daily_hourly_from_cache(request: Request):
    """
    从缓存表导出日前、实时、价差数据

    导出格式（每个日期占4行，空一行隔开）:
    - 第1行: 日前节点电价 (00:00 - 23:00)
    - 第2行: 实时节点电价 (00:00 - 23:00)
    - 第3行: 价差 (00:00 - 23:00)
    - 第4行: 空行
    """
    try:
        # 从 FormData 获取 dates
        form_data = await request.form()
        dates_str = form_data.get("dates")
        if not dates_str:
            return JSONResponse(status_code=400, content={"error": "缺少 dates 参数"})

        try:
            date_list = json.loads(dates_str)
        except:
            return JSONResponse(status_code=400, content={"error": "dates 格式错误"})

        table_name = "cache_daily_hourly"
        tables = db_manager.get_tables()

        if table_name not in tables:
            df = pd.DataFrame()
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            output.seek(0)

            from fastapi.responses import StreamingResponse
            import urllib.parse
            filename = f"日前实时价差数据_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            encoded_filename = urllib.parse.quote(filename)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )

        # 验证日期格式
        valid_dates = []
        for d in date_list:
            try:
                pd.to_datetime(d)
                valid_dates.append(d)
            except:
                pass

        if not valid_dates:
            return JSONResponse(status_code=400, content={"error": "日期格式错误"})

        with db_manager.engine.connect() as conn:
            placeholders = ", ".join([f":d{i}" for i in range(len(valid_dates))])
            params = {f"d{i}": d for i, d in enumerate(valid_dates)}

            # 查询日前、实时、价差数据
            sql = text(f"""
                SELECT record_date, hour, price_da, price_rt, price_diff
                FROM {table_name}
                WHERE record_date IN ({placeholders})
                ORDER BY record_date, hour
            """)

            result = conn.execute(sql, params).fetchall()

            if not result:
                df = pd.DataFrame()
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                output.seek(0)

                from fastapi.responses import StreamingResponse
                import urllib.parse
                filename = f"日前实时价差数据_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                encoded_filename = urllib.parse.quote(filename)
                return StreamingResponse(
                    iter([output.getvalue()]),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
                )

            # 转换为DataFrame
            data = []
            for row in result:
                d = dict(row._mapping)
                if hasattr(d.get('record_date'), 'strftime'):
                    d['record_date'] = str(d['record_date'])
                data.append(d)

            df = pd.DataFrame(data)

            # 构建导出格式：每个日期4行（日前、实时、价差、空行）
            export_rows = []
            hour_cols = [f'{h:02d}:00' for h in range(24)]

            # 按日期分组
            dates = df['record_date'].unique()

            for date in dates:
                day_data = df[df['record_date'] == date]

                # 初始化24小时的日前、实时、价差数据
                price_da_row = [np.nan] * 24
                price_rt_row = [np.nan] * 24
                price_diff_row = [np.nan] * 24

                for _, row in day_data.iterrows():
                    hour = int(row['hour'])
                    if 0 <= hour < 24:
                        price_da_row[hour] = row['price_da']
                        price_rt_row[hour] = row['price_rt']
                        price_diff_row[hour] = row['price_diff']

                # 添加日前行
                export_rows.append([date, '日前节点电价'] + price_da_row)
                # 添加实时行
                export_rows.append(['', '实时节点电价'] + price_rt_row)
                # 添加价差行
                export_rows.append(['', '价差'] + price_diff_row)
                # 添加空行
                export_rows.append([''] * 26)

            # 创建DataFrame
            header_cols = ['日期', '类型'] + hour_cols
            export_df = pd.DataFrame(export_rows, columns=header_cols)

            # 生成Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                export_df.to_excel(writer, index=False, sheet_name='日前实时价差')

                # 获取工作表
                worksheet = writer.sheets['日前实时价差']

                # 设置列宽
                worksheet.column_dimensions['A'].width = 12
                worksheet.column_dimensions['B'].width = 12
                for h in range(24):
                    col_letter = chr(67 + h)  # C=67
                    worksheet.column_dimensions[col_letter].width = 8

            output.seek(0)

            from fastapi.responses import StreamingResponse
            import urllib.parse
            filename = f"日前实时价差数据_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            encoded_filename = urllib.parse.quote(filename)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/daily-averages/export-dr-compare", summary="导出日前-实时-均值对比表（带颜色样式）")
async def export_dr_compare(request: Request):
    """
    导出日前-实时-均值对比表

    表格格式：
    - 行：时刻（0:00 - 23:00）
    - 列：每个日期包含日前、实时两个子列
    - 颜色：实时>日前红色，实时<日前绿色
    - 背景色：重点时段(18:00-20:00)黄色，算数均价行黄色
    """
    try:
        def weekday_cn(d: datetime.date) -> str:
            mapping = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            return mapping[d.weekday()]

        # 从 FormData 获取 dates
        form_data = await request.form()
        dates_str = form_data.get("dates")
        if not dates_str:
            return JSONResponse(status_code=400, content={"error": "缺少 dates 参数"})

        try:
            date_list = json.loads(dates_str)
        except:
            return JSONResponse(status_code=400, content={"error": "dates 格式错误"})

        # 验证日期格式
        valid_dates = []
        for d in date_list:
            try:
                pd.to_datetime(d)
                valid_dates.append(d)
            except:
                pass

        if not valid_dates:
            return JSONResponse(status_code=400, content={"error": "日期格式错误"})

        # 按日期升序，保证导出列顺序稳定
        valid_dates = sorted(valid_dates, key=lambda x: pd.to_datetime(x))

        # 直接从缓存表查询
        table_name = "cache_daily_hourly"
        tables = db_manager.get_tables()

        if table_name not in tables:
            df = pd.DataFrame()
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            output.seek(0)

            from fastapi.responses import StreamingResponse
            import urllib.parse
            filename = f"日前实时均值对比_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            encoded_filename = urllib.parse.quote(filename)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )

        with db_manager.engine.connect() as conn:
            placeholders = ", ".join([f":d{i}" for i in range(len(valid_dates))])
            params = {f"d{i}": d for i, d in enumerate(valid_dates)}

            # 查询日前、实时数据
            sql = text(f"""
                SELECT record_date, hour, price_da, price_rt
                FROM {table_name}
                WHERE record_date IN ({placeholders})
                ORDER BY record_date, hour
            """)

            result = conn.execute(sql, params).fetchall()

            if not result:
                df = pd.DataFrame()
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                output.seek(0)

                from fastapi.responses import StreamingResponse
                import urllib.parse
                filename = f"日前实时均值对比_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                encoded_filename = urllib.parse.quote(filename)
                return StreamingResponse(
                    iter([output.getvalue()]),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
                )

            # 构建映射：data_map[date][hour] = (da, rt)
            data_map = {}
            for record_date, hour, price_da, price_rt in result:
                date_key = str(record_date)
                try:
                    hour_int = int(hour)
                except (TypeError, ValueError):
                    continue
                if hour_int < 0 or hour_int > 23:
                    continue
                if date_key not in data_map:
                    data_map[date_key] = {}
                da_val = None if pd.isna(price_da) else float(price_da)
                rt_val = None if pd.isna(price_rt) else float(price_rt)
                data_map[date_key][hour_int] = (da_val, rt_val)

            # 标题：同月则用“YYYY年M月价格合计”，否则用“价格合计”
            parsed_dates = [pd.to_datetime(d).date() for d in valid_dates]
            year_months = {(d.year, d.month) for d in parsed_dates}
            if len(year_months) == 1:
                only_year, only_month = next(iter(year_months))
                title_text = f"{only_year}年{only_month}月价格合计"
            else:
                title_text = "价格合计"

        # 生成Excel并应用样式
        output = BytesIO()

        # 定义颜色样式（深色字体 + 浅色底）
        red_font = Font(color="9C0006")
        green_font = Font(color="006100")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 算数均价行背景
        header_font = Font(bold=True)
        title_font = Font(bold=True, size=16)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side = Side(style="thin", color="D0D0D0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        wb = Workbook()
        ws = wb.active
        ws.title = title_text[:31]

        total_cols = 1 + len(valid_dates) * 2

        # 第一行：标题（合并单元格）
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = title_font
        title_cell.alignment = center
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

        # 第二、三行：多级表头（日期 + 子列）
        ws.cell(row=2, column=1, value="时刻").font = header_font
        ws.cell(row=2, column=1).alignment = center
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)

        for i, date_str in enumerate(valid_dates):
            date_obj = pd.to_datetime(date_str).date()
            header_text = f"{date_obj.month}月{date_obj.day}日({weekday_cn(date_obj)})"
            start_col = 2 + i * 2

            ws.cell(row=2, column=start_col, value=header_text).font = header_font
            ws.cell(row=2, column=start_col).alignment = center
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 1)

            ws.cell(row=3, column=start_col, value="日前").font = header_font
            ws.cell(row=3, column=start_col).alignment = center
            ws.cell(row=3, column=start_col + 1, value="实时").font = header_font
            ws.cell(row=3, column=start_col + 1).alignment = center

        # 数据：第4行开始，0:00 - 23:00 共 24 行
        data_start_row = 4
        # 逐列统计算数均价（按列独立、24 时刻简单平均）
        col_values: dict[tuple[str, str], list[float]] = {}
        for date_str in valid_dates:
            col_values[(date_str, "da")] = []
            col_values[(date_str, "rt")] = []

        for hour in range(24):
            row_idx = data_start_row + hour
            ws.cell(row=row_idx, column=1, value=f"{hour}:00").alignment = center

            for i, date_str in enumerate(valid_dates):
                start_col = 2 + i * 2
                da, rt = data_map.get(date_str, {}).get(hour, (None, None))

                da_cell = ws.cell(row=row_idx, column=start_col, value=da)
                rt_cell = ws.cell(row=row_idx, column=start_col + 1, value=rt)
                da_cell.number_format = "0.00"
                rt_cell.number_format = "0.00"

                if da is not None:
                    col_values[(date_str, "da")].append(da)
                    da_cell.font = green_font
                    da_cell.fill = green_fill
                if rt is not None:
                    col_values[(date_str, "rt")].append(rt)
                    rt_cell.font = green_font
                    rt_cell.fill = green_fill

                # 颜色语义：对同一时刻的一组“日前/实时”进行对比
                # - 较大：红色字体 + 红色底
                # - 较小：绿色字体 + 绿色底
                # - 无对比（相等）：默认绿色
                if da is not None and rt is not None:
                    if da > rt:
                        da_cell.font = red_font
                        da_cell.fill = red_fill
                    elif rt > da:
                        rt_cell.font = red_font
                        rt_cell.fill = red_fill

        # 最后一行：算数均价（整行黄色背景）
        avg_row_idx = data_start_row + 24
        ws.cell(row=avg_row_idx, column=1, value="算数均价").alignment = center
        for i, date_str in enumerate(valid_dates):
            start_col = 2 + i * 2
            da_vals = col_values[(date_str, "da")]
            rt_vals = col_values[(date_str, "rt")]
            da_mean = float(np.mean(da_vals)) if da_vals else None
            rt_mean = float(np.mean(rt_vals)) if rt_vals else None

            da_cell = ws.cell(row=avg_row_idx, column=start_col, value=da_mean)
            rt_cell = ws.cell(row=avg_row_idx, column=start_col + 1, value=rt_mean)
            da_cell.number_format = "0.00"
            rt_cell.number_format = "0.00"

            # 算数均价行：背景固定黄色，字体仍按对比红/绿（无对比默认绿）
            if da_mean is not None:
                da_cell.font = green_font
            if rt_mean is not None:
                rt_cell.font = green_font
            if da_mean is not None and rt_mean is not None:
                if da_mean > rt_mean:
                    da_cell.font = red_font
                elif rt_mean > da_mean:
                    rt_cell.font = red_font

        for col in range(1, total_cols + 1):
            ws.cell(row=avg_row_idx, column=col).fill = yellow_fill

        # 列宽/边框/对齐
        ws.column_dimensions["A"].width = 10
        for col in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

        for row in range(2, avg_row_idx + 1):
            for col in range(1, total_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center

        # 冻结窗格：保留标题/表头与时刻列
        ws.freeze_panes = ws["B4"]

        wb.save(output)

        output.seek(0)

        from fastapi.responses import StreamingResponse
        import urllib.parse
        filename = f"{title_text}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        encoded_filename = urllib.parse.quote(filename)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/daily-averages/export-new-energy-forecast", summary="导出新能源D日预测（从缓存表查询）")
async def export_new_energy_forecast(request: Request):
    """
    导出新能源D日预测（缓存表 cache_daily_hourly）

    - 输入：dates (JSON list)
    - 输出：Excel，sheet: 现货新能源D日预测
    """
    try:
        form_data = await request.form()
        dates_str = form_data.get("dates")
        if not dates_str:
            return JSONResponse(status_code=400, content={"error": "缺少 dates 参数"})

        try:
            date_list = json.loads(dates_str)
        except Exception:
            return JSONResponse(status_code=400, content={"error": "dates 格式错误"})

        # 验证日期格式并排序
        valid_dates = []
        for d in date_list:
            try:
                pd.to_datetime(d)
                valid_dates.append(d)
            except Exception:
                pass

        if not valid_dates:
            return JSONResponse(status_code=400, content={"error": "日期格式错误"})

        valid_dates = sorted(valid_dates, key=lambda x: pd.to_datetime(x))

        table_name = "cache_daily_hourly"
        if table_name not in db_manager.get_tables():
            return JSONResponse(status_code=404, content={"error": "缓存表不存在，请先生成缓存数据", "table": table_name})

        with db_manager.engine.connect() as conn:
            placeholders = ", ".join([f":d{i}" for i in range(len(valid_dates))])
            params = {f"d{i}": d for i, d in enumerate(valid_dates)}
            sql = text(f"""
                SELECT record_date, hour, spot_ne_d_forecast
                FROM {table_name}
                WHERE record_date IN ({placeholders})
                ORDER BY record_date, hour
            """)
            rows = conn.execute(sql, params).fetchall()

        if not rows:
            df_empty = pd.DataFrame()
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_empty.to_excel(writer, index=False, sheet_name="现货新能源D日预测")
            output.seek(0)

            filename = f"新能源D日预测_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            import urllib.parse
            encoded_filename = urllib.parse.quote(filename)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )

        df = pd.DataFrame([dict(r._mapping) for r in rows])
        # 统一 hour 为 int，方便 pivot；无效 hour 行直接丢弃
        df["hour"] = pd.to_numeric(df["hour"], errors="coerce")
        df = df.dropna(subset=["hour"])
        df["hour"] = df["hour"].astype(int)
        df = df[(df["hour"] >= 0) & (df["hour"] <= 23)]

        def build_pivot(value_col: str) -> pd.DataFrame:
            p = df.pivot_table(index="hour", columns="record_date", values=value_col, aggfunc="first")
            p = p.reindex(range(24))  # 确保 0-23 全量行
            p.insert(0, "时刻", [f"{h:02d}:00" for h in range(24)])
            p = p.reset_index(drop=True)
            # 列名统一转为字符串日期，避免 Excel 显示成 Timestamp
            p.columns = [str(c) for c in p.columns]
            return p

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            build_pivot("spot_ne_d_forecast").to_excel(writer, index=False, sheet_name="现货新能源D日预测")
        output.seek(0)

        filename = f"新能源D日预测_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})
