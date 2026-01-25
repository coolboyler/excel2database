#!/usr/bin/env python3
"""
Import Tianlang January spreadsheets into strategy-related tables used by:
  - 报价申报 (/strategy_quote)
  - 策略复盘 (/strategy_review)

Inputs (as provided by user):
  - 天朗1月.xlsx: contains per-day monthly forecast hourly, strategy coefficients, and (optional) declared hourly
  - 天朗一月实际分时.xlsx: contains per-day actual hourly

Key requirement:
  January monthly forecast should be stored independently and must not be affected by the
  D-7/14/21 weighted forecast (because user portfolio changed in January).
  We store these per-day forecasts into strategy_forecast_hourly so the app can prefer it.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date as Date, datetime, time as Time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

import sys
from pathlib import Path as _Path

# Ensure repo root is importable when running as a script from ./scripts.
REPO_ROOT = str(_Path(__file__).resolve().parents[1])
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

import api


@dataclass
class DayBundle:
    forecast: Optional[List[float]] = None  # 24
    coeff: Optional[List[float]] = None  # 24
    declared: Optional[List[float]] = None  # 24


def _as_date(v) -> Optional[Date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, Date):
        return v
    return None


def _as_float(v) -> Optional[float]:
    if v is None:
        return None
    # Excel error strings
    if isinstance(v, str):
        s = v.strip()
        if not s or s.upper() in ("#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!"):
            return None
        try:
            return float(s)
        except Exception:
            return None
    try:
        return float(v)
    except Exception:
        return None


def _find_hour_columns(ws) -> Tuple[int, int]:
    """
    Find the start column for 00:00 and return (start_col, end_col) inclusive for 24 hours.
    Tianlang sheet uses row 1 with a datetime.time(0,0) at the 00:00 column.
    """
    header_row = 1
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, Time) and v.hour == 0 and v.minute == 0:
            return c, c + 23
        if isinstance(v, str) and v.strip() in ("00:00", "0:00", "00:00:00"):
            return c, c + 23
    raise ValueError("Cannot find 00:00 hour column in header row")


def parse_tianlang_forecast_coeff_declared(path: Path) -> Dict[Date, DayBundle]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        start_col, end_col = _find_hour_columns(ws)
        out: Dict[Date, DayBundle] = {}
        current_date: Optional[Date] = None
        for r in range(2, ws.max_row + 1):
            d = _as_date(ws.cell(r, 1).value)
            if d is not None:
                current_date = d
            d = current_date
            if d is None:
                continue  # haven't seen a date yet
            label = ws.cell(r, 3).value
            if not isinstance(label, str):
                continue
            label = label.strip()
            if label not in ("月度预测电量", "策略系数", "日前申报电量"):
                continue

            vals = [_as_float(ws.cell(r, c).value) for c in range(start_col, end_col + 1)]
            if len(vals) != 24:
                continue

            b = out.get(d) or DayBundle()
            if label == "月度预测电量":
                # Require 24 numeric points; otherwise skip this day.
                if all(v is not None for v in vals):
                    b.forecast = [float(v) for v in vals]  # type: ignore[arg-type]
            elif label == "策略系数":
                # Coefficients may contain blanks/errors for some hours; keep None for missing.
                b.coeff = [None if v is None else float(v) for v in vals]
            elif label == "日前申报电量":
                # Declared may be missing for some dates; only keep if complete.
                if all(v is not None for v in vals):
                    b.declared = [float(v) for v in vals]  # type: ignore[arg-type]
            out[d] = b
        return out
    finally:
        wb.close()


def parse_tianlang_actual_hourly(path: Path) -> Dict[Date, List[float]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        out: Dict[Date, List[float]] = {}
        for r in range(1, ws.max_row + 1):
            d = _as_date(ws.cell(r, 1).value)
            if d is None:
                continue
            # Layout observed:
            # col1=date, col2=some daily value (not reliably equal to sum),
            # col3..col26 = 24 hourly values (00..23)
            vals = [_as_float(ws.cell(r, c).value) for c in range(3, 27)]
            if len(vals) != 24:
                continue
            if any(v is None for v in vals):
                continue
            out[d] = [float(v) for v in vals]  # type: ignore[arg-type]
        return out
    finally:
        wb.close()


def main() -> int:
    forecast_path = Path("/Users/cayron/work/excel2sql/天朗1月.xlsx")
    actual_path = Path("/Users/cayron/work/excel2sql/天朗一月实际分时.xlsx")
    if not forecast_path.exists():
        raise SystemExit(f"Missing file: {forecast_path}")
    if not actual_path.exists():
        raise SystemExit(f"Missing file: {actual_path}")

    api._ensure_strategy_tables()

    bundles = parse_tianlang_forecast_coeff_declared(forecast_path)
    actual_map = parse_tianlang_actual_hourly(actual_path)

    # Upsert forecast hourly (explicit monthly forecast)
    forecast_rows = []
    coeff_rows = []
    declared_rows = []
    actual_rows = []
    settle_rows = []

    for d, b in sorted(bundles.items(), key=lambda kv: kv[0]):
        if b.forecast:
            for h, v in enumerate(b.forecast):
                forecast_rows.append(
                    {"record_date": d, "hour": int(h), "forecast_energy": float(v), "source": "tianlang_jan_excel"}
                )
        if b.coeff:
            for h, v in enumerate(b.coeff):
                coeff_rows.append({"record_date": d, "hour": int(h), "coeff": None if v is None else float(v)})
        if b.declared:
            for h, v in enumerate(b.declared):
                declared_rows.append({"record_date": d, "hour": int(h), "declared_energy": float(v)})

    for d, hourly in sorted(actual_map.items(), key=lambda kv: kv[0]):
        for h, v in enumerate(hourly):
            row = {"record_date": d, "hour": int(h), "actual_energy": float(v)}
            actual_rows.append(row)
            settle_rows.append(row.copy())

    # Bulk upserts (single transaction per table), then refresh daily metrics once.
    with api.db_manager.engine.connect() as conn:
        with conn.begin():
            if forecast_rows:
                conn.execute(
                    api.text(
                        """
                        INSERT INTO strategy_forecast_hourly (record_date, hour, forecast_energy, source)
                        VALUES (:record_date, :hour, :forecast_energy, :source)
                        ON DUPLICATE KEY UPDATE
                            forecast_energy=VALUES(forecast_energy),
                            source=COALESCE(VALUES(source), source)
                        """
                    ),
                    forecast_rows,
                )
            if coeff_rows:
                conn.execute(
                    api.text(
                        """
                        INSERT INTO strategy_hourly_coeff (record_date, hour, coeff)
                        VALUES (:record_date, :hour, :coeff)
                        ON DUPLICATE KEY UPDATE coeff=VALUES(coeff)
                        """
                    ),
                    coeff_rows,
                )
            if declared_rows:
                conn.execute(
                    api.text(
                        """
                        INSERT INTO strategy_declared_hourly (record_date, hour, declared_energy)
                        VALUES (:record_date, :hour, :declared_energy)
                        ON DUPLICATE KEY UPDATE declared_energy=VALUES(declared_energy)
                        """
                    ),
                    declared_rows,
                )
            if actual_rows:
                conn.execute(
                    api.text(
                        """
                        INSERT INTO strategy_actual_hourly (record_date, hour, actual_energy)
                        VALUES (:record_date, :hour, :actual_energy)
                        ON DUPLICATE KEY UPDATE actual_energy=VALUES(actual_energy)
                        """
                    ),
                    actual_rows,
                )
            if settle_rows:
                conn.execute(
                    api.text(
                        """
                        INSERT INTO strategy_settlement_actual_hourly (record_date, hour, actual_energy)
                        VALUES (:record_date, :hour, :actual_energy)
                        ON DUPLICATE KEY UPDATE actual_energy=VALUES(actual_energy)
                        """
                    ),
                    settle_rows,
                )

    touched_dates = sorted(set(list(bundles.keys()) + list(actual_map.keys())))
    if touched_dates:
        api._refresh_daily_metrics_for_dates(touched_dates)

    print("Imported:")
    print(f"- forecast days: {sum(1 for b in bundles.values() if b.forecast)}  rows: {len(forecast_rows)}")
    print(f"- coeff days:    {sum(1 for b in bundles.values() if b.coeff)}  rows: {len(coeff_rows)}")
    print(f"- declared days: {sum(1 for b in bundles.values() if b.declared)}  rows: {len(declared_rows)}")
    print(f"- actual days:   {len(actual_map)}  rows: {len(actual_rows)}")
    print(f"- refreshed daily_metrics for {len(touched_dates)} days")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
